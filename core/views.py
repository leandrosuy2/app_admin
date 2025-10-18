from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, permission_required
from django.contrib.auth import authenticate, login  
from .models import Devedor, Empresa, Titulo, Acordo, Parcelamento, UserAccessLog, MensagemWhatsapp, TabelaRemuneracao, TabelaRemuneracaoLista, EmailEnvio, EmailTemplate, Boleto
from django.apps import AppConfig
from django.core.paginator import Paginator
import logging
import base64 
import binascii
from datetime import date, datetime
from django.urls import reverse
from django.contrib import messages
from .models import Empresa, Parcelamento, FollowUp
from django.db import connection
from django.http import HttpResponseNotFound, JsonResponse, HttpResponse, FileResponse
from dateutil.relativedelta import relativedelta
from django.utils.dateformat import format
from django.utils import translation
from django.utils import timezone
from django.db.models import F, Q, Sum 
from django.views.decorators.http import require_POST
from core.models import Acordo, TipoDocTitulo, Agendamento, FollowUp, UsersLojistas, Boleto
from django.utils.timezone import make_aware, now
import re
from .utils import consultar_cnpj_via_scraping
import time
from django.contrib.auth.models import User, Group, Permission
from django.contrib.contenttypes.models import ContentType
from django.contrib.auth.hashers import make_password
from django.utils.translation import gettext_lazy as _
from django.utils.translation import activate
from django.views.decorators.csrf import csrf_exempt
import json
from weasyprint import HTML
from django.template.loader import render_to_string
import tempfile
import traceback
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from io import BytesIO
import pandas as pd
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation


from django.contrib.auth.decorators import login_required, permission_required

from num2words import num2words
import os
from django.conf import settings
from .forms import MensagemWhatsappForm
from django.core.exceptions import ValidationError
from decimal import Decimal, InvalidOperation 


import uuid
import bcrypt

import json
import logging
import requests


from django.shortcuts import get_object_or_404, render
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.decorators import login_required
from django.db import connection

from datetime import date
import os, requests, re
from decimal import Decimal



from urllib.parse import urljoin

 


import logging

logger = logging.getLogger(__name__)


def group_required(group_ids, redirect_url=None, message=None):
    def decorator(view_func):
        def _wrapped_view(request, *args, **kwargs):
            if not request.user.is_authenticated:
                return HttpResponseRedirect(reverse('login'))
            if request.user.groups.filter(id=1).exists():
                context = {
                    'message': "ATENÇÃO, esta é a tela administrativa, você tem acesso ao portal do operador.",
                    'redirect_url': 'https://operador.appone.com.br/'
                }
                return render(request, 'redirect_with_message.html', context)
            elif not any(request.user.groups.filter(id=group_id).exists() for group_id in group_ids):
                return render(request, 'errors/access_denied.html', status=403)
            return view_func(request, *args, **kwargs)
        return _wrapped_view
    return decorator


def home_redirect(request):
    if request.user.is_authenticated:
        return redirect('dashboard')  # Redireciona apenas para usuários autenticados
    return redirect('login')  # Redireciona para login se não estiver autenticado

    

class CoreConfig(AppConfig):
    default_auto_field = 'django.db.models.BigAutoField'
    name = 'core'

 
def format_whatsapp_number(phone):
    """Formata o número de telefone para o padrão do WhatsApp (sem caracteres especiais e com prefixo 55)."""
    if not phone:
        return None
    # Remove caracteres não numéricos
    phone = re.sub(r'\D', '', phone)
    # Adiciona o código do Brasil (55) se o número não começar com ele
    if not phone.startswith('55'):
        phone = f'55{phone}'
    return phone





@login_required
@group_required([2])
def dashboard(request):
    hoje = now().date()
    query = request.GET.get('query', '').strip()
    usuarios = User.objects.filter(is_active=True).order_by('username') 

    # Relatórios rápidos
    titulos_pendentes = Titulo.objects.filter(Q(statusBaixa=0) | Q(statusBaixa__isnull=True)).count()
    titulos_quitados = Titulo.objects.filter(statusBaixa=2).count()
    titulos_negociados = Titulo.objects.filter(statusBaixa=3).count()
    total_clientes = Devedor.objects.count()

    # Consulta para "Negociados Hoje"
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT SUM(valor) AS total_negociados_hoje
            FROM titulo
            WHERE statusBaixa = 3 AND created_at LIKE CONCAT(CURDATE(), '%');
        """)
        result = cursor.fetchone()
        negociados_hoje = Decimal(result[0]) if result[0] is not None else Decimal('0.00')
        negociados_hoje = round(negociados_hoje, 2)

    # Consulta para "Quitados Hoje"
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT SUM(COALESCE(valorRecebido, 0)) AS total_quitados_hoje
            FROM titulo
            WHERE data_baixa = CURDATE();
        """)
        result = cursor.fetchone()
        quitados_hoje = Decimal(result[0]) if result[0] is not None else Decimal('0.00')
        quitados_hoje = round(quitados_hoje, 2)

    # Detalhes de "Quitados Hoje"
    quitados_hoje_detalhes = Titulo.objects.raw("""
        SELECT titulo.id, devedores.nome, COALESCE(devedores.cpf, devedores.cnpj) AS cpf_cnpj,
               core_empresa.nome_fantasia, titulo.data_baixa, titulo.valorRecebido
        FROM titulo
        INNER JOIN devedores ON titulo.devedor_id = devedores.id
        INNER JOIN core_empresa ON devedores.empresa_id = core_empresa.id
        WHERE titulo.data_baixa = %s
    """, [hoje])

    quitados_hoje_detalhes_data = [{
        'nome': q.nome,
        'cpf_cnpj': q.cpf_cnpj,
        'nome_fantasia': q.nome_fantasia,
        'data_baixa': q.data_baixa.strftime('%d/%m/%Y') if q.data_baixa else '-',
        'valorRecebido': f'R$ {float(q.valorRecebido):,.2f}' if q.valorRecebido else 'R$ 0.00'
    } for q in quitados_hoje_detalhes]

    # Detalhes de "Negociados Hoje"
    negociados_hoje_detalhes = Titulo.objects.raw("""
        SELECT titulo.id, devedores.nome, COALESCE(devedores.cpf, devedores.cnpj) AS cpf_cnpj,
               core_empresa.nome_fantasia, titulo.created_at, titulo.valor
        FROM titulo
        INNER JOIN devedores ON titulo.devedor_id = devedores.id
        INNER JOIN core_empresa ON devedores.empresa_id = core_empresa.id
        WHERE titulo.statusBaixa = 3 AND DATE(titulo.created_at) = %s
    """, [hoje])

    negociados_hoje_detalhes_data = [{
        'nome': n.nome,
        'cpf_cnpj': n.cpf_cnpj,
        'nome_fantasia': n.nome_fantasia,
        'data_negociacao': n.created_at.strftime('%d/%m/%Y %H:%M') if n.created_at else '-',
        'valor': f'R$ {float(n.valor):,.2f}' if n.valor else 'R$ 0.00'
    } for n in negociados_hoje_detalhes]

    # Contagem de negociados em atraso
    negociados_em_atraso_count = Titulo.objects.filter(
        statusBaixa=3,
        dataVencimento__lt=hoje
    ).count()

    # Parcelamentos pendentes atrasados ou vencendo hoje
    parcelamentos_atrasados = Parcelamento.objects.filter(
        Q(
            Q(status='Pendente') &
            Q(data_vencimento_parcela__lte=hoje) &
            ~Q(acordo__titulo__ultima_acao=hoje)
        )
    ).select_related(
        'acordo', 'acordo__titulo', 'acordo__titulo__devedor', 'acordo__titulo__devedor__empresa'
    ).annotate(
        qtde_prc=F('acordo__qtde_prc')
    )

    # Últimas movimentações
    ultimos_movimentos = Acordo.objects.select_related('devedor', 'titulo').order_by('-id')[:10].values(
        'id',
        'devedor__nome',
        'titulo_id',
        'entrada',
        'data_entrada',
        'contato',
    )

    # Últimos clientes cadastrados
    ultimos_clientes = Devedor.objects.order_by('-id')[:10].values(
        'id',
        'nome',
        'cpf',
        'cnpj',
        'created_at',
        'nome_fantasia',
    )

    # Agendamentos do dia corrente
    agendamentos_hoje = Agendamento.objects.filter(
        Q(data_retorno__date=hoje) & Q(status='Pendente')
    ).select_related('devedor', 'empresa').values(
        'id',
        'devedor__nome',
        'devedor__cpf',
        'devedor__cnpj',
        'empresa__nome_fantasia',
        'telefone',
        'data_retorno',
        'data_abertura',
        'assunto',
        'operador',
        'status'
    )

    # Filtro de busca para pendentes
    # Filtro de busca para pendentes
    search_filter = ""
    params = []
    username = request.user.username  # obtém o username do usuário logado

    if query:
        # Permite buscar sem filtrar pelo operador logado
        search_filter = """
        AND (
            devedores.nome LIKE %s OR
            core_empresa.nome_fantasia LIKE %s OR
            devedores.cpf LIKE %s OR
            devedores.cnpj LIKE %s
        )
        """
        params.extend([f"%{query}%"] * 4)  # Adiciona os filtros de busca à lista de parâmetros
    else:
        # Aplica o filtro para o operador logado apenas quando não há pesquisa
        search_filter = "AND titulo.operador = %s"
        params.append(username)  # Adiciona o operador logado aos parâmetros

    agenda_pendentes_query = f"""
    SELECT    
        titulo.id AS id,
        devedores.nome,    
        core_empresa.nome_fantasia AS nome_fantasia_credor,
        devedores.nome_mae,
        titulo.operador,
        devedores.cpf,
        devedores.cnpj,
        devedores.rg,    
        devedores.telefone1,
        devedores.razao_social
    FROM 
        devedores, titulo, core_empresa
    WHERE
        titulo.devedor_id = devedores.id 
        AND devedores.empresa_id = core_empresa.id 
        AND (titulo.statusBaixa=0 OR titulo.statusBaixa IS NULL)
        AND (titulo.ultima_acao IS NULL OR DATE(titulo.ultima_acao) != CURDATE())
        and core_empresa.status_empresa =1
        {search_filter}
    GROUP BY 
        titulo.id,
        devedores.nome, 
        core_empresa.nome_fantasia, 
        devedores.nome_mae, 
        devedores.cpf, 
        devedores.cnpj, 
        devedores.rg,
        titulo.juros,        
        devedores.telefone1
    ORDER BY 
        titulo.id DESC
    """

    agenda_pendentes = Titulo.objects.raw(agenda_pendentes_query, params)


    # Paginação para Agenda de Pendentes
    paginator_pendentes = Paginator(agenda_pendentes, 10)
    page_number_pendentes = request.GET.get('page')
    agenda_pendentes_paginated = paginator_pendentes.get_page(page_number_pendentes)

    # Filtro de busca para negociados
    negociados_em_atraso_query = f"""
    SELECT
        MIN(titulo.id) AS id,
        core_empresa.id AS empresa_id,
        devedores.nome AS devedor_nome,
        core_empresa.nome_fantasia AS empresa_nome,
        devedores.nome_mae AS nome_mae,
        titulo.devedor_id AS devedor_id,
        MIN(titulo.dataVencimento) AS data_vencimento,
        SUM(titulo.valor) AS valor_total
    FROM 
        titulo
    JOIN 
        devedores ON titulo.devedor_id = devedores.id
    JOIN 
        core_empresa ON devedores.empresa_id = core_empresa.id
    WHERE 
        titulo.statusBaixa = 3 
        AND titulo.dataVencimento < CURRENT_DATE and core_empresa.status_empresa =1
        
    GROUP BY 
        core_empresa.id, 
        devedores.nome, 
        devedores.nome_mae, 
        titulo.devedor_id, 
        core_empresa.nome_fantasia
    """
    negociados_em_atraso = Titulo.objects.raw(negociados_em_atraso_query)

    # Paginação para Negociados em Atraso
    paginator_negociados = Paginator(list(negociados_em_atraso), 10)
    page_number_negociados = request.GET.get('page_negociados')
    negociados_paginated = paginator_negociados.get_page(page_number_negociados)

    # Contexto final
    context = {
        'usuarios': usuarios,
        'titulos_pendentes': titulos_pendentes,
        'titulos_quitados': titulos_quitados,
        'titulos_negociados': titulos_negociados,
        'total_clientes': total_clientes,
        'negociados_em_atraso_count': negociados_em_atraso_count,
        'parcelamentos_atrasados': parcelamentos_atrasados,
        'ultimos_movimentos': ultimos_movimentos,
        'ultimos_clientes': ultimos_clientes,
        'agendamentos_hoje': agendamentos_hoje,
        'agenda_pendentes_paginated': agenda_pendentes_paginated,
        'negociados_paginated': negociados_paginated,
        'query': query,
        'quitados_hoje': quitados_hoje,
        'negociados_hoje': negociados_hoje,
        'quitados_hoje_detalhes': quitados_hoje_detalhes_data,
        'negociados_hoje_detalhes': negociados_hoje_detalhes_data,
    }

    return render(request, 'dashboard.html', context)


@csrf_exempt
@login_required
@group_required([2])
def ranking_operadores(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        data_inicio = data.get('data_inicio', '')
        data_fim = data.get('data_fim', '')
        operador = data.get('operador', '')

        # Consulta ajustada para agrupar por operador e somar os valores
        query = """
            SELECT 
                titulo.operador,
                COUNT(titulo.id) AS total_titulos,
                SUM(titulo.valorRecebido) AS total_valor_recebido
            FROM 
                titulo
            JOIN 
                devedores ON titulo.devedor_id = devedores.id
            JOIN 
                core_empresa ON devedores.empresa_id = core_empresa.id
            WHERE 
                titulo.statusBaixa = 2
        """
        params = []

        if data_inicio:
            query += " AND titulo.data_baixa >= %s"
            params.append(data_inicio)
        if data_fim:
            query += " AND titulo.data_baixa <= %s"
            params.append(data_fim)
        if operador:
            query += " AND titulo.operador = %s"
            params.append(operador)

        # Agrupar por operador e ordenar pelo total_valor_recebido decrescente
        query += """
            GROUP BY 
                titulo.operador
            ORDER BY 
                total_valor_recebido DESC
        """

        with connection.cursor() as cursor:
            cursor.execute(query, params)
            rows = cursor.fetchall()

            # Calcula o total geral
            total_query = """
                SELECT SUM(titulo.valorRecebido)
                FROM 
                    titulo
                JOIN 
                    devedores ON titulo.devedor_id = devedores.id
                JOIN 
                    core_empresa ON devedores.empresa_id = core_empresa.id
                WHERE 
                    titulo.statusBaixa = 2
            """
            total_params = []
            if data_inicio:
                total_query += " AND titulo.data_baixa >= %s"
                total_params.append(data_inicio)
            if data_fim:
                total_query += " AND titulo.data_baixa <= %s"
                total_params.append(data_fim)
            if operador:
                total_query += " AND titulo.operador = %s"
                total_params.append(operador)

            cursor.execute(total_query, total_params)
            total = cursor.fetchone()[0] or 0.0

        # Formatar os resultados com posição no ranking
        resultados = []
        for idx, row in enumerate(rows, start=1):
            resultados.append({
                'posicao': idx,
                'operador': row[0] or 'Sem Operador',
                'total_titulos': row[1],
                'total_valor_recebido': float(row[2]) if row[2] is not None else 0.0,
            })

        return JsonResponse({
            'resultados': resultados,
            'total': float(total)
        }, safe=False)

    return JsonResponse({'error': 'Método não permitido'}, status=405)

@login_required
@group_required([2])
def listar_grupos(request):
    grupos = Group.objects.all()
    return render(request, 'grupos_listar.html', {'grupos': grupos})

# Criar os grupos e permissões (executar uma vez ou em um script separado)
@login_required
@group_required([2])
def criar_grupos():
    # Criar ou obter os grupos
    admin_group, _ = Group.objects.get_or_create(name='Admin')
    lojista_group, _ = Group.objects.get_or_create(name='Lojista')
    operador_group, _ = Group.objects.get_or_create(name='Operador')

    print("Grupos criados ou já existentes:")
    print(f" - {admin_group}")
    print(f" - {lojista_group}")
    print(f" - {operador_group}")

@login_required
@group_required([2])
def editar_grupo(request, grupo_id):
    grupo = Group.objects.get(id=grupo_id)
    todas_permissoes = Permission.objects.all()  # Todas as permissões disponíveis

    # Associa permissões ao grupo (verifica se estão atribuídas)
    permissoes_detalhadas = [
        {
            'id': permissao.id,
            'codename': permissao.codename,
            'traduzido': permissao.name,
            'atribuida': grupo.permissions.filter(id=permissao.id).exists()  # Verifica associação
        }
        for permissao in todas_permissoes
    ]

    if request.method == 'POST':
        for permissao in todas_permissoes:
            input_name = f"permissoes_{permissao.id}"
            atribuir = request.POST.get(input_name) == "sim"

            if atribuir and not grupo.permissions.filter(id=permissao.id).exists():
                grupo.permissions.add(permissao)  # Adiciona permissão
            elif not atribuir and grupo.permissions.filter(id=permissao.id).exists():
                grupo.permissions.remove(permissao)  # Remove permissão

        return redirect('listar_grupos')  # Redireciona após salvar

    return render(request, 'grupos_editar.html', {
        'grupo': grupo,
        'permissoes_detalhadas': permissoes_detalhadas
    })

import json  # Adicione esta linha


@csrf_exempt
def finalizar_titulo(request, titulo_id):
    if request.method == "POST":
        titulo = get_object_or_404(Titulo, id=titulo_id)
        titulo.ultima_acao = now().date()
        titulo.save()
    return JsonResponse({"status": "success", "message": "Título finalizado com sucesso!"})
    return JsonResponse({"status": "error", "message": "Método não permitido."}, status=405)

@csrf_exempt  # Permite AJAX, mas use o CSRF Token corretamente no cabeçalho
def atualizar_permissao(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            group_id = data.get('group_id')
            permission_id = data.get('permission_id')
            action = data.get('action')

            # Certifique-se de que os IDs são válidos
            grupo = Group.objects.get(id=group_id)
            permissao = Permission.objects.get(id=permission_id)

            if action == "sim":
                grupo.permissions.add(permissao)
            elif action == "nao":
                grupo.permissions.remove(permissao)
            else:
                return JsonResponse({"success": False, "error": "Ação inválida."})

            return JsonResponse({"success": True})
        except Exception as e:
            return JsonResponse({"success": False, "error": str(e)})
    return JsonResponse({"success": False, "error": "Método inválido."})





def permission_denied_view(request, exception):
    return render(request, '403.html', status=403)



def salvar_permissoes(request, grupo_id):
    grupo = Group.objects.get(id=grupo_id)
    if request.method == "POST":
        permissoes = Permission.objects.all()
        for permissao in permissoes:
            # Obtém valor do formulário
            valor = request.POST.get(f'permissoes_{permissao.id}', 'nao')
            if valor == "sim":
                grupo.permissions.add(permissao)  # Adiciona permissão
            else:
                grupo.permissions.remove(permissao)  # Remove permissão
    return redirect('listar_grupos')



def listar_permissoes_ptbr():
    permissoes_por_modelo = {}
    permissoes = Permission.objects.select_related('content_type').all()

    for permissao in permissoes:
        modelo = permissao.content_type.model
        app_label = permissao.content_type.app_label
        nome = permissao.name

        descricao_traduzida = traduzir_permissao(nome)
        permissoes_por_modelo.setdefault(f"{app_label} - {modelo}", []).append(descricao_traduzida)

    return permissoes_por_modelo

def traduzir_permissao(permissao):
    # Traduções das permissões padrões do Django
    traducao = {
        'Can add': 'Pode adicionar',
        'Can change': 'Pode editar',
        'Can delete': 'Pode excluir',
        'Can view': 'Pode visualizar',
        # Adicione outras traduções personalizadas aqui se necessário
    }

    # Busca e substitui o padrão "Can <ação> <modelo>"
    for termo_en, termo_pt in traducao.items():
        if termo_en in permissao:
            return permissao.replace(termo_en, termo_pt)
    return permissao  # Retorna o original se não encontrar tradução
    


def listar_permissoes_view(request):
    permissoes = listar_permissoes_ptbr()
    return render(request, 'listar_permissoes.html', {'permissoes': permissoes})
    
    
   
from django.contrib.auth.models import User, Group
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.decorators import login_required

# ...

def adicionar_usuario(request):
    groups = Group.objects.all()
    if request.method == 'POST':
        name = (request.POST.get('name') or '').strip()
        username = (request.POST.get('username') or '').strip()
        email = (request.POST.get('email') or '').strip()
        password = request.POST.get('password') or ''
        password_confirmation = request.POST.get('password_confirmation') or ''
        group_name = request.POST.get('group') or ''

        # Validações simples
        if not name:
            messages.error(request, "Informe o Nome.")
        if password != password_confirmation:
            messages.error(request, "As senhas não conferem.")

        if messages.get_messages(request):
            # Reexibe o formulário com valores preenchidos
            return render(request, 'usuarios_adicionar.html', {
                'groups': groups,
                'name': name,
                'username': username,
                'email': email,
                'group_selected': group_name,
            })

        try:
            user = User.objects.create_user(username=username, email=email, password=password)
            user.first_name = name  # salva o "Nome"
            user.save()

            if group_name:
                group = Group.objects.get(name=group_name)
                user.groups.add(group)

            return redirect('listar_usuarios')
        except Exception as e:
            messages.error(request, f"Erro ao criar usuário: {e}")

            return render(request, 'usuarios_adicionar.html', {
                'groups': groups,
                'name': name,
                'username': username,
                'email': email,
                'group_selected': group_name,
            })

    return render(request, 'usuarios_adicionar.html', {'groups': groups})


@login_required
@group_required([2])
def listar_usuarios(request):
    usuarios = User.objects.all()
    return render(request, 'usuarios_listar.html', {'usuarios': usuarios})


@login_required
@group_required([2])
def editar_usuario(request, user_id):
    user = get_object_or_404(User, id=user_id)
    groups = Group.objects.all()

    if request.method == 'POST':
        name = (request.POST.get('name') or '').strip()
        username = (request.POST.get('username') or '').strip()
        email = (request.POST.get('email') or '').strip()
        password = request.POST.get('password') or ''
        password_confirmation = request.POST.get('password_confirmation') or ''
        group_name = request.POST.get('group') or ''

        if password and password != password_confirmation:
            messages.error(request, "As senhas não conferem.")

        if not name:
            messages.error(request, "Informe o Nome.")

        if messages.get_messages(request):
            # reexibe com dados preenchidos
            return render(request, 'usuarios_editar.html', {
                'user': user,
                'groups': groups,
                'group_selected': group_name or (user.groups.first().name if user.groups.exists() else None),
                'name': name,
                'username': username,
                'email': email,
            })

        try:
            user.first_name = name
            user.username = username
            user.email = email
            if password:
                user.set_password(password)
            user.save()

            # Atualiza grupo
            user.groups.clear()
            if group_name:
                group = Group.objects.get(name=group_name)
                user.groups.add(group)

            return redirect('listar_usuarios')
        except Exception as e:
            messages.error(request, f"Erro ao editar usuário: {e}")
            return render(request, 'usuarios_editar.html', {
                'user': user,
                'groups': groups,
                'group_selected': group_name or (user.groups.first().name if user.groups.exists() else None),
                'name': name,
                'username': username,
                'email': email,
            })

    return render(request, 'usuarios_editar.html', {
        'user': user,
        'groups': groups,
        'group_selected': user.groups.first().name if user.groups.exists() else None,
    })


@login_required
@group_required([2])
def excluir_usuario(request, user_id):
    user = get_object_or_404(User, id=user_id)
    user.delete()
    #messages.success(request, f'Usuário {user.username} excluído com sucesso!')
    return redirect('listar_usuarios')




@login_required
@group_required([2])
def detalhar_parcela(request, parcela_id):
    parcela = get_object_or_404(Parcelamento, id=parcela_id)
    return render(request, 'detalhar_parcela.html', {'parcela': parcela})

# Devedores - Listar



    
def consult_api(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        devedores_ids = data.get('devedores', [])

        access_token = 'nG1dFp8huYPdgkhpWbAvg0TPYMRLx90B41eOLaoq'  # Replace with your actual access token

        for devedor_id in devedores_ids:
            devedor = get_object_or_404(Devedor, id=devedor_id)

            # Check if CPF or CNPJ is available
            cpf = re.sub(r'\D', '', devedor.cpf) if devedor.cpf else None
            cnpj = re.sub(r'\D', '', devedor.cnpj) if devedor.cnpj else None

            # Determine the correct endpoint based on CPF or CNPJ
            if cpf:
                url = f'https://api.lemit.com.br/api/v1/consulta/pessoa/{cpf}'
            elif cnpj:
                url = f'https://api.lemit.com.br/api/v1/consulta/empresa/{cnpj}'
            else:
                return JsonResponse({'success': False, 'message': 'Neither CPF nor CNPJ is available'})

            headers = {
                'Authorization': f'Bearer {access_token}',
                'Content-Type': 'application/json'
            }

            try:
                response = requests.get(url, headers=headers)
                response.raise_for_status()

                api_data = response.json()

                if 'pessoa' in api_data:
                    pessoa_data = api_data['pessoa']
                elif 'empresa' in api_data:
                    pessoa_data = api_data['empresa']
                else:
                    return JsonResponse({'success': False, 'message': 'Invalid API response structure'})

                devedor.nome = pessoa_data.get('nome', devedor.nome)
                devedor.nome_mae = pessoa_data.get('nome_mae', devedor.nome_mae)

                celulares = pessoa_data.get('celulares', [])
                fixos = pessoa_data.get('fixos', [])

                # Update phone fields, filling empty fields first
                phones = [f"{c['ddd']}{c['numero']}" for c in celulares] + [f"{f['ddd']}{f['numero']}" for f in fixos]

                phone_fields = [
                    'telefone1', 'telefone2', 'telefone3', 'telefone4', 'telefone5',
                    'telefone6', 'telefone7', 'telefone8', 'telefone9', 'telefone10'
                ]

                # Fill empty phone fields first
                empty_fields = [field for field in phone_fields if getattr(devedor, field) in [None, '']]
                for phone in phones:
                    if empty_fields:
                        setattr(devedor, empty_fields.pop(0), phone)
                    else:
                        # If no empty fields, replace existing fields one by one
                        for field in phone_fields:
                            setattr(devedor, field, phone)
                            phone_fields.remove(field)
                            break

                devedor.save()

            except requests.exceptions.HTTPError as http_err:
                return JsonResponse({'success': False,
                                     'message': f'HTTP error occurred: {http_err}. Response: {response.content.decode()}'})
            except Exception as err:
                return JsonResponse({'success': False, 'message': f'Other error occurred: {err}'})

        return JsonResponse({'success': True})
    return JsonResponse({'success': False, 'message': 'Invalid request method'})    

@login_required
@group_required([2])
def baixar_modelo_devedor(request):
    # Consulta as core_empresa.e tipos de documentos no banco
    empresas = Empresa.objects.values_list("nome_fantasia", flat=True)
    tipos_doc = TipoDocTitulo.objects.values_list("id", "name")  # Supondo que essa seja a tabela de tipos de documentos

    # Cria um Workbook com duas abas
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Devedores e Títulos"

    # Defina os campos do modelo
    campos = [
        "empresa_nome_fantasia", "tipo_pessoa", "cpf", "cnpj", "nome", "nome_mae", "rg",
        "razao_social", "nome_fantasia", "nome_socio", "telefone", "telefone1", "telefone2",
        "telefone3", "telefone4", "telefone5", "telefone6", "telefone7", "telefone8",
        "telefone9", "telefone10", "observacao", "cep", "endereco", "bairro", "uf", "cidade", "email1",
        # Campos de Título
        "num_titulo", "dataEmissao", "dataVencimento", "valor", "tipo_doc_id"
    ]
    ws1.append(campos)

    # Adiciona a aba com as empresas
    ws2 = wb.create_sheet("Empresas")
    for empresa in empresas:
        ws2.append([empresa])

    # Adiciona uma aba com os tipos de documento
    ws3 = wb.create_sheet("TiposDoc")
    for tipo in tipos_doc:
        ws3.append([f"{tipo[0]} - {tipo[1]}"])  # Exibe o ID e o nome do documento

    # Adiciona uma aba para as opções de tipo_pessoa
    ws4 = wb.create_sheet("Opções")
    ws4.append(["Tipo Pessoa"])
    ws4.append(["Física"])
    ws4.append(["Jurídica"])

    # Cria validação de dados (drop-down) para a coluna `empresa_nome_fantasia`
    empresa_dv = DataValidation(
        type="list",
        formula1=f"'Empresas'!$A$1:$A${len(empresas)}",
        allow_blank=False,
        showErrorMessage=True
    )
    ws1.add_data_validation(empresa_dv)
    for row in range(2, 1002):
        empresa_dv.add(ws1[f"A{row}"])  # Aplica à coluna `empresa_nome_fantasia`

    # Cria validação de dados para `tipo_pessoa`
    tipo_pessoa_dv = DataValidation(
        type="list",
        formula1=f"'Opções'!$A$2:$A$3",
        allow_blank=False,
        showErrorMessage=True
    )
    ws1.add_data_validation(tipo_pessoa_dv)
    for row in range(2, 1002):
        tipo_pessoa_dv.add(ws1[f"B{row}"])  # Aplica à coluna `tipo_pessoa`

    # Cria validação de dados para `tipo_doc_id`
    tipo_doc_dv = DataValidation(
        type="list",
        formula1=f"'TiposDoc'!$A$1:$A${len(tipos_doc)}",
        allow_blank=False,
        showErrorMessage=True
    )
    ws1.add_data_validation(tipo_doc_dv)
    for row in range(2, 1002):
        tipo_doc_dv.add(ws1[f"AG{row}"])  # Aplica à coluna `tipo_doc_id` (coluna AG)

    # Gera o arquivo Excel
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="modelo_importacao_devedor_titulo.xlsx"'
    wb.save(response)
    return response





@login_required
@group_required([2])
def importar_devedor(request):
    if request.method == "POST" and request.FILES.get("arquivo"):
        try:
            # Lê o arquivo enviado
            arquivo = request.FILES["arquivo"]
            df = pd.read_excel(arquivo)

            for _, row in df.iterrows():
                # Busca a empresa pelo nome_fantasia
                empresa = Empresa.objects.filter(nome_fantasia=row["empresa_nome_fantasia"]).first()
                if not empresa:
                    messages.error(request, f"Empresa não encontrada: {row['empresa_nome_fantasia']}")
                    continue

                # Formata as datas para o formato ISO (yyyy-mm-dd)
                data_emissao = None
                data_vencimento = None

                try:
                    if not pd.isna(row.get("dataEmissao")):
                        data_emissao = pd.to_datetime(row["dataEmissao"], dayfirst=True).date()
                except Exception as e:
                    logger.error(f"Erro ao converter dataEmissao: {row.get('dataEmissao')} - {e}")
                    messages.error(request, f"Erro ao converter dataEmissao: {e}")
                    continue

                try:
                    if not pd.isna(row.get("dataVencimento")):
                        data_vencimento = pd.to_datetime(row["dataVencimento"], dayfirst=True).date()
                except Exception as e:
                    logger.error(f"Erro ao converter dataVencimento: {row.get('dataVencimento')} - {e}")
                    messages.error(request, f"Erro ao converter dataVencimento: {e}")
                    continue

                # Cria ou recupera o devedor
                try:
                    devedor, created = Devedor.objects.get_or_create(
                        empresa=empresa,
                        tipo_pessoa=row["tipo_pessoa"],
                        cpf=row.get("cpf", None),
                        cnpj=row.get("cnpj", None),
                        defaults={
                            "nome": row["nome"],
                            "nome_mae": row["nome_mae"],
                            "rg": row["rg"],
                            "razao_social": row["razao_social"],
                            "nome_fantasia": row["nome_fantasia"],
                            "nome_socio": row["nome_socio"],
                            "telefone": row["telefone"],
                            "telefone1": row["telefone1"],
                            "telefone2": row["telefone2"],
                            "telefone3": row["telefone3"],
                            "telefone4": row["telefone4"],
                            "telefone5": row["telefone5"],
                            "telefone6": row["telefone6"],
                            "telefone7": row["telefone7"],
                            "telefone8": row["telefone8"],
                            "telefone9": row["telefone9"],
                            "telefone10": row["telefone10"],
                            "observacao": row["observacao"],
                            "cep": row["cep"],
                            "endereco": row["endereco"],
                            "bairro": row["bairro"],
                            "uf": row["uf"],
                            "cidade": row["cidade"],
                            "email1": row["email1"],
                        },
                    )
                except Exception as e:
                    logger.error(f"Erro ao criar ou recuperar devedor {row.get('nome')}: {e}")
                    messages.error(request, f"Erro ao criar ou recuperar devedor: {e}")
                    continue

                # Criar o título associado ao devedor
                # Criar o título associado ao devedor
                if not pd.isna(row.get("num_titulo")):
                    tipo_doc_id = row.get("tipo_doc_id")
                    if pd.isna(tipo_doc_id) or not tipo_doc_id:
                        logger.error(f"Tipo de documento ausente ou inválido para o título: {row['num_titulo']}")
                        messages.error(request, f"Tipo de documento ausente ou inválido para o título: {row['num_titulo']}")
                        continue

                    try:
                        # Extrai apenas o ID numérico (assumindo formato "2 - Cheque")
                        tipo_doc_id = int(str(tipo_doc_id).split('-')[0].strip())

                        Titulo.objects.create(
                            devedor=devedor,
                            num_titulo=row["num_titulo"],
                            dataEmissao=data_emissao,
                            dataVencimento=data_vencimento,
                            valor=row["valor"],
                            tipo_doc_id=tipo_doc_id,  # Agora é garantido que seja um inteiro
                            acordoComfirmed=0,  # Valor padrão para o campo acordoComfirmed
                        )
                        #logger.info(f"Título criado com sucesso para o devedor: {row['nome']}")
                    except Exception as e:
                        logger.error(f"Erro ao criar título para o devedor {row['nome']} - {e}")
                        messages.error(request, f"Erro ao criar título para o devedor {row['nome']} - {e}")
                        continue


            #messages.success(request, "Importação concluída com sucesso.")
        except Exception as e:
            logger.error(f"Erro geral durante a importação: {e}")
            messages.error(request, f"Erro geral durante a importação: {e}")
    return redirect("listar_devedores")




def agendamentos_cadastrar(request):
    if request.method == 'POST':
        try:
            devedor_id = request.POST.get('devedor_id')
            empresa_id = request.POST.get('empresa_id')
            telefone = request.POST.get('telefone')  # Captura o telefone
            data_abertura = request.POST.get('data_abertura')
            data_retorno = request.POST.get('data_retorno')
            assunto = request.POST.get('assunto')
            operador = request.POST.get('operador')

            # Validação básica
            if not devedor_id or not empresa_id:
                messages.error(request, "Devedor e Empresa são obrigatórios.")
                return redirect('agendamentos_cadastrar')

            # Atualiza o telefone no devedor, se aplicável
            devedor = Devedor.objects.get(id=devedor_id)
            if telefone:
                devedor.telefone = telefone
                devedor.save()

            # Criação do agendamento
            Agendamento.objects.create(
                devedor=devedor,
                empresa_id=empresa_id,
                data_abertura=data_abertura,
                data_retorno=data_retorno,
                assunto=assunto,
                operador=operador
            )

           # messages.success(request, "Agendamento criado com sucesso.")
            return redirect('agendamentos_listar')

        except Exception as e:
            messages.error(request, f"Erro ao criar agendamento: {e}")
    return render(request, 'agendamentos_criar.html', {'devedores': Devedor.objects.all()})



def login_view(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)

        if user is not None:
            if user.groups.filter(id=2).exists():
                login(request, user)
                return redirect('dashboard')
            elif user.groups.filter(id=1).exists():
                login(request, user)
                # Enviamos a URL de redirecionamento e a mensagem para o template
                return render(request, 'redirect.html', {
                    'message': "ATENÇÃO, esta é a tela administrativa, você tem acesso ao portal do operador.",
                    'redirect_url': 'https://operador.appone.com.br/'
                })
            else:
                return render(request, 'login.html', {'error': 'Acesso negado. Você não pertence ao grupo autorizado.'})
        else:
            return render(request, 'login.html', {'error': 'Credenciais inválidas. Verifique seu usuário e senha.'})

    return render(request, 'login.html')


    


# views.py
import json
from django.shortcuts import render, redirect
from django.urls import reverse
from django.core.paginator import Paginator
from django.db import connection
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse, HttpResponseBadRequest
from django.contrib import messages

from .decorators import group_required
from .models import Devedor


@login_required
@group_required([2])
def listar_devedores(request):
    """
    Lista devedores (apenas quem tem pelo menos 1 título), com pesquisa, filtro por status,
    exclusões (massa/uma/por filtro), totais coerentes e paginação.
    """

    def build_where_and_params(q):
        where_parts = ["COALESCE(e.status_empresa, 1) = 1"]
        params = []
        if q:
            like = f"%{q}%"
            where_parts.append(
                """
                (
                    d.nome           COLLATE utf8mb4_0900_ai_ci LIKE %s OR
                    d.cpf            COLLATE utf8mb4_0900_ai_ci LIKE %s OR
                    d.cnpj           COLLATE utf8mb4_0900_ai_ci LIKE %s OR
                    d.telefone       COLLATE utf8mb4_0900_ai_ci LIKE %s OR
                    e.nome_fantasia  COLLATE utf8mb4_0900_ai_ci LIKE %s OR
                    d.nome_fantasia  COLLATE utf8mb4_0900_ai_ci LIKE %s OR
                    d.razao_social   COLLATE utf8mb4_0900_ai_ci LIKE %s
                )
                """
            )
            params += [like] * 7
        return " AND ".join(where_parts), params

    def status_clause(status_filter):
        if status_filter == "Negociado":
            return " AND COALESCE(spd.any_negociado, 0) = 1"
        if status_filter == "Quitado":
            return " AND COALESCE(spd.any_negociado, 0) = 0 AND COALESCE(spd.any_quitado, 0) = 1"
        if status_filter == "Pendente":
            return " AND COALESCE(spd.any_negociado, 0) = 0 AND COALESCE(spd.any_quitado, 0) = 0"
        return ""

    base_cte = """
    WITH status_por_devedor AS (
        SELECT
            t.devedor_id,
            CASE
                WHEN MAX(CASE WHEN (t.statusBaixa=3 OR t.statusBaixaGeral=3) THEN 1 ELSE 0 END)=1 THEN 3
                WHEN MAX(CASE WHEN (t.statusBaixa=2 OR t.statusBaixaGeral=2) THEN 1 ELSE 0 END)=1 THEN 2
                ELSE 0
            END AS status_baixa_num,
            MAX(CASE WHEN (t.statusBaixa=3 OR t.statusBaixaGeral=3) THEN 1 ELSE 0 END) AS any_negociado,
            MAX(CASE WHEN (t.statusBaixa=2 OR t.statusBaixaGeral=2) THEN 1 ELSE 0 END) AS any_quitado,
            COUNT(DISTINCT t.id) AS qtd_titulos,
            MIN(t.id) AS titulo_id_exemplo
        FROM titulo t
        GROUP BY t.devedor_id
    )
    """

    # POST: ações
    if request.method == "POST":
        op = (request.POST.get("op") or "").strip()
        next_url = request.POST.get("next") or reverse("listar_devedores")

        if op == "mass_delete":
            ids = request.POST.getlist("ids")
            if ids:
                Devedor.objects.filter(id__in=ids).delete()
                messages.success(request, f"{len(ids)} devedor(es) excluído(s).")
            return redirect(next_url)

        if op == "delete_one":
            devedor_id = request.GET.get("devedor_id") or request.POST.get("devedor_id")
            if devedor_id:
                Devedor.objects.filter(id=devedor_id).delete()
                messages.success(request, f"Devedor {devedor_id} excluído.")
            return redirect(next_url)

        if op == "delete_filtered":
            confirm = (request.POST.get("confirm") or "").strip().upper()
            if confirm == "EXCLUIR":
                q_post = (request.POST.get("q") or "").strip()
                status_post = (request.POST.get("status") or "").strip()
                where_sql, params = build_where_and_params(q_post)
                st_where = status_clause(status_post)

                sql_ids = base_cte + f"""
                SELECT d.id
                FROM devedores d
                JOIN core_empresa e              ON d.empresa_id = e.id
                LEFT JOIN status_por_devedor spd ON spd.devedor_id = d.id
                WHERE {where_sql}{st_where}
                  AND COALESCE(spd.qtd_titulos, 0) > 0
                """
                with connection.cursor() as cur:
                    cur.execute(sql_ids, params)
                    ids = [r[0] for r in cur.fetchall()]
                if ids:
                    Devedor.objects.filter(id__in=ids).delete()
                    messages.success(request, f"{len(ids)} devedor(es) excluído(s) pelo filtro.")
            else:
                messages.error(request, "Digite EXCLUIR para confirmar.")
            return redirect(next_url)

    # GET: listagem
    q = (request.GET.get("q") or "").strip()
    status_filter = (request.GET.get("status") or "").strip()

    where_sql, params = build_where_and_params(q)
    status_where = status_clause(status_filter)

    list_sql = base_cte + f"""
    SELECT
        COALESCE(spd.status_baixa_num, 0) AS status_baixa_num,
        d.id               AS devedor_id,
        d.nome             AS devedor_nome,
        d.cpf              AS devedor_cpf,
        d.cnpj             AS devedor_cnpj,
        e.nome_fantasia    AS empresa_nome,
        spd.titulo_id_exemplo,
        d.nome_fantasia,
        d.razao_social,
        COALESCE(spd.qtd_titulos, 0) AS qtd_titulos
    FROM devedores d
    JOIN core_empresa e              ON d.empresa_id = e.id
    LEFT JOIN status_por_devedor spd ON spd.devedor_id = d.id
    WHERE {where_sql}{status_where}
      AND COALESCE(spd.qtd_titulos, 0) > 0
    ORDER BY d.id
    """

    with connection.cursor() as cur:
        cur.execute(list_sql, params)
        rows = cur.fetchall()

    devedores = [
        {
            "id": r[1],
            "titulo_id": r[6],  # existe (tem pelo menos 1 título)
            "nome_fantasia": r[7],
            "razao_social": r[8],
            "nome": r[2],
            "cpf": r[3],
            "cnpj": r[4] or "Não informado",
            "empresa": r[5],
            "quantidade_titulos": r[9],
            "status_baixa": {0: "Pendente", 2: "Quitado", 3: "Negociado"}.get(r[0], "Desconhecido"),
        }
        for r in rows
    ]

    totals_sql = base_cte + f"""
    SELECT
        CASE
            WHEN COALESCE(spd.status_baixa_num, 0) = 3 THEN 'Negociado'
            WHEN COALESCE(spd.status_baixa_num, 0) = 2 THEN 'Quitado'
            ELSE 'Pendente'
        END AS status_txt,
        COUNT(*) AS qtd
    FROM devedores d
    JOIN core_empresa e              ON d.empresa_id = e.id
    LEFT JOIN status_por_devedor spd ON spd.devedor_id = d.id
    WHERE {where_sql}
      AND COALESCE(spd.qtd_titulos, 0) > 0
    GROUP BY status_txt
    """
    with connection.cursor() as cur:
        cur.execute(totals_sql, params)
        tot_rows = cur.fetchall()

    totals = {"pendentes": 0, "negociados": 0, "quitados": 0, "total": 0}
    for status_txt, qtd in tot_rows:
        totals["total"] += qtd
        if status_txt == "Pendente":
            totals["pendentes"] = qtd
        elif status_txt == "Negociado":
            totals["negociados"] = qtd
        elif status_txt == "Quitado":
            totals["quitados"] = qtd

    page_obj = Paginator(devedores, 10).get_page(request.GET.get("page"))
    return render(
        request,
        "devedores_listar.html",
        {"page_obj": page_obj, "query": q, "status": status_filter, "totals": totals},
    )


# ---- ALIAS DE COMPATIBILIDADE ----
@login_required
@group_required([2])
def devedores_listar(request):
    """Alias para manter compatibilidade com o nome antigo."""
    return listar_devedores(request)



# core/views.py
from django.views.decorators.http import require_POST
from django.db import transaction, connection
from django.shortcuts import redirect
from django.urls import reverse
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from .decorators import group_required

def _table_exists(cur, table):
    cur.execute("""
        SELECT COUNT(*) FROM information_schema.TABLES
        WHERE TABLE_SCHEMA = DATABASE() AND TABLE_NAME=%s
    """, [table])
    return cur.fetchone()[0] > 0

def _cols(cur, table):
    cur.execute("""
        SELECT COLUMN_NAME FROM information_schema.COLUMNS
        WHERE TABLE_SCHEMA = DATABASE() AND TABLE_NAME=%s
    """, [table])
    return {r[0] for r in cur.fetchall()}

@login_required
@group_required([2])
@require_POST
def refazer_devedor(request, devedor_id: int):
    next_url = request.POST.get("next") or reverse("listar_devedores")

    with transaction.atomic(), connection.cursor() as cur:
        # --- títulos desse devedor ---
        cur.execute("SELECT id FROM titulo WHERE devedor_id=%s", [devedor_id])
        titulo_ids = [r[0] for r in cur.fetchall()]

        # --- reset de TÍTULOS (compatível com snake_case e camelCase) ---
        tcols = _cols(cur, "titulo")
        set_parts = []

        # status
        for c in ("statusBaixa", "status_baixa"): 
            if c in tcols: set_parts.append(f"{c}=0")
        for c in ("statusBaixaGeral", "status_baixa_geral"):
            if c in tcols: set_parts.append(f"{c}=0")

        # datas de baixa
        for c in ("data_baixa", "dataBaixa"):
            if c in tcols: set_parts.append(f"{c}=NULL")

        # forma de pagamento
        for c in ("forma_pag_Id", "forma_pag_id"):
            if c in tcols: set_parts.append(f"{c}=NULL")

        # valores recebidos
        for c in ("valor_recebido", "valorRecebido"):
            if c in tcols: set_parts.append(f"{c}=0")

        # protocolos e anexos
        for c in ("protocolo", "protocolo_gerado", "codigo_protocolo", "comprovante"):
            if c in tcols: set_parts.append(f"{c}=NULL")

        if set_parts:
            sql = f"UPDATE titulo SET {', '.join(set_parts)} WHERE devedor_id=%s"
            cur.execute(sql, [devedor_id])

        # --- apagar registros relacionados, se existirem ---
        # core_acordo
        if _table_exists(cur, "core_acordo"):
            ac_cols = _cols(cur, "core_acordo")
            where = []
            params = []
            if "devedor_id" in ac_cols:
                where.append("devedor_id=%s"); params.append(devedor_id)
            if "titulo_id" in ac_cols and titulo_ids:
                where.append("titulo_id IN (" + ",".join(["%s"]*len(titulo_ids)) + ")")
                params += titulo_ids
            if where:
                cur.execute("DELETE FROM core_acordo WHERE " + " OR ".join(where), params)

        # core_parcelamento
        if _table_exists(cur, "core_parcelamento"):
            pc_cols = _cols(cur, "core_parcelamento")
            where = []
            params = []
            if "devedor_id" in pc_cols:
                where.append("devedor_id=%s"); params.append(devedor_id)
            if "titulo_id" in pc_cols and titulo_ids:
                where.append("titulo_id IN (" + ",".join(["%s"]*len(titulo_ids)) + ")")
                params += titulo_ids
            if where:
                cur.execute("DELETE FROM core_parcelamento WHERE " + " OR ".join(where), params)

        # (opcional) outras tabelas relacionadas: core_entrada, etc. — só adicionar como acima

    messages.success(request, "Devedor revertido para Pendente e registros de negociação/baixa limpos.")
    return redirect(next_url)




# Adicionar Devedor
from decimal import Decimal
from datetime import date
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect, get_object_or_404
from django.db import connection
from .decorators import group_required  # se você usa controle por grupo
from core.models import Empresa, Devedor, Titulo, TipoDocTitulo  # ajuste o caminho dos seus models

@login_required
@group_required([2])
def adicionar_devedor(request):
    """
    Mesma rota: cria o Devedor e já cria o Título na mesma submissão do formulário.
    """
    empresas   = Empresa.objects.all().order_by('razao_social')
    tipos_docs = TipoDocTitulo.objects.all().order_by('name')

    # todos os campos do devedor que você usa no model
    fields = [
        'cpf', 'cnpj', 'nome', 'nome_mae', 'rg', 'razao_social',
        'nome_fantasia', 'nome_socio', 'cpf_socio', 'rg_socio',
        'telefone', 'telefone1', 'telefone2', 'telefone3', 'telefone4', 'telefone5',
        'telefone6', 'telefone7', 'telefone8', 'telefone9', 'telefone10',
        'cep', 'endereco', 'bairro', 'uf', 'cidade', 'email1', 'email2', 'observacao'
    ]

    if request.method == 'POST':
        data = request.POST.copy()
        errors = {}

        # === validações essenciais
        empresa_id  = data.get('empresa_id')
        tipo_pessoa = data.get('tipo_pessoa')
        if not empresa_id:
            errors['empresa_id'] = 'Selecione a empresa.'
        if tipo_pessoa not in ('F', 'J'):
            errors['tipo_pessoa'] = 'Tipo de pessoa inválido.'

        if tipo_pessoa == 'F':
            if not data.get('nome'):
                errors['nome'] = 'Nome é obrigatório.'
            if not data.get('cpf'):
                errors['cpf'] = 'CPF é obrigatório.'
        else:
            if not data.get('razao_social'):
                errors['razao_social'] = 'Razão social é obrigatória.'
            if not data.get('cnpj'):
                errors['cnpj'] = 'CNPJ é obrigatório.'

        # título obrigatório
        obrigatorios_titulo = [
            ('num_titulo',     'Número do Título'),
            ('tipo_doc_id',    'Tipo de Documento'),
            ('data_emissao',   'Data de Emissão'),
            ('data_vencimento','Data de Vencimento'),
            ('valor',          'Valor'),
        ]
        for k, rotulo in obrigatorios_titulo:
            if not data.get(k):
                errors[k] = f'{rotulo} é obrigatório.'

        # se houver erro, re-renderiza com os valores preenchidos
        if errors:
            initial = {k: data.get(k) for k in fields + [
                'empresa_id','tipo_pessoa',
                'num_titulo','tipo_doc_id','data_emissao','data_vencimento','valor'
            ]}
            return render(request, 'devedores_adicionar.html', {
                'empresas': empresas, 'tipos_docs': tipos_docs,
                'errors': errors, 'initial': initial
            })

        # ==== cria devedor
        empresa = get_object_or_404(Empresa, id=empresa_id)
        devedor_data = {field: (data.get(field) or None) for field in fields}
        devedor = Devedor.objects.create(
            empresa=empresa,
            tipo_pessoa=tipo_pessoa,
            **devedor_data
        )

        # ==== cria título
        tipo_doc = get_object_or_404(TipoDocTitulo, id=data['tipo_doc_id'])
        valor_str = data['valor'].replace('.', '').replace(',', '.')
        try:
            valor_decimal = Decimal(valor_str)
        except Exception:
            valor_decimal = Decimal('0.00')

        Titulo.objects.create(
            empresa=empresa,
            devedor=devedor,
            num_titulo=data['num_titulo'],
            valor=valor_decimal,
            dataVencimento=data['data_vencimento'],
            dataEmissao=data['data_emissao'],
            tipo_doc=tipo_doc,
            statusBaixa=data.get('status_baixa', 0),
        )

        # vai para a listagem do devedor, como você já faz
        return redirect('listar_titulos_por_devedor', devedor_id=devedor.id)

    # GET
    return render(request, 'devedores_adicionar.html', {
        'empresas': empresas,
        'tipos_docs': tipos_docs,
        'fields': fields,
        'initial': {}
    })



@login_required
@group_required([2])
def adicionar_titulo_pg_devedor(request, devedor_id):
    devedor = get_object_or_404(Devedor, id=devedor_id)
    empresas = Empresa.objects.all()
    tipos_docs = TipoDocTitulo.objects.all()

    if request.method == 'POST':
        data = request.POST
        tipo_doc = TipoDocTitulo.objects.get(id=data['tipo_doc_id'])

        # Converte o valor para Decimal no formato correto
        valor_formatado = data['valor'].replace('.', '').replace(',', '.')
        valor_decimal = Decimal(valor_formatado)

        Titulo.objects.create(
            empresa=devedor.empresa,  # Usa a empresa associada ao devedor
            devedor=devedor,
            num_titulo=data['num_titulo'],
            valor=valor_decimal,
            dataVencimento=data['data_vencimento'],
            dataEmissao=data['data_emissao'],  # Adiciona a data de emissão
            tipo_doc=tipo_doc,
            statusBaixa=data.get('status_baixa', 0),
        )
        # messages.success(request, 'Título adicionado com sucesso.')
        return redirect('listar_titulos_por_devedor', devedor_id=devedor.id)

    return render(request, 'titulos_adicionar_pg_devedor.html', {
        'devedor': devedor,
        'empresas': empresas,
        'tipos_docs': tipos_docs,
    })



@login_required
@group_required([2])
def editar_devedor(request, id):
    devedor = get_object_or_404(Devedor, id=id)
    empresas = Empresa.objects.all()

    fields = [
        'id', 'tipo_pessoa', 'cpf', 'cnpj', 'nome', 'nome_mae', 'rg', 'razao_social',
        'nome_fantasia', 'nome_socio', 'cpf_socio', 'rg_socio', 'cep', 'endereco',
        'bairro', 'uf', 'cidade', 'email1', 'email2', 'email3', 'observacao',
        'telefone','telefone1', 'telefone2', 'telefone3', 'telefone4', 'telefone5',
        'telefone6', 'telefone7', 'telefone8', 'telefone9', 'telefone10',
        'telefone_valido','telefone1_valido','telefone2_valido','telefone3_valido',
        'telefone4_valido','telefone5_valido','telefone6_valido','telefone7_valido',
        'telefone8_valido', 'telefone9_valido', 'telefone10_valido',
    ]

    valido_fields = {
        'telefone_valido', 'telefone1_valido', 'telefone2_valido',
        'telefone3_valido', 'telefone4_valido', 'telefone5_valido',
        'telefone6_valido', 'telefone7_valido', 'telefone8_valido',
        'telefone9_valido', 'telefone10_valido',
    }

    devedor_data = {field: getattr(devedor, field, '') for field in fields}
    changes = []

    if request.method == 'POST':
        empresa_id = request.POST.get('empresa_id')
        tipo_pessoa = request.POST.get('tipo_pessoa')
        valid_options = {'SIM', 'NÃO', 'NAO VERIFICADO'}

        for field in fields:
            new_value = request.POST.get(field, '').strip()
            old_value = getattr(devedor, field, None)
            if field in valido_fields:
                new_value = new_value.upper() if new_value in valid_options else 'NAO VERIFICADO'

            if new_value != (old_value if old_value is not None else ''):
                setattr(devedor, field, new_value or None)
                changes.append(f"{field.capitalize()} alterado de '{old_value}' para '{new_value or 'vazio'}'.")

        if empresa_id:
            devedor.empresa_id = empresa_id
        devedor.tipo_pessoa = tipo_pessoa
        devedor.save()

        return JsonResponse({
            'success': True,
            'message': 'Devedor atualizado com sucesso!',
            'changes': changes
        }, status=200)

    # Inclui a lista de números no contexto
    return render(request, 'devedores_editar.html', {
        'devedor': devedor,
        'devedor_data': devedor_data,
        'empresas': empresas,
        'numeros_telefones': range(1, 11),  # Lista de números de 1 a 10
    })




# Dados da API
# Dados da API
API_URL = 'https://api.validocadastro.com.br/json/service.aspx'
CHAVE_ACESSO = 'TkYNXlaJrdIv3m5HBl21PK2i/r2WPGMP2rSLDdOY5Gdof+rU9r6aNCgKtR4hepe4' 

def limpar_cpf_cnpj(cpf_cnpj):
    """ Remove caracteres especiais do CPF/CNPJ """
    return ''.join(filter(str.isdigit, cpf_cnpj))

def buscar_dados_api_cliente(request, devedor_id):
    cpf_cnpj = request.GET.get('cpf', '').strip()  # Primeiro tenta pegar o CPF
    if not cpf_cnpj:  # Se CPF estiver vazio, tenta buscar pelo CNPJ
        cpf_cnpj = request.GET.get('cnpj', '').strip()

    if not cpf_cnpj:
        return JsonResponse({'success': False, 'message': 'CPF ou CNPJ não fornecido.'})

    cpf_cnpj = limpar_cpf_cnpj(cpf_cnpj)

    if not cpf_cnpj.isdigit():
        return JsonResponse({'success': False, 'message': 'CPF ou CNPJ inválido. Insira apenas números.'})

    tipo_pessoa = "F" if len(cpf_cnpj) == 11 else "J"
    data = {
        "CodigoProduto": "332",
        "Versao": "20180521",
        "ChaveAcesso": CHAVE_ACESSO,
        "Parametros": {
            "TipoPessoa": tipo_pessoa,
            "CPFCNPJ": cpf_cnpj
        }
    }

    headers = {'Content-Type': 'application/json'}
    response = requests.post(API_URL, json=data, headers=headers)

    if response.status_code != 200:
        return JsonResponse({'success': False, 'message': 'Erro na comunicação com a API.'})

    result = response.json()
    status = result.get('HEADER', {}).get('INFORMACOES_RETORNO', {}).get('STATUS_RETORNO', {}).get('CODIGO')

    if status == '1':
        try:
            with connection.cursor() as cursor:  # 🔹 Usa a conexão do Django
                sql = "INSERT INTO consultas (cpfcnpj, consulta_data) VALUES (%s, %s)"
                consulta_data = response.text
                cursor.execute(sql, [cpf_cnpj, consulta_data])
            
            return JsonResponse({'success': True, 'message': 'Consulta concluída com sucesso. Resultados salvos no banco de dados.'})
        
        except Exception as err:
            return JsonResponse({'success': False, 'message': f'Erro ao conectar ao banco de dados: {str(err)}'})
    
    else:
        erro_descricao = result.get('HEADER', {}).get('INFORMACOES_RETORNO', {}).get('STATUS_RETORNO', {}).get('DESCRICAO', 'Erro desconhecido')
        return JsonResponse({'success': False, 'message': f'Erro na consulta: {erro_descricao}'})




def normalizar_cpf_cnpj(valor):
    """Remove caracteres especiais do CPF/CNPJ para garantir compatibilidade com o banco."""
    return re.sub(r'\D', '', valor)  # Remove tudo que não for número

@csrf_exempt
def salvar_dados_api_cadastro(request):
    if request.method == 'POST':
        cpf_cnpj = request.POST.get('cpfcnpj', '').strip()
        if not cpf_cnpj:
            return JsonResponse({'success': False, 'message': 'CPF ou CNPJ não fornecido.'})
        
        cpf_cnpj_normalizado = normalizar_cpf_cnpj(cpf_cnpj)
        
        try:
            with connection.cursor() as cursor:
                # Busca os dados na tabela `consultas` pelo CPF/CNPJ
                cursor.execute("SELECT consulta_data FROM consultas WHERE cpfcnpj = %s", [cpf_cnpj_normalizado])
                resultado = cursor.fetchone()

                if not resultado:
                    return JsonResponse({'success': False, 'message': 'Nenhum dado encontrado para esse CPF/CNPJ.'})

                data = json.loads(resultado[0])
                cred_cadastral = data.get("CREDCADASTRAL", {})
                dados_receita = cred_cadastral.get("DADOS_RECEITA_FEDERAL", {})

                nome = dados_receita.get("NOME", "").strip()
                nome_mae = dados_receita.get("NOME_MAE", "").strip()

                telefones = []
                for tipo in ["TELEFONE_FIXO", "TELEFONE_CELULAR"]:
                    if tipo in cred_cadastral:
                        for tel in cred_cadastral[tipo].get("TELEFONES", []):
                            ddd = tel.get("DDD", "").strip()
                            numero = tel.get("NUM_TELEFONE", "").strip()
                            if ddd and numero:
                                telefones.append(f"({ddd}) {numero}")

                while len(telefones) < 10:
                    telefones.append(None)

                # Atualiza os dados na tabela `devedores`
                update_query = """
                    UPDATE devedores
                    SET nome_socio = %s,
                        nome_mae = %s,
                        telefone1 = %s,
                        telefone2 = %s,
                        telefone3 = %s,
                        telefone4 = %s,
                        telefone5 = %s,
                        telefone6 = %s,
                        telefone7 = %s,
                        telefone8 = %s,
                        telefone9 = %s,
                        telefone10 = %s
                    WHERE REPLACE(REPLACE(REPLACE(cpf, '.', ''), '-', ''), '/', '') = %s;
                """
                params = (nome, nome_mae, telefones[0], telefones[1], telefones[2], telefones[3], telefones[4],
                          telefones[5], telefones[6], telefones[7], telefones[8], telefones[9], cpf_cnpj_normalizado)

                cursor.execute(update_query, params)
                connection.commit()

                if cursor.rowcount > 0:
                    return JsonResponse({'success': True, 'message': 'Dados atualizados com sucesso!'})
                else:
                    return JsonResponse({'success': False, 'message': 'Nenhuma linha foi atualizada. O CPF/CNPJ pode não existir no banco.'})

        except Exception as e:
            return JsonResponse({'success': False, 'message': f'Erro ao salvar dados: {str(e)}'})

    return JsonResponse({'success': False, 'message': 'Método inválido.'})






@login_required
@group_required([2])
# Excluir Devedor
def excluir_devedor(request, id):
    devedor = get_object_or_404(Devedor, id=id)
    if request.method == 'POST':
        devedor.delete()
       # messages.success(request, 'Devedor excluído com sucesso.')
        return redirect('listar_devedores')
    return render(request, 'devedores_excluir.html', {'devedor': devedor})
    




@login_required
@group_required([2])
def titulos_listar(request):
    # Obtém os parâmetros de busca e filtro
    query = request.GET.get('q', '')
    status_filter = request.GET.get('status', '')

    # Mapeamento de status
    status_map = {
        "Pendente": (0, "NULL"),  # Considera 0 e NULL para pendentes
        "Quitado": 2,
        "Negociado": 3,
    }

    # Construindo a consulta SQL base
    query_sql = """
        SELECT 
            titulo.id, 
            core_empresa.razao_social, 
            titulo.num_titulo, 
            titulo.valor, 
            titulo.dataVencimento AS data_vencimento,
            titulo.data_baixa, 
            titulo.statusBaixa AS status_baixa,
            devedores.nome AS devedor_nome,
            devedores.cpf AS devedor_cpf,
            core_empresa.nome_fantasia,            
            devedores.cnpj
        FROM 
            titulo
        INNER JOIN devedores ON titulo.devedor_id = devedores.id
        INNER JOIN core_empresa ON devedores.empresa_id = core_empresa.id
        WHERE 1=1 and core_empresa.status_empresa =1
    """

    # Adiciona condição de busca, se aplicável
    query_params = []
    if query:
        query_sql += """
            AND (
                titulo.num_titulo LIKE %s OR
                devedores.nome LIKE %s OR
                devedores.cpf LIKE %s OR
                core_empresa.razao_social LIKE %s OR
                core_empresa.nome_fantasia LIKE %s OR
                devedores.cnpj LIKE %s
            )
        """
        query_params = [f'%{query}%'] * 6

    # Adiciona condição de filtro por status, se aplicável
    if status_filter:
        if status_filter == "Pendente":
            query_sql += " AND (titulo.statusBaixa = 0 OR titulo.statusBaixa IS NULL)"
        else:
            query_sql += " AND titulo.statusBaixa = %s"
            query_params.append(status_map[status_filter])

    # Ordena resultados
    query_sql += " ORDER BY titulo.id DESC"

    # Executa a consulta
    with connection.cursor() as cursor:
        cursor.execute(query_sql, query_params)
        rows = cursor.fetchall()

    # Mapeia os resultados
    titulos = [
        {
            'id': row[0],
            'razao_social': row[1],
            'num_titulo': row[2],
            'valor': row[3],
            'data_vencimento': row[4],
            'data_baixa': row[5],
            'status_baixa': row[6],
            'devedor_nome': row[7],
            'devedor_cpf': row[8],  # Adicionando o CPF
            'nome_fantasia': row[9],            
            'cnpj': row[10],
        }
        for row in rows
    ]

    # Configura paginação
    paginator = Paginator(titulos, 30)  # Limita 30 registros por página
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'titulos_listar.html', {
        'page_obj': page_obj, 
        'query': query, 
        'status': status_filter
    })



from datetime import date
from decimal import Decimal
from django.contrib.auth.decorators import login_required
from django.db import connection
from django.shortcuts import render
from .decorators import group_required  # se você usa no admin

@login_required
@group_required([2])
def listar_titulos_por_devedor(request, devedor_id):
    """
    Lista títulos do devedor, com totals e protocolo (se existir).
    Calcula DIAS DE ATRASO no SELECT (quando não baixado).
    O campo 'juros' exibido é o que está gravado no banco.
    """
    # Verifica qual coluna de protocolo existe
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT COLUMN_NAME
            FROM information_schema.COLUMNS
            WHERE TABLE_SCHEMA = DATABASE() AND TABLE_NAME = 'titulo'
        """)
        cols = {r[0] for r in cursor.fetchall()}

    protocolo_expr = None
    if 'protocolo' in cols:
        protocolo_expr = 'titulo.protocolo'
    elif 'codigo_protocolo' in cols:
        protocolo_expr = 'titulo.codigo_protocolo'
    elif 'protocolo_gerado' in cols:
        protocolo_expr = 'titulo.protocolo_gerado'

    extra_protocolo = f", {protocolo_expr} AS protocolo" if protocolo_expr else ""

    # Busca os títulos (DIA DE ATRASO calculado)
    with connection.cursor() as cursor:
        cursor.execute(f"""
            SELECT
                titulo.id                   AS titulo_id,
                core_empresa.razao_social   AS empresa_razao_social,
                titulo.num_titulo           AS numero_titulo,
                titulo.valor                AS valor_titulo,
                titulo.dataVencimento       AS data_vencimento,
                titulo.data_baixa           AS data_baixa,
                titulo.statusBaixa          AS status_baixa,
                devedores.nome              AS nome_devedor,
                titulo.valorRecebido        AS valor_recebido,
                titulo.forma_pag_Id         AS forma_pag_id,
                titulo.idTituloRef          AS id_titulo_ref,
                COALESCE(titulo.juros, 0)   AS juros,

                /* DIAS DE ATRASO: conta somente se não baixado e houver vencimento */
                CASE
                  WHEN (titulo.data_baixa IS NOT NULL OR titulo.statusBaixa = 2)
                    THEN 0
                  WHEN titulo.dataVencimento IS NULL
                    THEN 0
                  ELSE GREATEST(DATEDIFF(CURDATE(), titulo.dataVencimento), 0)
                END AS dias_atraso

                {extra_protocolo}
            FROM titulo
            INNER JOIN devedores     ON titulo.devedor_id    = devedores.id
            INNER JOIN core_empresa  ON devedores.empresa_id = core_empresa.id
            WHERE devedores.id = %s
        """, [devedor_id])

        cols_desc = [c[0] for c in cursor.description]
        rows = [dict(zip(cols_desc, r)) for r in cursor.fetchall()]

    forma_pagamento_map = {
        0: "Pix", 1: "Dinheiro", 2: "Cartão de Débito", 3: "Cartão de Crédito",
        4: "Cheque", 5: "Depósito em Conta", 6: "Pagamento na Loja",
        7: "Boleto Bancário", 8: "Duplicata",
    }

    titulos_principais = []
    titulos_entrada = []
    entrada_ids = set()

    total_quitado = 0.0
    total_negociado = 0.0
    total_pendente = 0.0

    for r in rows:
        valor          = float(r.get('valor_titulo') or 0)
        valor_recebido = float(r.get('valor_recebido') or 0)
        juros          = float(r.get('juros') or 0)
        status_baixa   = r.get('status_baixa') if r.get('status_baixa') is not None else 0

        titulo_dict = {
            'id': r['titulo_id'],
            'razao_social': r['empresa_razao_social'],
            'num_titulo': r['numero_titulo'],
            'valor': valor,
            'data_vencimento': r.get('data_vencimento'),
            'data_baixa': r.get('data_baixa'),
            'status_baixa': status_baixa,
            'devedor_nome': r['nome_devedor'],
            'valor_recebido': valor_recebido,
            'forma_pagamento': forma_pagamento_map.get(r.get('forma_pag_id'), "Não definido"),
            'juros': juros,  # vindo do BD
            'dias_atraso': r.get('dias_atraso', 0),
            'valor_com_juros': valor + juros,
        }
        if 'protocolo' in r:
            titulo_dict['protocolo'] = r.get('protocolo')

        # Totais
        if status_baixa == 2:          # Quitado
            total_quitado += valor_recebido + juros
        elif status_baixa == 3:        # Negociado
            total_negociado += valor + juros
        else:                          # Pendente
            total_pendente += valor + juros

        # Entrada: sem referência e com baixa/negociação
        if r.get('id_titulo_ref') is None and status_baixa > 1:
            titulos_entrada.append(titulo_dict)
            entrada_ids.add(r['titulo_id'])

        titulos_principais.append(titulo_dict)

    total_negociado_em_aberto = sum(
        (t['valor_com_juros'] or t['valor'] or 0)
        for t in titulos_principais
        if t.get('status_baixa') == 3 and not t.get('data_baixa')
    )

    today = date.today()

    return render(request, 'titulos_listar_por_devedor.html', {
        'titulos': titulos_principais,
        'titulos_entrada': titulos_entrada,
        'entrada_ids': entrada_ids,
        'devedor_id': devedor_id,
        'total_quitado': total_quitado,
        'total_negociado': total_negociado,
        'total_pendente': total_pendente,
        'total_negociado_em_aberto': total_negociado_em_aberto,
        'today': today,
    })




def negociacao_devedor(request, devedor_id):
    devedor = get_object_or_404(Devedor, id=devedor_id)

    # Calcular os totais
    total_quitado = Titulo.objects.filter(devedor=devedor, statusBaixa=2).aggregate(Sum('valorRecebido'))['valorRecebido__sum'] or 0
    total_negociado = Titulo.objects.filter(devedor=devedor, statusBaixa=3).aggregate(Sum('valor'))['valor__sum'] or 0
    total_pendente = Titulo.objects.filter(devedor=devedor, statusBaixa__in=[0, None]).aggregate(Sum('valor'))['valor__sum'] or 0

    # Debugging - Printando no console
    print(f"Total Quitado: {total_quitado}")
    print(f"Total Negociado: {total_negociado}")
    print(f"Total Pendente: {total_pendente}")

    # Passando o contexto para o template
    context = {
        'devedor': devedor,
        'total_quitado': total_quitado,
        'total_negociado': total_negociado,
        'total_pendente': total_pendente,
    }
    return render(request, 'negociacao_devedor.html', context)



@login_required
@group_required([2])    
def adicionar_titulo(request):
    if request.method == 'POST':
        data = request.POST
        empresa = Empresa.objects.get(id=data['empresa_id'])
        devedor = Devedor.objects.get(id=data['devedor_id'])

        Titulo.objects.create(
            empresa=empresa,
            devedor=devedor,
            num_titulo=data['num_titulo'],
            valor=data['valor'],
            data_vencimento=data['data_vencimento'],
            statusBaixa=data.get('status_baixa', 0),
        )
        #messages.success(request, 'Título adicionado com sucesso.')
        return redirect('titulos_listar')

    empresas = Empresa.objects.all()
    devedores = Devedor.objects.all()
    return render(request, 'titulos_adicionar.html', {'empresas': empresas, 'devedores': devedores})


@login_required
@group_required([2])
def editar_titulo(request, id):
    titulo = get_object_or_404(Titulo, id=id)

    # Restringir edição para statusBaixa = 0, null ou 3
    """
    Essa restrição era aplicaca para não editar os tiulos quitados, porém foi  solicitado 
    pelo cliente a permissão de alterar os titulos quitados
    if titulo.statusBaixa not in [0, None, 3]:
        messages.error(request, "Este título não pode ser editado.")
        return redirect('detalhes_devedor', titulo_id=titulo.id)  # 🔥 Correção aqui!
    """ 

    if request.method == "POST":
        novo_valor = request.POST.get("valor")
        nova_data_vencimento = request.POST.get("dataVencimento")

        try:
            titulo.valor = float(novo_valor.replace(',', '.'))  # Convertendo vírgula para ponto
            titulo.dataVencimento = datetime.strptime(nova_data_vencimento, "%Y-%m-%d").date()
            titulo.save()
            messages.success(request, "Título atualizado com sucesso!")
        except ValueError:
            messages.error(request, "Erro ao atualizar o título. Verifique os valores inseridos.")

        return redirect('detalhes_devedor', titulo_id=titulo.id)  # 🔥 Correção aqui também!

    return render(request, 'editar_titulo.html', {'titulo': titulo})





    
    
@login_required
@group_required([2])    
def excluir_titulo(request, id):
    titulo = get_object_or_404(Titulo, id=id)
    if request.method == 'POST':
        titulo.delete()
        #messages.success(request, 'Título excluído com sucesso.')
        return redirect('titulos_listar')

    return render(request, 'titulos_excluir.html', {'titulo': titulo})


# views.py
import json
from django.shortcuts import render, get_object_or_404
from django.http import JsonResponse
from django.views.decorators.http import require_POST
from django.views.decorators.csrf import ensure_csrf_cookie
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.db.models import Q

from .decorators import group_required
from .models import Empresa

def to_bool(v):
    if isinstance(v, bool):
        return v
    try:
        return int(v) != 0
    except Exception:
        s = str(v).strip().lower()
        return s in ("1", "true", "t", "yes", "y", "on")

@login_required
@group_required([2])
@ensure_csrf_cookie           # <- garante o cookie de CSRF na página da lista
def listar_empresas(request):
    query = (request.GET.get('q') or '').strip()
    empresas = Empresa.objects.all()
    if query:
        empresas = empresas.filter(
            Q(id__icontains=query) |
            Q(razao_social__icontains=query) |
            Q(nome_fantasia__icontains=query) |
            Q(cnpj__icontains=query)
        )

    page_obj = Paginator(empresas.order_by('id'), 10).get_page(request.GET.get('page'))
    return render(request, 'empresas_listar.html', {'page_obj': page_obj, 'query': query})

@login_required
@group_required([2])
@require_POST
def alterar_status_empresa(request, id):
    empresa = get_object_or_404(Empresa, id=id)

    try:
        payload = json.loads(request.body.decode('utf-8') or '{}')
    except Exception:
        payload = {}

    if "status_empresa" in payload:
        new_status = to_bool(payload.get("status_empresa"))
    else:
        current = to_bool(empresa.status_empresa)
        new_status = not current

    # salva como 1/0 (funciona p/ BooleanField/IntegerField/CharField)
    try:
        empresa.status_empresa = 1 if new_status else 0
    except Exception:
        empresa.status_empresa = '1' if new_status else '0'

    empresa.save(update_fields=["status_empresa"])
    return JsonResponse({"success": True, "status_empresa": bool(new_status)})

from decimal import Decimal, InvalidOperation
from django.shortcuts import render, get_object_or_404
from django.utils import timezone
from core.models import Empresa, TabelaRemuneracaoLista

def _to_decimal_br(value):
    if value in (None, ""): return Decimal("0")
    s = str(value).strip().replace(".", "").replace(",", ".")
    try:
        return Decimal(s)
    except InvalidOperation:
        return Decimal("0")

def gerar_contrato_lojista(request, id):
    empresa = get_object_or_404(Empresa, id=id)

    dtref = getattr(empresa, "created_at", None) or timezone.now()
    empresa.data_formatada = dtref.strftime("%d/%m/%Y")

    # valor adesão normalizado (se quiser usar no texto por extenso você mantém sua função)
    empresa.valor_adesao = _to_decimal_br(empresa.valor_adesao)

    # Busca e decide se usa tabela dinâmica
    qs = list(TabelaRemuneracaoLista.objects.all().order_by("de_dias"))

    def has_percent(x):
        try:
            return x.percentual not in (None, "") and Decimal(str(x.percentual)) != Decimal("0")
        except Exception:
            return False

    use_dynamic_table = any(has_percent(r) for r in qs)

    context = {
        "empresa": empresa,                 # passa o objeto inteiro
        "remuneracoes": qs,                 # linhas (se houver)
        "use_dynamic_table": use_dynamic_table,  # flag p/ template
    }
    return render(request, "contrato_template_lojista.html", context)

@login_required
@group_required([2])
def gerar_ficha_lojista(request, id):
    activate('pt-br')
    empresa = get_object_or_404(Empresa, id=id)

    # data de cadastro (fallback para hoje)
    try:
        data_formatada = timezone.localtime(empresa.created_at).strftime("%d/%m/%Y") if empresa.created_at else timezone.localdate().strftime("%d/%m/%Y")
    except Exception:
        data_formatada = timezone.localdate().strftime("%d/%m/%Y")

    # valor por extenso
    try:
        bruto = (str(empresa.valor_adesao) if empresa.valor_adesao is not None else "0").strip()
        normalizado = bruto.replace('.', '').replace(',', '.')
        valor_adesao_float = float(normalizado)
        valor_extenso = valor_por_extenso(valor_adesao_float)
    except Exception:
        valor_extenso = ""

    context = {
        "empresa": {
            "razao_social": empresa.razao_social,
            "cnpj": empresa.cnpj,
            "nome_fantasia": empresa.nome_fantasia,
            "endereco": empresa.endereco,
            "numero": getattr(empresa, "numero", ""),
            "bairro": getattr(empresa, "bairro", ""),
            "cidade": empresa.cidade,
            "uf": empresa.uf,
            "cep": empresa.cep,
            "telefone": empresa.telefone,
            "email": empresa.email,
            "inscricao_estadual": getattr(empresa, "ie", ""),
            "contratante_nome": empresa.nome_contato,
            "contratante_cpf": empresa.cpf_contato,
            "valor_adesao": empresa.valor_adesao,
            "valor_adesao_extenso": valor_extenso,
            "data_cadastro": data_formatada,

            # >>> PIX <<<
            "nome_favorecido_pix": getattr(empresa, "nome_favorecido_pix", ""),
            "tipo_pix": getattr(empresa, "tipo_pix", ""),
            "banco": getattr(empresa, "banco", ""),  # sua “chave pix” está nesse campo
        },
        "dados_ficha": {
            "valor_mensal": "R$ isento",
            "site": "www.negociarcobrancas.com.br",
            "fidelidade": (
                "Os títulos deverão permanecer no sistema de cobrança por no mínimo (30) Trinta dias, "
                "para solicitação de cancelamento, após a data de inclusão. A exclusão neste período será "
                "considerada como baixa e aplicado o Percentual de Cobrança por Títulos."
            ),
            "observacao": (
                "Os títulos lançados no sistema Negociar cobranças, terão uma carência de (3) três dias "
                "para solicitação de baixa sem custo, não havendo contato por telefone ou SMS."
            ),
        },
        # se você já calcula isso em outro lugar, pode manter
        "cidade_data": request.GET.get("cd", ""),  # opcional: permite sobrepor pela URL
    }

    return render(request, "ficha_template_lojista.html", context)

# core/views.py
from django.shortcuts import render, get_object_or_404
from django.http import HttpResponse
from django.template.loader import render_to_string
from django.utils import timezone
from django.contrib.auth.decorators import login_required
from django.templatetags.static import static
from .models import Empresa
from .utils import valor_por_extenso
from .decorators import group_required  # se você já usa
from weasyprint import HTML, CSS

def _contrato_context(request, empresa):
    # Data formatada
    try:
        data_fmt = timezone.localtime(empresa.created_at).strftime("%d/%m/%Y") if getattr(empresa, "created_at", None) else timezone.localdate().strftime("%d/%m/%Y")
    except Exception:
        data_fmt = timezone.localdate().strftime("%d/%m/%Y")

    # Valor por extenso
    try:
        bruto = (str(empresa.valor_adesao) if empresa.valor_adesao is not None else "0").strip()
        normalizado = bruto.replace('.', '').replace(',', '.')
        valor_adesao_float = float(normalizado)
        valor_ext = valor_por_extenso(valor_adesao_float)
    except Exception:
        valor_ext = ""

    # LOGO ABSOLUTA:
    # 1) se Empresa tiver campo logo (ImageField), usa-o
    logo_url = None
    if hasattr(empresa, "logo") and getattr(empresa.logo, "url", None):
        try:
            logo_url = request.build_absolute_uri(empresa.logo.url)
        except Exception:
            logo_url = None
    # 2) fallback para static/img/logo.png (coloque seu arquivo lá)
    if not logo_url:
        logo_url = request.build_absolute_uri(static("img/logo.png"))
    # 3) fallback final (dashboard)
    if not logo_url:
        logo_url = request.build_absolute_uri(f"{settings.MEDIA_URL}logos/logo_empresa.jpeg")

    return {
        "empresa": {
            "id": empresa.id,
            "razao_social": empresa.razao_social,
            "cnpj": empresa.cnpj,
            "endereco": empresa.endereco,
            "bairro": getattr(empresa, "bairro", ""),
            "cidade": empresa.cidade,
            "uf": empresa.uf,
            "valor_adesao": empresa.valor_adesao,
            "valor_adesao_extenso": valor_ext,
            "data_formatada": data_fmt,
        },
        "logo_url": logo_url,  # << aqui
    }

@login_required
@group_required([2])
def contrato_cobranca(request, id):
    empresa = get_object_or_404(Empresa, id=id)
    context = _contrato_context(request, empresa)
    return render(request, "contrato_prestacao.html", context)

@login_required
@group_required([2])
def contrato_cobranca_pdf(request, id):
    empresa = get_object_or_404(Empresa, id=id)
    context = _contrato_context(request, empresa)
    html = render_to_string("contrato_prestacao.html", context, request=request)

    pdf = HTML(string=html, base_url=request.build_absolute_uri("/")).write_pdf(
        stylesheets=[CSS(string='@page { size: A4; margin: 18mm 14mm; }')]
    )
    resp = HttpResponse(pdf, content_type="application/pdf")
    resp["Content-Disposition"] = f'inline; filename="contrato_{empresa.id}.pdf"'
    return resp





from decimal import Decimal, InvalidOperation
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.shortcuts import get_object_or_404, render, redirect

from core.decorators import group_required
from core.models import Empresa, TabelaRemuneracao

def _only_digits(s: str) -> str:
    return ''.join(ch for ch in (s or '') if ch.isdigit())

def _to_decimal(br_value: str) -> Decimal:
    raw = (br_value or '').strip().replace('.', '').replace(',', '.')
    return Decimal(raw or '0')

@login_required
@group_required([2])
def editar_empresa(request, id):
    empresa = get_object_or_404(Empresa, id=id)
    tabelas = TabelaRemuneracao.objects.all()

    if request.method == 'POST':
        # Básicos
        empresa.razao_social        = request.POST.get('razao_social', empresa.razao_social).strip()
        empresa.nome_fantasia       = request.POST.get('nome_fantasia', empresa.nome_fantasia).strip()
        empresa.cnpj                = _only_digits(request.POST.get('cnpj')) or empresa.cnpj
        empresa.nome_contato        = request.POST.get('nome_contato', empresa.nome_contato).strip()
        empresa.cpf_contato         = _only_digits(request.POST.get('cpf_contato')) or empresa.cpf_contato

        # Contatos
        empresa.telefone            = _only_digits(request.POST.get('telefone')) or ''
        empresa.celular             = _only_digits(request.POST.get('celular')) or ''
        empresa.whatsapp_financeiro = _only_digits(request.POST.get('whatsapp_financeiro')) or ''
        empresa.email               = (request.POST.get('email') or '').strip()
        empresa.email_financeiro    = (request.POST.get('email_financeiro') or '').strip()

        # Endereço
        empresa.cep                 = _only_digits(request.POST.get('cep')) or ''
        empresa.endereco            = (request.POST.get('endereco') or '').strip()
        empresa.numero              = (request.POST.get('numero') or '').strip()
        empresa.bairro              = (request.POST.get('bairro') or '').strip()
        empresa.uf                  = (request.POST.get('uf') or '').strip().upper()[:2]
        empresa.cidade              = (request.POST.get('cidade') or '').strip()

        # Equipe
        empresa.operador            = (request.POST.get('operador') or '').strip()
        empresa.supervisor          = (request.POST.get('supervisor') or '').strip()
        empresa.gerente             = (request.POST.get('gerente') or '').strip()

        # Dados bancários / PIX
        empresa.banco_nome          = (request.POST.get('banco_nome') or '').strip()
        empresa.agencia             = _only_digits(request.POST.get('agencia')) or ''
        empresa.conta               = _only_digits(request.POST.get('conta')) or ''
        empresa.chave_pix           = (request.POST.get('chave_pix') or '').strip()
        empresa.nome_favorecido_pix = (request.POST.get('nome_favorecido_pix') or '').strip()
        empresa.tipo_pix            = (request.POST.get('tipo_pix') or '').strip()

        # Implantação / Negociação
        try:
            empresa.valor_adesao            = _to_decimal(request.POST.get('valor_adesao'))
            empresa.desconto_total_avista   = _to_decimal(request.POST.get('desconto_total_avista'))
            empresa.desconto_total_aprazo   = _to_decimal(request.POST.get('desconto_total_aprazo'))
        except (InvalidOperation, ValueError):
            messages.error(request, 'Algum valor numérico está em formato inválido.')
            return render(request, 'empresas_editar.html', {'empresa': empresa, 'tabelas': tabelas})

        empresa.qtd_parcelas = int(request.POST.get('qtd_parcelas') or 0)

        # Plano
        plano_id = request.POST.get('plano')
        if plano_id:
            try:
                empresa.plano = TabelaRemuneracao.objects.get(id=plano_id)
            except TabelaRemuneracao.DoesNotExist:
                messages.error(request, 'Plano inválido.')
                return render(request, 'empresas_editar.html', {'empresa': empresa, 'tabelas': tabelas})

        # Logo
        novo_logo = request.FILES.get('logo')
        if novo_logo:
            empresa.logo = novo_logo

        # (opcional) status via form
        status_empresa = request.POST.get('status_empresa')
        if status_empresa is not None:
            empresa.status_empresa = (status_empresa == "True")

        try:
            empresa.save()
            messages.success(request, 'Empresa editada com sucesso.')
            return redirect('listar_empresas')
        except Exception as e:
            messages.error(request, f'Erro ao salvar: {e}')

    return render(request, 'empresas_editar.html', {'empresa': empresa, 'tabelas': tabelas})


@login_required
@group_required([2])
def alterar_status_empresa(request, id):
    if request.method == "POST":
        empresa = get_object_or_404(Empresa, id=id)

        try:
            # Alternar o status da empresa
            empresa.status_empresa = not empresa.status_empresa
            empresa.save()
            
            return JsonResponse({"success": True, "status_empresa": empresa.status_empresa})
        except Exception as e:
            return JsonResponse({"success": False, "error": str(e)})

    return JsonResponse({"success": False, "error": "Método inválido"})
    
@login_required
@group_required([2])
def excluir_empresa(request, id):
    empresa = get_object_or_404(Empresa, id=id)
    dependentes = Devedor.objects.filter(empresa=empresa)
    if request.method == 'POST':
        dependentes.delete()  # Exclui dependentes
        empresa.delete()  # Exclui empresa
        messages.success(request, 'Empresa e dependentes excluídos com sucesso.')
        return redirect('listar_empresas')
    return render(request, 'empresas_excluir.html', {
        'empresa': empresa,
        'dependentes': dependentes,
    })


def validar_cnpj(cnpj):
    """
    Valida o formato e a estrutura do CNPJ.
    """
    cnpj = re.sub(r'\D', '', cnpj)  # Remove caracteres não numéricos
    if len(cnpj) != 14:
        return False
    
    # Validação básica para números sequenciais
    if cnpj in (c * 14 for c in "0123456789"):
        return False
    
    # Cálculo dos dígitos verificadores
    def calcular_digito(cnpj, peso):
        soma = sum(int(cnpj[i]) * peso[i] for i in range(len(peso)))
        resto = soma % 11
        return '0' if resto < 2 else str(11 - resto)
    
    primeiro_digito = calcular_digito(cnpj[:12], [5, 4, 3, 2, 9, 8, 7, 6, 5, 4, 3, 2])
    segundo_digito = calcular_digito(cnpj[:13], [6, 5, 4, 3, 2, 9, 8, 7, 6, 5, 4, 3, 2])
    
    return cnpj[12] == primeiro_digito and cnpj[13] == segundo_digito
    
    
import smtplib
from decimal import Decimal, InvalidOperation  # ✅ Importação corrigida
from django.contrib.auth.decorators import login_required
from django.shortcuts import redirect, render
from django.contrib import messages
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from .models import Empresa, TabelaRemuneracao, EmailEnvio, EmailTemplate






COPIA_EMAIL = "nortecheck-to@hptmail.com"  # Sempre enviar cópia para este e-mail

def buscar_config_email(tipo_envio):
    """Busca as configurações de e-mail do banco para um tipo específico de envio."""
    return EmailEnvio.objects.filter(tipo_envio=tipo_envio).first()

def buscar_template_email(tipo_envio):
    """Busca o template de e-mail do banco para um tipo específico de envio."""
    return EmailTemplate.objects.filter(tipo_envio=tipo_envio).first()

def substituir_variaveis(template, dados):
    """Substitui variáveis dinâmicas no template de e-mail."""
    for chave, valor in dados.items():
        template = template.replace(f"{{{{{chave}}}}}", str(valor))
    return template

def enviar_email_tipo_envio(tipo_envio, destinatario, dados):
    """Envia um e-mail baseado no tipo de envio e nas informações fornecidas, sempre copiando um e-mail fixo."""
    print(f"🔍 Buscando configurações e template para '{tipo_envio}'...")

    config_email = buscar_config_email(tipo_envio)
    template_email = buscar_template_email(tipo_envio)

    if not config_email or not template_email:
        print(f"❌ Falha: Configuração ou template não encontrado para '{tipo_envio}'")
        return False

    corpo_email = substituir_variaveis(template_email.mensagem, dados)

    msg = MIMEMultipart()
    msg['From'] = config_email.email
    msg['To'] = destinatario
    msg['Cc'] = COPIA_EMAIL  # ✅ Adicionando a cópia
    msg['Subject'] = f"Notificação: {tipo_envio}"

    msg.attach(MIMEText(corpo_email, 'plain'))

    destinatarios = [destinatario, COPIA_EMAIL]  # Lista com destinatário e cópia

    try:
        print(f"📧 Enviando e-mail para {destinatario} e cópia para {COPIA_EMAIL} via {config_email.servidor_smtp}:{config_email.porta}...")
        
        server = smtplib.SMTP_SSL(config_email.servidor_smtp, config_email.porta)
        server.login(config_email.email, config_email.senha)
        server.sendmail(config_email.email, destinatarios, msg.as_string())
        server.quit()
        
        print(f"✅ E-mail '{tipo_envio}' enviado com sucesso para {destinatario} e {COPIA_EMAIL}!")
        return True
    except Exception as e:
        print(f"❌ Erro ao enviar e-mail '{tipo_envio}': {e}")
        return False


from decimal import Decimal, InvalidOperation
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect

from core.decorators import group_required
from core.models import Empresa, TabelaRemuneracao

def _only_digits(s: str) -> str:
    return ''.join(ch for ch in (s or '') if ch.isdigit())

def _to_decimal(br_value: str) -> Decimal:
    raw = (br_value or '').strip().replace('.', '').replace(',', '.')
    return Decimal(raw or '0')

@login_required
@group_required([2])
def adicionar_empresa(request):
    tabelas = TabelaRemuneracao.objects.all().order_by('nome')

    if request.method == 'POST':
        try:
            # Dados básicos
            nome_fantasia       = (request.POST.get('nome_fantasia') or '').strip()
            razao_social        = (request.POST.get('razao_social') or '').strip()
            nome_contato        = (request.POST.get('nome_contato') or '').strip()
            cpf_contato         = _only_digits(request.POST.get('cpf_contato'))
            cnpj                = _only_digits(request.POST.get('cnpj'))

            telefone            = _only_digits(request.POST.get('telefone'))
            celular             = _only_digits(request.POST.get('celular'))
            whatsapp_financeiro = _only_digits(request.POST.get('whatsapp_financeiro'))

            email               = (request.POST.get('email') or '').strip()
            email_financeiro    = (request.POST.get('email_financeiro') or '').strip()

            cep                 = _only_digits(request.POST.get('cep'))
            endereco            = (request.POST.get('endereco') or '').strip()
            numero              = (request.POST.get('numero') or '').strip()
            bairro              = (request.POST.get('bairro') or '').strip()
            cidade              = (request.POST.get('cidade') or '').strip()
            uf                  = (request.POST.get('uf') or '').strip().upper()[:2]
            ie                  = (request.POST.get('ie') or '').strip()

            operador            = (request.POST.get('operador') or '').strip()
            supervisor          = (request.POST.get('supervisor') or '').strip()
            gerente             = (request.POST.get('gerente') or '').strip()

            # Bancários / PIX
            banco_nome          = (request.POST.get('banco_nome') or '').strip()
            agencia             = _only_digits(request.POST.get('agencia'))
            conta               = _only_digits(request.POST.get('conta'))
            chave_pix           = (request.POST.get('chave_pix') or '').strip()
            nome_favorecido_pix = (request.POST.get('nome_favorecido_pix') or '').strip()
            tipo_pix            = (request.POST.get('tipo_pix') or '').strip()

            # Taxas / negociação
            try:
                valor_adesao = _to_decimal(request.POST.get('valor_adesao'))  # "Implantação"
            except (InvalidOperation, ValueError):
                messages.error(request, 'O valor de Implantação está em formato inválido.')
                return render(request, 'empresas_adicionar.html', {'tabelas': tabelas})

            qtd_parcelas            = int(request.POST.get('qtd_parcelas') or 0)
            desconto_total_avista   = _to_decimal(request.POST.get('desconto_total_avista'))
            desconto_total_aprazo   = _to_decimal(request.POST.get('desconto_total_aprazo'))

            plano_id = request.POST.get('plano')
            logo_file = request.FILES.get('logo')

            # Validação mínima
            if not (nome_fantasia and razao_social and cnpj and plano_id):
                messages.error(request, 'Preencha CNPJ, Razão Social, Nome Fantasia e Plano.')
                return render(request, 'empresas_adicionar.html', {'tabelas': tabelas})

            try:
                plano = TabelaRemuneracao.objects.get(id=plano_id)
            except TabelaRemuneracao.DoesNotExist:
                messages.error(request, 'Plano inválido.')
                return render(request, 'empresas_adicionar.html', {'tabelas': tabelas})

            empresa = Empresa.objects.create(
                nome_fantasia=nome_fantasia,
                razao_social=razao_social,
                cnpj=cnpj,
                nome_contato=nome_contato,
                cpf_contato=cpf_contato,

                telefone=telefone,
                celular=celular,
                whatsapp_financeiro=whatsapp_financeiro,

                email=email,
                email_financeiro=email_financeiro,

                cep=cep,
                endereco=endereco,
                numero=numero,
                bairro=bairro,
                cidade=cidade,
                uf=uf,
                ie=ie,

                operador=operador,
                supervisor=supervisor,
                gerente=gerente,

                # Bancários
                banco_nome=banco_nome,
                agencia=agencia,
                conta=conta,
                chave_pix=chave_pix,
                nome_favorecido_pix=nome_favorecido_pix,
                tipo_pix=tipo_pix,

                # Taxas / negociação
                valor_adesao=valor_adesao,                 # Implantação
                qtd_parcelas=qtd_parcelas,
                desconto_total_avista=desconto_total_avista,
                desconto_total_aprazo=desconto_total_aprazo,

                plano=plano,
                logo=logo_file,
            )

            messages.success(request, 'Empresa adicionada com sucesso.')
            return redirect('listar_empresas')

        except Exception as e:
            messages.error(request, f'Erro ao adicionar empresa: {e}')

    return render(request, 'empresas_adicionar.html', {'tabelas': tabelas})


def consultar_cnpj_view(request):
    cnpj = request.GET.get('cnpj', '').strip()
    if not cnpj:
        return JsonResponse({"erro": "CNPJ não fornecido"}, status=400)

    resultado = consultar_cnpj_via_scraping(cnpj)
    return JsonResponse(resultado)


def consultar_com_espera(cnpj):
    time.sleep(5)  # Espera 5 segundos entre as consultas
    return consultar_cnpj_via_scraping(cnpj)



from datetime import date, datetime
from decimal import Decimal, ROUND_HALF_UP
from dateutil.relativedelta import relativedelta

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect, get_object_or_404

# --- imports do seu projeto (ajuste o caminho se necessário) ---
from .decorators import group_required
from core.models import Titulo  # ajuste se o model estiver noutro app

# --- FORMAS DE PAGAMENTO (opcional: importe de um lugar central) ---
FORMAS_PAGAMENTO_CHOICES = [
    (0, "Pix"),
    (1, "Dinheiro"),
    (2, "Cartão de Débito"),
    (3, "Cartão de Crédito"),
    (4, "Cheque"),
    (5, "Depósito em Conta"),
    (6, "Pagamento na Loja"),
    (7, "Boleto Bancário"),
    (8, "Duplicata"),
    (9, "Recebimento pelo credor"),
]
FORMAS_PAGAMENTO = dict(FORMAS_PAGAMENTO_CHOICES)

# --- Envio de e-mail (se não existir, vira no-op) ---
try:
    from .utils import enviar_email_tipo_envio  # ajuste o caminho se preciso
except Exception:
    def enviar_email_tipo_envio(*args, **kwargs):
        return False


@login_required
@group_required([2])
def realizar_acordo(request, titulo_id):
    """
    Gera acordo para um título:
      - calcula juros por atraso (8% a.m. pró-rata por dia);
      - marca título principal como negociado (statusBaixa=3);
      - cria N parcelas filhas (statusBaixa=3) sem forma de pagamento;
      - forma de pagamento é salva apenas na ENTRADA.
    """
    # Carrega o título (precisa estar vinculado a um devedor)
    titulo = get_object_or_404(Titulo, id=titulo_id, devedor_id__isnull=False)

    # --- Cálculo de juros por atraso ---
    diferenca_dias = (
        (date.today() - titulo.dataVencimento).days
        if titulo.dataVencimento else 0
    )
    juros_mensais = Decimal("0.08")  # 8% a.m. — ajuste se necessário
    valor_base = Decimal(str(titulo.valor or 0))

    if diferenca_dias > 0:
        juros_totais = (valor_base * juros_mensais) * Decimal(diferenca_dias) / Decimal(30)
    else:
        juros_totais = Decimal("0.00")

    juros_totais = juros_totais.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)

    # persiste o campo juros no título base (mantém histórico)
    titulo.juros = float(juros_totais)
    titulo.save(update_fields=["juros"])

    data_vencimento_formatada = (
        titulo.dataVencimento.strftime('%d/%m/%Y') if titulo.dataVencimento else None
    )

    if request.method == 'POST':
        data = request.POST
        try:
            # --- forma de pagamento da entrada (obrigatória) ---
            forma_pag_entrada_raw = data.get('forma_pag_entrada', '')
            try:
                forma_pag_entrada = int(forma_pag_entrada_raw)
            except (TypeError, ValueError):
                raise ValueError("Selecione a forma de pagamento da entrada.")

            if forma_pag_entrada not in FORMAS_PAGAMENTO:
                raise ValueError("Forma de pagamento da entrada inválida.")

            # --- valores vindos do form ---
            entrada = Decimal(data.get('entrada', '0') or '0').quantize(
                Decimal("0.01"), rounding=ROUND_HALF_UP
            )
            qtde_prc = int(data.get('qtde_prc', '0') or 0)
            venc_primeira_parcela = data.get('venc_primeira_parcela')

            # listas de parcelas (aceita 'parcelas_valores[]' ou 'parcelas_valor[]')
            valores_raw = data.getlist('parcelas_valores[]') or data.getlist('parcelas_valor[]')
            datas_list = data.getlist('parcelas_data[]')

            valores_list = [
                Decimal(v or '0').quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
                for v in valores_raw
            ]

            # --- validações ---
            if qtde_prc <= 0:
                raise ValueError("Informe a quantidade de parcelas.")
            if len(valores_list) != qtde_prc or len(datas_list) != qtde_prc:
                raise ValueError("Quantidade de parcelas não confere com os valores/datas enviados.")
            if any(v <= 0 for v in valores_list):
                raise ValueError("Todas as parcelas devem ter valor maior que zero.")

            soma_parcelas = sum(valores_list).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)

            # --- Atualiza título principal (entrada) ---
            titulo.statusBaixa = 3
            titulo.valorRecebido = float(entrada)
            titulo.total_acordo = float(entrada)          # se o campo existir
            titulo.valor_parcela = None                   # se o campo existir
            titulo.qtde_parcelas = None                   # se o campo existir
            titulo.forma_pag_Id = forma_pag_entrada       # apenas ENTRADA
            titulo.primeiro_vencimento = venc_primeira_parcela  # se o campo existir
            titulo.juros = float(juros_totais)
            titulo.save()

            # --- Cria parcelas filhas (status 3), sem forma_pag_Id ---
            for i in range(qtde_prc):
                if datas_list[i]:
                    data_venc = datetime.strptime(datas_list[i], '%Y-%m-%d').date()
                else:
                    base = (
                        datetime.strptime(venc_primeira_parcela, '%Y-%m-%d').date()
                        if venc_primeira_parcela else date.today()
                    )
                    data_venc = base + relativedelta(months=i)

                kwargs = dict(
                    idTituloRef=titulo.id,
                    num_titulo=titulo.num_titulo,
                    tipo_doc_id=titulo.tipo_doc_id,
                    dataEmissao=date.today(),
                    dataVencimento=data_venc,
                    dataVencimentoReal=data_venc,
                    dataVencimentoPrimeira=(
                        venc_primeira_parcela if i == 0 else None
                    ),
                    valor=float(valores_list[i]),
                    qtde_parcelas=qtde_prc,
                    nPrc=i + 1,
                    statusBaixa=3,
                    devedor_id=titulo.devedor_id,
                )

                # Se sua coluna 'forma_pag_Id' NÃO aceitar NULL nas parcelas,
                # descomente a linha abaixo para herdar da entrada:
                # kwargs["forma_pag_Id"] = forma_pag_entrada

                Titulo.objects.create(**kwargs)

            messages.success(request, "Acordo realizado com sucesso!")

            # --- E-mail opcional para a empresa do devedor ---
            email_destinatario = (
                getattr(getattr(titulo.devedor, 'empresa', None), 'email', None)
            )
            if email_destinatario:
                total_geral = (entrada + soma_parcelas).quantize(Decimal("0.01"))
                try:
                    enviar_email_tipo_envio("Negociacao", email_destinatario, {
                        "core_empresa.nome_contato": getattr(titulo.devedor.empresa, 'nome_contato', '') or "Cliente",
                        "core_empresa.nome_fantasia": getattr(titulo.devedor.empresa, 'nome_fantasia', ''),
                        "titulo.id": titulo.id,
                        "titulo.valor": f"R$ {total_geral:,.2f}",
                        "titulo.valorNegociado": f"R$ {entrada:,.2f}",
                        "titulo.total_acordo": f"R$ {entrada:,.2f}",
                        "titulo.entrada": f"R$ {entrada:,.2f}",
                        "devedores.nome": titulo.devedor.nome,
                    })
                except Exception:
                    # não quebra o fluxo se falhar
                    pass

            return redirect('listar_titulos_por_devedor', titulo.devedor.id)

        except ValueError as e:
            messages.error(request, f"Erro nos valores fornecidos: {e}")
        except Exception as e:
            messages.error(request, f"Erro inesperado: {e}")

    # GET: renderiza tela do acordo
    valor_total_com_juros = float((valor_base + juros_totais).quantize(Decimal("0.01")))
    context = {
        "titulo": titulo,
        "juros_totais": float(juros_totais),
        "diferenca_dias": diferenca_dias,
        "valor_total_com_juros": valor_total_com_juros,
        "data_vencimento_formatada": data_vencimento_formatada,
        "FORMAS_PAGAMENTO": FORMAS_PAGAMENTO_CHOICES,  # para o <select> no template
    }
    return render(request, "realizar_acordo.html", context)






def buscar_email_empresa(core_empresa_id):
    """Busca o e-mail da empresa pelo core_empresa.id, garantindo apenas um resultado."""
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT email, email_financeiro
            FROM core_empresa
            WHERE id = %s
            LIMIT 1
        """, [core_empresa_id])
        
        row = cursor.fetchone()
        if row:
            return row[0] or row[1]  # Retorna o e-mail principal ou, se não existir, o financeiro

    return None  # Se não encontrar nada, retorna None

def buscar_email_empresa(core_empresa_id, devedor_id):
    """Busca o e-mail da empresa pelo `core_empresa.id` ou, se não existir, pelo `devedor.empresa_id`."""
    if core_empresa_id:
        query = "SELECT email, email_financeiro FROM core_empresa WHERE id = %s LIMIT 1"
        params = [core_empresa_id]
    else:
        print(f"⚠️ `core_empresa_id` está NULL. Tentando buscar via `devedor_id`: {devedor_id}")
        query = """
            SELECT core_empresa.email, core_empresa.email_financeiro 
            FROM devedores
            INNER JOIN core_empresa ON devedores.empresa_id = core_empresa.id
            WHERE devedores.id = %s
            LIMIT 1
        """
        params = [devedor_id]

    with connection.cursor() as cursor:
        cursor.execute(query, params)
        row = cursor.fetchone()
        if row:
            return row[0] or row[1]  # Retorna o e-mail principal ou, se não existir, o financeiro

    return None  # Se não encontrar nada, retorna None

@login_required
@group_required([2])
def quitar_parcela(request, titulo_id):
    titulo = get_object_or_404(Titulo, id=titulo_id)

    if request.method == 'POST':
        try:
            valor_recebido = float(request.POST.get('valorRecebido'))
            data_baixa = request.POST.get('dataBaixa')
            forma_pagamento = int(request.POST.get('formaPagamento'))

            # Atualizar o título
            titulo.valorRecebido = valor_recebido
            titulo.data_baixa = data_baixa
            titulo.forma_pag_Id = forma_pagamento  # Salvar a forma de pagamento no banco
            titulo.statusBaixa = 2  # Alterar status para Quitado
            titulo.save()

            # Buscar e-mail da empresa usando `core_empresa.id` ou `devedor.empresa_id`
            email_destinatario = buscar_email_empresa(titulo.empresa_id, titulo.devedor.id)

            if email_destinatario:
                print(f"🔄 Tentando enviar e-mail de quitação para {email_destinatario}...")

                sucesso_email = enviar_email_tipo_envio("Quitação Parcela", email_destinatario, {
                    "core_empresa.nome_contato": titulo.devedor.empresa.nome_contato or "Cliente",
                    "core_empresa.nome_fantasia": titulo.devedor.empresa.nome_fantasia,
                    "titulo.id": titulo.id,
                    "titulo.valorRecebido": f"R$ {titulo.valorRecebido:,.2f}",
                    "devedores.nome": titulo.devedor.nome
                })

                if sucesso_email:
                    print(f"✅ E-mail de quitação enviado com sucesso para {email_destinatario}!")
                else:
                    print("⚠️ Falha ao enviar e-mail de quitação.")
            else:
                print(f"⚠️ Nenhum e-mail encontrado para a empresa (ID: {titulo.empresa_id}) ou via devedor (ID: {titulo.devedor.id}).")

            messages.success(request, f"Parcela {titulo.num_titulo} quitada com sucesso!")

        except Exception as e:
            messages.error(request, f"Erro ao quitar parcela: {e}")
            print(f"❌ Erro ao processar quitação da parcela: {e}")

    return redirect('listar_titulos_por_devedor', titulo.devedor_id)

    
    





def default_acordo(request):
    if request.user.is_authenticated:
        acordo = Acordo.objects.first()  # Substitua com sua lógica de seleção
        return {'acordo_id': acordo.id if acordo else None}
    return {}
    
@login_required
@group_required([2])
def gerar_pdf(request, titulo_id):
    """
    Gera um PDF detalhado para o título especificado usando ReportLab.
    """
    try:
        # Obtenha os dados do título e acordo
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT 
                    devedores.nome AS devedor_nome,
                    devedores.cpf,
                    devedores.cnpj,
                    titulo.valor AS valor_titulo,
                    titulo.juros,
                    core_empresa.nome_fantasia AS empresa_nome_fantasia,
                    core_empresa.cnpj AS empresa_cnpj,
                    core_acordo.valor_total_negociacao,
                    core_acordo.entrada,
                    core_acordo.qtde_prc,
                    core_acordo.data_entrada,
                    core_acordo.venc_primeira_parcela,
                    core_acordo.contato,
                    core_acordo.id AS acordo_id
                FROM 
                    devedores
                INNER JOIN core_empresa ON devedores.empresa_id = core_empresa.id
                INNER JOIN titulo ON titulo.devedor_id = devedores.id
                INNER JOIN core_acordo ON core_acordo.titulo_id = titulo.id
                WHERE titulo.id = %s
            """, [titulo_id])
            acordo_data = cursor.fetchone()

        if not acordo_data:
            return HttpResponse("Acordo não encontrado.", status=404)

        # Mapear os dados para um dicionário
        acordo = {
            'devedor_nome': acordo_data[0],
            'cpf': acordo_data[1],
            'cnpj': acordo_data[2],
            'valor_titulo': acordo_data[3],
            'juros': acordo_data[4],
            'empresa_nome_fantasia': acordo_data[5],
            'empresa_cnpj': acordo_data[6],
            'valor_total_negociacao': acordo_data[7],
            'entrada': acordo_data[8],
            'qtde_prc': acordo_data[9],
            'data_entrada': acordo_data[10],
            'venc_primeira_parcela': acordo_data[11],
            'contato': acordo_data[12],
            'acordo_id': acordo_data[13],
        }

        # Obter as parcelas do acordo
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT parcela_numero, data_vencimento, valor
                FROM core_parcelamento
                WHERE acordo_id = %s
            """, [acordo['acordo_id']])
            parcelas = cursor.fetchall()

        # Criar um buffer para o PDF
        buffer = BytesIO()
        pdf = canvas.Canvas(buffer, pagesize=letter)
        width, height = letter

        # Adicionar título do PDF
        pdf.setFont("Helvetica-Bold", 12)
        pdf.drawCentredString(width / 2, height - 40, "ACORDO EXTRAJUDICIAL DE RENEGOCIAÇÃO DE DÍVIDA")

        # Adicionar detalhes do acordo
        pdf.setFont("Helvetica", 10)
       # pdf.drawString(50, height - 100, f"Nome do cliente: {acordo['devedor_nome']}")
       # pdf.drawString(50, height - 120, f"CPF/CNPJ: {acordo['cpf'] or acordo['cnpj']}")
       #  pdf.drawString(50, height - 140, f"Empresa: {acordo['empresa_nome_fantasia']}")
       # pdf.drawString(50, height - 160, f"CNPJ Empresa: {acordo['empresa_cnpj']}")
       # pdf.drawString(50, height - 180, f"Valor da Dívida: R$ {acordo['valor_titulo']:.2f}")
       # pdf.drawString(50, height - 200, f"Juros: R$ {acordo['juros']:.2f}")
       # Adicionar introdução ao contrato
       #pdf.drawString(50, height - 100, "ACORDO EXTRAJUDICIAL DE RENEGOCIAÇÃO DE DÍVIDA:")
        pdf.drawString(
            50, height - 70,
            f"Eu, {acordo['devedor_nome']}, portador do CPF/CNPJ {acordo['cpf'] or acordo['cnpj']}, confirmo a"
        )
        pdf.drawString(
            50, height - 90,
            f"Renegociação da dívida descrita acima em favor da empresa {acordo['empresa_nome_fantasia']},"
        )
        pdf.drawString(
            50, height - 110,
            f"De CNPJ {acordo['empresa_cnpj']}. Firmo este Contrato de Confissão e Renegociação de Dívida."
        )
        pdf.drawString(50, height - 130, f"Valor Total da Negociação: R$ {acordo['valor_total_negociacao']:.2f}")
        pdf.drawString(50, height - 150 , f"Entrada: R$ {acordo['entrada']:.2f}")
        pdf.drawString(50, height - 170, f"Quantidade de Parcelas: {acordo['qtde_prc']}")
        data_entrada_formatada = datetime.strptime(str(acordo['data_entrada']), '%Y-%m-%d').strftime('%d/%m/%Y')
        venc_primeira_parcela_formatada = datetime.strptime(str(acordo['venc_primeira_parcela']), '%Y-%m-%d').strftime('%d/%m/%Y')

        pdf.drawString(50, height - 190, f"Data da Entrada: {data_entrada_formatada}")
        pdf.drawString(50, height - 210, f"Vencimento da Primeira Parcela: {venc_primeira_parcela_formatada}")
       # pdf.drawString(50, height - 190, f"Data da Entrada: {acordo['data_entrada']}")
       # pdf.drawString(50, height - 210, f"Vencimento da Primeira Parcela: {acordo['venc_primeira_parcela']}")
        pdf.drawString(50, height - 230, f"Contato: {acordo['contato']}")

        
        

        # Adicionar tabela de parcelas
       # pdf.drawString(50, height - 440, "Parcelas:")
        pdf.line(50, height - 250, width - 50, height - 250)
        pdf.drawString(50, height - 270, "Parcela")
        pdf.drawString(150, height - 270, "Data de Vencimento")
        pdf.drawString(300, height - 270, "Valor")
        y = height - 290

        for parcela in parcelas:
            pdf.drawString(50, y, str(parcela[0]))
            pdf.drawString(150, y, parcela[1].strftime('%d/%m/%Y'))
            pdf.drawString(300, y, f"R$ {parcela[2]:.2f}")
            y -= 20

        # Assinatura
        pdf.drawString(50, y - 30, "Confirmo a renegociação nos termos acima.")
        pdf.line(70, y - 70, width - 70, y - 70)
        pdf.drawCentredString(width / 2, y - 80, f"{acordo['devedor_nome']}")
        pdf.drawCentredString(width / 2, y - 100, f"Assinatura")

        # Finalizar o PDF
        pdf.showPage()
        pdf.save()

        # Obter o conteúdo do PDF do buffer
        buffer.seek(0)
        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = f'inline; filename="acordo_titulo_{titulo_id}.pdf"'
        return response

    except Exception as e:
        # Log do erro para depuração
        print(f"Erro ao gerar PDF: {e}")
        return HttpResponse(f"Erro ao gerar PDF: {str(e)}", status=500)


from django.db import connection
from django.core.paginator import Paginator
from django.shortcuts import render, get_object_or_404, redirect
from django.http import HttpResponseForbidden
from django.contrib import messages
from datetime import datetime
from decimal import Decimal, InvalidOperation

# Models usados
# from .models import Titulo  # já deve existir
# from .decorators import lojista_login_required  # já deve existir

# ----------------- Helpers -----------------
def _parse_valor_ptbr(s: str) -> Decimal:
    s = (s or '').strip()
    if not s:
        return Decimal('0')
    s = s.replace('.', '').replace(',', '.')
    try:
        return Decimal(s)
    except InvalidOperation:
        return Decimal('0')

def _parse_date(s: str):
    s = (s or '').strip()
    if not s:
        return None
    for fmt in ('%Y-%m-%d', '%d/%m/%Y'):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            pass
    return None

# ----------------- Listagem -----------------
@login_required
@group_required([2])
def listar_acordos(request):
    empresa_id_sessao = request.session.get('empresa_id_sessao')
    query = request.GET.get('q', '')

    sql = """
      SELECT 
        t.id AS titulo_id,
        t.valorRecebido,
        t.data_baixa,
        t.qtde_parcelas,
        t.total_acordo,
        t.dataVencimentoPrimeira,
        d.telefone1 AS contato,
        d.nome AS devedor_nome,
        e.nome_fantasia AS empresa_nome,
        d.cpf, d.cnpj,
        t.comprovante, t.contrato
      FROM titulo t
      JOIN devedores d ON t.devedor_id = d.id
      JOIN core_empresa e ON d.empresa_id = e.id
      WHERE t.idTituloRef IS NULL
        AND (t.statusBaixa = 2 OR t.statusBaixa = 3)
        AND d.empresa_id = %s
    """
    params = [empresa_id_sessao]
    if query:
        sql += """
          AND (
            d.nome LIKE %s OR
            e.nome_fantasia LIKE %s OR
            d.cpf LIKE %s OR
            d.cnpj LIKE %s
          )
        """
        q = f'%{query}%'
        params.extend([q, q, q, q])

    sql += " ORDER BY t.id DESC"

    with connection.cursor() as cur:
        cur.execute(sql, params)
        rows = cur.fetchall()

    acordos = []
    for r in rows:
        acordo = {
            'titulo_id': r[0],
            'valorRecebido': r[1],
            'data_baixa': r[2].strftime('%d/%m/%Y') if r[2] else '',
            'qtde_parcelas': r[3],
            'total_acordo': r[4],
            'dataVencimentoPrimeira': r[5].strftime('%d/%m/%Y') if r[5] else '',
            'contato': r[6],
            'devedor_nome': r[7],
            'empresa_nome': r[8],
            'cpf': r[9],
            'cnpj': r[10],
            'contrato': r[12],
        }

        # Parcelas (filhas)
        with connection.cursor() as c2:
            c2.execute("""
              SELECT id, valor, dataVencimento, data_baixa, statusBaixa, comprovante, contrato
              FROM titulo
              WHERE idTituloRef = %s
              ORDER BY dataVencimento ASC, id ASC
            """, [acordo['titulo_id']])
            ps = c2.fetchall()

        parcelas = []
        for p in ps:
            dv = p[2]
            db = p[3]
            parcelas.append({
                'id': p[0],
                'valor': p[1],
                'data_vencimento_br': dv.strftime('%d/%m/%Y') if dv else '',
                'data_vencimento_iso': dv.strftime('%Y-%m-%d') if dv else '',
                'data_baixa_br': db.strftime('%d/%m/%Y') if db else '',
                'data_baixa_iso': db.strftime('%Y-%m-%d') if db else '',
                'status_baixa': p[4],     # 0/1/2/3
                'comprovante': p[5],
                'contrato': p[6],
            })
        acordo['parcelas'] = parcelas
        acordos.append(acordo)

    paginator = Paginator(acordos, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    return render(request, 'acordos_listar.html', {'page_obj': page_obj, 'query': query})

# ----------------- Editar parcela -----------------
@login_required
@group_required([2])
def parcela_editar(request, parcela_id: int):
    from .models import Titulo  # garante import local
    empresa_id = request.session.get('empresa_id_sessao')

    parcela = get_object_or_404(
        Titulo.objects.select_related('devedor'),
        id=parcela_id,
        devedor__empresa_id=empresa_id
    )
    if request.method != 'POST':
        return HttpResponseForbidden('Método não permitido')

    valor = _parse_valor_ptbr(request.POST.get('valor'))
    data_venc = _parse_date(request.POST.get('data_vencimento'))
    status_baixa = request.POST.get('status_baixa') or '0'
    data_baixa = _parse_date(request.POST.get('data_baixa'))

    try:
        status_baixa_int = int(status_baixa)
    except Exception:
        status_baixa_int = 0

    if status_baixa_int == 2 and not data_baixa:
        messages.error(request, 'Informe a data de pagamento para marcar como Quitado.')
        return redirect(request.POST.get('next') or 'acordos_listar')

    parcela.valor = valor
    parcela.dataVencimento = data_venc
    parcela.statusBaixa = status_baixa_int
    parcela.data_baixa = data_baixa if status_baixa_int == 2 else None
    parcela.save(update_fields=['valor','dataVencimento','statusBaixa','data_baixa'])

    messages.success(request, f'Parcela #{parcela.id} atualizada com sucesso!')
    return redirect(request.POST.get('next') or 'acordos_listar')

# ----------------- Excluir parcela -----------------
@login_required
@group_required([2])
def parcela_excluir(request, parcela_id: int):
    from .models import Titulo
    empresa_id = request.session.get('empresa_id_sessao')

    parcela = get_object_or_404(
        Titulo.objects.select_related('devedor'),
        id=parcela_id,
        devedor__empresa_id=empresa_id
    )
    if request.method != 'POST':
        return HttpResponseForbidden('Método não permitido')

    if parcela.statusBaixa == 2:
        messages.error(request, 'Parcela quitada não pode ser excluída.')
        return redirect(request.POST.get('next') or 'acordos_listar')

    if not parcela.idTituloRef:
        messages.error(request, 'Este registro não é uma parcela (é o acordo principal).')
        return redirect(request.POST.get('next') or 'acordos_listar')

    parcela.delete()
    messages.success(request, f'Parcela #{parcela_id} excluída com sucesso.')
    return redirect(request.POST.get('next') or 'acordos_listar')

from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.db import connection
from django.shortcuts import render

# mantém seu decorator de grupos
@login_required
@group_required([2])
def listar_acordos(request):
    q = (request.GET.get('q') or '').strip()
    operador_f = (request.GET.get('operador') or '').strip()
    supervisor_f = (request.GET.get('supervisor') or '').strip()

    # Nome do usuário para travar a visibilidade
    usuario_nome = (request.user.get_full_name() or request.user.username).strip()

    sql = """
        SELECT 
            t.id                            AS titulo_id,
            t.valorRecebido,
            t.data_baixa,
            t.qtde_parcelas,
            t.total_acordo,
            t.dataVencimentoPrimeira,
            d.telefone1                    AS contato,
            COALESCE(NULLIF(d.nome,''), NULLIF(d.nome_fantasia,''), d.razao_social) AS devedor_nome,
            e.nome_fantasia                AS empresa_nome,
            d.cpf,
            d.cnpj,
            t.comprovante,
            t.contrato,
            e.operador,
            e.supervisor
        FROM titulo t
        INNER JOIN devedores d   ON t.devedor_id = d.id
        INNER JOIN core_empresa e ON d.empresa_id = e.id
        WHERE t.idTituloRef IS NULL
          AND (t.statusBaixa = 2 OR t.statusBaixa = 3)
          AND e.status_empresa = 1
    """

    params = []

    # Pesquisa livre
    if q:
        sql += """
          AND (
              d.nome LIKE %s OR d.nome_fantasia LIKE %s OR d.razao_social LIKE %s OR
              e.nome_fantasia LIKE %s OR d.cpf LIKE %s OR d.cnpj LIKE %s
          )
        """
        params += [f'%{q}%'] * 6

    # Visibilidade por operador/supervisor
    if not (request.user.is_staff or request.user.is_superuser):
        # Usuário comum vê acordos onde ele é operador OU supervisor
        sql += " AND (LOWER(e.operador) = LOWER(%s) OR LOWER(COALESCE(e.supervisor,'')) = LOWER(%s))"
        params += [usuario_nome, usuario_nome]
    else:
        # Admin pode filtrar
        if operador_f:
            sql += " AND LOWER(e.operador) = LOWER(%s)"
            params.append(operador_f)
        if supervisor_f:
            sql += " AND LOWER(e.supervisor) = LOWER(%s)"
            params.append(supervisor_f)

    sql += " ORDER BY t.data_baixa DESC, t.id DESC"

    # Busca títulos/acordos
    with connection.cursor() as cur:
        cur.execute(sql, params)
        rows = cur.fetchall()

    acordos = [
        {
            'titulo_id': row[0],
            'valorRecebido': row[1],
            'data_baixa': row[2].strftime('%d/%m/%Y') if row[2] else '',
            'qtde_parcelas': row[3],
            'total_acordo': row[4],
            'dataVencimentoPrimeira': row[5].strftime('%d/%m/%Y') if row[5] else '',
            'contato': row[6],
            'devedor_nome': row[7],
            'empresa_nome': row[8],
            'cpf': row[9],
            'cnpj': row[10],
            'contrato': row[12],
            'operador': row[13],
            'supervisor': row[14],
        }
        for row in rows
    ]

    # Carrega parcelas de cada título
    for acordo in acordos:
        with connection.cursor() as cur:
            cur.execute("""
                SELECT id, valor, dataVencimento, data_baixa, statusBaixa, comprovante, contrato
                FROM titulo
                WHERE idTituloRef = %s
                ORDER BY dataVencimento ASC
            """, [acordo['titulo_id']])
            parcelas = cur.fetchall()

        acordo['parcelas'] = [
            {
                'id': p[0],
                'valor': p[1],
                'data_vencimento': p[2].strftime('%d/%m/%Y') if p[2] else '',
                'data_baixa': p[3].strftime('%d/%m/%Y') if p[3] else '',
                'status': ('Quitado' if p[4] == 2 else ('Negociado' if p[4] == 3 else 'Pendente')),
                'status_baixa': p[4],
                'comprovante': p[5],
                'contrato': p[6],
            }
            for p in parcelas
        ]

    # Paginação
    paginator = Paginator(acordos, 10)
    page_obj = paginator.get_page(request.GET.get('page'))

    return render(
        request,
        'acordos_listar.html',
        {
            'page_obj': page_obj,
            'query': q,
            'operador': operador_f,
            'supervisor': supervisor_f,
            'trava_operador': not (request.user.is_staff or request.user.is_superuser),
        }
    )


def valor_por_extenso(valor):
    unidades = [
        '', 'um', 'dois', 'três', 'quatro', 'cinco', 'seis', 'sete', 'oito', 'nove'
    ]
    dezenas = [
        '', 'dez', 'vinte', 'trinta', 'quarenta', 'cinquenta', 'sessenta', 'setenta', 'oitenta', 'noventa'
    ]
    centenas = [
        '', 'cento', 'duzentos', 'trezentos', 'quatrocentos', 'quinhentos', 'seiscentos', 'setecentos', 'oitocentos', 'novecentos'
    ]
    especiais = {
        10: 'dez', 11: 'onze', 12: 'doze', 13: 'treze', 14: 'quatorze',
        15: 'quinze', 16: 'dezesseis', 17: 'dezessete', 18: 'dezoito', 19: 'dezenove'
    }

    def numero_por_extenso(n):
        if n == 0:
            return 'zero'
        elif n < 10:
            return unidades[n]
        elif n < 20:
            return especiais[n]
        elif n < 100:
            dezena, unidade = divmod(n, 10)
            return dezenas[dezena] + (f' e {unidades[unidade]}' if unidade else '')
        elif n < 1000:
            centena, resto = divmod(n, 100)
            if n == 100:
                return 'cem'
            return centenas[centena] + (f' e {numero_por_extenso(resto)}' if resto else '')
        else:
            milhar, resto = divmod(n, 1000)
            milhar_extenso = f'{numero_por_extenso(milhar)} mil' if milhar > 1 else 'mil'
            return milhar_extenso + (f' e {numero_por_extenso(resto)}' if resto else '')

    reais, centavos = divmod(round(valor * 100), 100)
    reais_extenso = f'{numero_por_extenso(reais)} real{"s" if reais > 1 else ""}' if reais else ''
    centavos_extenso = f'{numero_por_extenso(centavos)} centavo{"s" if centavos > 1 else ""}' if centavos else ''

    if reais and centavos:
        return f'{reais_extenso} e {centavos_extenso}'
    return reais_extenso or centavos_extenso



def gerar_contrato(request, titulo_id):
    # Obter o título principal
    titulo = get_object_or_404(Titulo, id=titulo_id, idTituloRef__isnull=True)

    # Obter o devedor e empresa associados
    devedor = titulo.devedor
    empresa = devedor.empresa

    # Obter as parcelas associadas (títulos filhos)
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT id, valor, dataVencimento, data_baixa, nPrc
            FROM titulo
            WHERE idTituloRef = %s
            ORDER BY nPrc
        """, [titulo.id])
        parcelas = cursor.fetchall()

    # Função auxiliar para evitar erro com valores None
    def valor_extenso(valor):
        return valor_por_extenso(valor) if valor is not None else "Zero"

    # Preparar o contexto com os dados disponíveis
    context = {
        'devedores': {
            'nome': devedor.nome,
            'endereco': devedor.endereco,
            'cep': devedor.cep,
            'cidade': devedor.cidade,
            'uf': devedor.uf,
            'cpf': devedor.cpf,
        },
        'core_empresa': {
            'razao_social': empresa.razao_social,
            'endereco': empresa.endereco,
            'bairro': empresa.bairro,
            'cidade': empresa.cidade,
            'uf': empresa.uf,
            'cnpj': empresa.cnpj,
        },
        'titulo': {
            'valor_total_negociacao': titulo.total_acordo,
            'valor_total_negociacao_extenso': valor_extenso(titulo.total_acordo),
            'entrada': titulo.valorRecebido,
            'data_entrada': titulo.data_baixa,
            'entrada_extenso': valor_extenso(titulo.valorRecebido),
            'valor_por_parcela': titulo.parcelar_valor,
            'valor_por_parcela_extenso': valor_extenso(titulo.parcelar_valor),
            'qtde_prc': titulo.qtde_parcelas,
        },
        'parcelas': [
            {
                'parcela_numero': parcela[4],
                'data_vencimento_parcela': parcela[2],
                'valor': parcela[1],
                'valor_extenso': valor_extenso(parcela[1]),
            }
            for parcela in parcelas
        ],
    }

    # Renderizar o template
    return render(request, 'contrato_template.html', context)







@login_required
@group_required([2])
def realizar_baixa(request, titulo_id):
    print(f"Requisição recebida: {request.method}")
    titulo = get_object_or_404(Titulo, id=titulo_id)

    forma_pagamento_map = {
    0: "Pix",
    1: "Dinheiro",
    2: "Cartão de Débito",
    3: "Cartão de Crédito",
    4: "Cheque",
    5: "Depósito em Conta",
    6: "Pagamento na Loja",
    7: "Boleto Bancário",
    8: "Duplicata",
    9: "Recebimento pelo credor",  # <-- NOVO
}


    if request.method == 'POST':
        print("Received POST data:", request.POST)  # Debug incoming data

        try:
            tipo_baixa = request.POST.get('tipo_baixa')
            forma_pagamento_key = int(request.POST.get('forma_pagamento', 0))
            forma_pagamento = forma_pagamento_map.get(forma_pagamento_key, "Indefinido")

            print(f"Tipo de Baixa: {tipo_baixa}, Forma de Pagamento: {forma_pagamento}")  # Debug the type and payment method

            if tipo_baixa == 'Quitação':
                valor_quitacao = float(request.POST.get('valor_quitacao', 0))
                data_pagamento = request.POST.get('data_pagamento')

                print(f"Quitação: valor={valor_quitacao}, data_pagamento={data_pagamento}")

                # Atualiza os valores no modelo Titulo para Quitação
                Titulo.objects.filter(id=titulo.id).update(
                    data_baixa=data_pagamento,
                    valorRecebido=valor_quitacao,
                    forma_pag_Id=forma_pagamento_key,
                    statusBaixa=2  # Quitado
                )

                titulo.data_baixa = data_pagamento
                titulo.valorRecebido = valor_quitacao
                titulo.forma_pag_Id = forma_pagamento_key
                titulo.statusBaixa = 2
                titulo.save()

                #print(f"Título atualizado com sucesso: {titulo}")
               # messages.success(request, f"Baixa realizada com sucesso para Quitação via {forma_pagamento}!")

            elif tipo_baixa == 'Parcela':
                valor_parcela = float(request.POST.get('valor_parcela', 0))
                data_pagamento = request.POST.get('data_pagamento')

                print(f"Parcela: valor={valor_parcela}, data_pagamento={data_pagamento}")

                # Incrementa o valor recebido e atualiza o status do título
                Titulo.objects.filter(id=titulo.id).update(
                    valorRecebido=(titulo.valorRecebido or 0) + valor_parcela,
                    data_baixa=data_pagamento,
                    forma_pag_Id=forma_pagamento_key,
                    statusBaixa=1 if (titulo.valorRecebido or 0) + valor_parcela < titulo.valor else 2  # Parcial ou Quitado
                )

                titulo.valorRecebido = (titulo.valorRecebido or 0) + valor_parcela
                titulo.data_baixa = data_pagamento
                titulo.forma_pag_Id = forma_pagamento_key
                titulo.statusBaixa = 1 if titulo.valorRecebido < titulo.valor else 2
                titulo.save()

                print(f"Título atualizado para pagamento de parcela: {titulo}")
               # messages.success(request, f"Pagamento de parcela registrado com sucesso via {forma_pagamento}!")

            else:
                print("Tipo de Baixa inválido.")
                messages.error(request, "Tipo de Baixa inválido.")

            return redirect('titulos_listar')

        except Exception as e:
            print(f"Erro ao salvar Baixa: {e}")
            messages.error(request, f"Erro ao salvar Baixa: {e}")
            return redirect('realizar_baixa', titulo_id=titulo_id)

    diferenca_dias = (datetime.today().date() - titulo.dataVencimento).days
    juros_totais = (titulo.valor * 0.08 * (diferenca_dias / 30)) if diferenca_dias > 0 else 0

    context = {
        'titulo': titulo,
        'juros_totais': juros_totais,
        'data_vencimento_formatada': titulo.dataVencimento.strftime('%d/%m/%Y'),
    }
    return render(request, 'realizar_baixa.html', context)


    
    
    
@login_required
@group_required([2])    
def listar_parcelamentos(request):
    # Obter o termo de pesquisa
    query = request.GET.get('q', '')

    # Construir a consulta SQL com filtro, se aplicável
    sql_query = """
        SELECT 
            core_parcelamento.id, 
            core_parcelamento.parcela_numero, 
            core_parcelamento.valor,
            core_parcelamento.data_vencimento,
            core_parcelamento.data_vencimento_parcela,
            core_parcelamento.status, 
            core_parcelamento.created_at,
            core_parcelamento.acordo_id, 
            core_acordo.entrada, 
            core_acordo.qtde_prc, 
            core_acordo.contato, 
            core_acordo.titulo_id,
            devedores.nome AS devedor_nome,
            core_empresa.nome_fantasia AS empresa_nome_fantasia,
            core_parcelamento.forma_pagamento,
            devedores.cpf,
            devedores.cnpj
        FROM 
            core_parcelamento
        INNER JOIN 
            core_acordo 
        ON 
            core_parcelamento.acordo_id = core_acordo.id
        INNER JOIN 
            titulo
        ON 
            core_acordo.titulo_id = titulo.id
        INNER JOIN 
            devedores
        ON 
            titulo.devedor_id = devedores.id
        INNER JOIN 
            core_empresa
        ON 
            devedores.empresa_id = core_empresa.id
        WHERE 1=1
    """

    # Adicionar filtro baseado no termo de pesquisa
    params = []
    if query:
        sql_query += """
            AND (
                core_parcelamento.parcela_numero LIKE %s OR
                devedores.nome LIKE %s OR
                core_empresa.nome_fantasia LIKE %s OR
                core_acordo.contato LIKE %s OR
                core_acordo.titulo_id LIKE %s
            )
        """
        params = [f'%{query}%'] * 5

    sql_query += " ORDER BY core_parcelamento.parcela_numero ASC"

    # Executar a consulta
    with connection.cursor() as cursor:
        cursor.execute(sql_query, params)
        rows = cursor.fetchall()

    # Mapear os resultados para uma estrutura legível no template
    parcelamentos = [
        {
            "id": row[0],
            "parcela_numero": f"{row[1]}/{row[9]}",  # Formata parcela_numero / qtde_prc
            "valor": row[2],
            "data_vencimento": row[3],
            "data_vencimento_parcela": row[4],
            "status": row[5],
            "created_at": row[6],
            "acordo_id": row[7],
            "entrada": row[8],
            "qtde_prc": row[9],
            "contato": row[10],
            "titulo_id": row[11],
            "devedor_nome": row[12],
            "empresa_nome_fantasia": row[13],
            "agendamento_forma_pagamento": row[14],
            "cpf": row[15],
            "cnpj": row[16],
        }
        for row in rows
    ]

    # Paginação (20 itens por página)
    paginator = Paginator(parcelamentos, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Renderizar o template com os dados
    return render(
        request,
        'parcelamentos_listar.html',  # Nome do template para exibir os parcelamentos
        {'page_obj': page_obj, 'query': query}
    )



# core/views.py
import base64
from django.shortcuts import get_object_or_404, render
from django.contrib.auth.decorators import login_required
from core.decorators import group_required
from core.models import Titulo, Devedor, Empresa

@login_required
@group_required([2])
def gerar_recibo(request, titulo_id):
    titulo  = get_object_or_404(Titulo, id=titulo_id, statusBaixa=2)
    devedor = get_object_or_404(Devedor, id=titulo.devedor_id)
    empresa = get_object_or_404(Empresa, id=devedor.empresa_id)  # quem recebe

    forma_pagamento_map = {0:"Pix",1:"Dinheiro",2:"Cartão de Débito",3:"Cartão de Crédito",
                           4:"Cheque",5:"Depósito em Conta",6:"Pagamento na Loja",
                           7:"Boleto Bancário",8:"Duplicata"}
    forma_pagamento = forma_pagamento_map.get(getattr(titulo, "forma_pag_Id", None), "Não definido")

    base_str = f"{titulo.id}:{int(float(titulo.valorRecebido)*100)}:{titulo.data_baixa:%Y%m%d}" if titulo.data_baixa else f"{titulo.id}"
    autenticacao_token = base64.urlsafe_b64encode(base_str.encode()).decode()[:20]

    # >>> URL da logo (segura)
    logo_url = ""
    try:
        if getattr(empresa, "logo", None) and getattr(empresa.logo, "name", ""):
            # absoluta para evitar bloqueio em impressão
            logo_url = request.build_absolute_uri(empresa.logo.url)
    except Exception:
        logo_url = ""

    context = {
        "empresa": empresa,          # objeto inteiro (demais dados)
        "logo_url": logo_url,        # <-- use isto no template

        "devedor": {"nome": devedor.nome or devedor.razao_social or "",
                    "cpf_cnpj": devedor.cpf or devedor.cnpj or ""},
        "parcela": {"numero": getattr(titulo, "nPrc", None),
                    "qtde_total": getattr(titulo, "qtde_parcelas", None),
                    "data_vencimento": getattr(titulo, "dataVencimento", None),
                    "data_pagamento": getattr(titulo, "data_baixa", None),
                    "valor_pago": getattr(titulo, "valorRecebido", 0),
                    "forma_pagamento": forma_pagamento},
        "titulo": titulo,
        "numero_titulo": getattr(titulo, "num_titulo", getattr(titulo, "nDoc", titulo.id)),
        "observacao_acordo": getattr(titulo, "acordo_id", titulo.id),
        "consultor": getattr(empresa, "operador", "") or getattr(empresa, "supervisor", "") or "",
        "autenticacao_token": autenticacao_token,
        "data_autenticacao": (titulo.data_baixa if titulo.data_baixa else None),
    }
    return render(request, "recibo.html", context)



@login_required
@group_required([2])
@require_POST
def pagar_parcela(request, parcela_id):
    # Obter os dados do formulário
    valor_pago = float(request.POST.get('valor_pago', 0))
    data_baixa = request.POST.get('data_baixa')
    forma_pagamento = request.POST.get('forma_pagamento')  # Captura a forma de pagamento

    # Validar se a parcela existe no banco de dados
    with connection.cursor() as cursor:
        cursor.execute("SELECT valor FROM core_parcelamento WHERE id = %s", [parcela_id])
        parcela = cursor.fetchone()

        if not parcela:
            messages.error(request, "Parcela não encontrada.")
            return redirect('listar_parcelamentos')

        valor_original = parcela[0]

        # Atualizar a parcela com os dados fornecidos
        cursor.execute("""
            UPDATE core_parcelamento
            SET status = %s, data_baixa = %s, forma_pagamento = %s
            WHERE id = %s
        """, ["Quitado", data_baixa, forma_pagamento, parcela_id])

    # Exibir mensagem de sucesso e redirecionar
   # messages.success(request, f"Parcela {parcela_id} atualizada com sucesso.")
    return redirect('listar_parcelamentos')



from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.db.models import Q, Case, When, IntegerField
from django.shortcuts import render

@login_required
@group_required([2])
def listar_agendamentos(request):
    # Filtros vindos da URL
    query         = (request.GET.get('q') or '').strip()
    status_filter = (request.GET.get('status') or '').strip()
    operador_f    = (request.GET.get('operador') or '').strip()
    supervisor_f  = (request.GET.get('supervisor') or '').strip()
    page_number   = request.GET.get('page', 1)

    is_admin  = request.user.is_staff or request.user.is_superuser
    user_name = (request.user.get_full_name() or request.user.username).strip()

    qs = Agendamento.objects.select_related('devedor', 'empresa')

    # Busca livre
    if query:
        qs = qs.filter(
            Q(devedor__nome__icontains=query) |
            Q(devedor__nome_fantasia__icontains=query) |
            Q(devedor__razao_social__icontains=query) |
            Q(devedor__cpf__icontains=query) |
            Q(devedor__cnpj__icontains=query) |
            Q(empresa__nome_fantasia__icontains=query) |
            Q(telefone__icontains=query)
        )

    # Filtro por status
    if status_filter:
        qs = qs.filter(status=status_filter)

    # Visibilidade / filtros de Operador & Supervisor
    if not is_admin:
        # Usuário comum enxerga somente o que ele toca
        qs = qs.filter(
            Q(operador__iexact=user_name) |
            Q(empresa__operador__iexact=user_name) |
            Q(empresa__supervisor__iexact=user_name)
        )
        operador_f = user_name  # travado no form
    else:
        if operador_f:
            qs = qs.filter(
                Q(operador__iexact=operador_f) |
                Q(empresa__operador__iexact=operador_f)
            )
        if supervisor_f:
            qs = qs.filter(empresa__supervisor__iexact=supervisor_f)

    # Ordenação: prioridade de status + data de retorno (mais recentes primeiro)
    qs = qs.annotate(
        status_priority=Case(
            When(status='Pendente',   then=3),
            When(status='Agendado',   then=2),
            When(status='Finalizado', then=1),
            default=0,
            output_field=IntegerField(),
        )
    ).order_by('-status_priority', '-data_retorno', '-id')

    # Paginação
    paginator = Paginator(qs, 10)
    agendamentos_paginados = paginator.get_page(page_number)

    # Opções para selects (somente admin)
    operadores, supervisores = [], []
    if is_admin:
        ops1 = Agendamento.objects.exclude(operador__isnull=True).exclude(operador='') \
                                  .values_list('operador', flat=True).distinct()
        ops2 = Agendamento.objects.exclude(empresa__operador__isnull=True).exclude(empresa__operador='') \
                                  .values_list('empresa__operador', flat=True).distinct()
        operadores = sorted(set(list(ops1) + list(ops2)))

        supervisores = list(
            Agendamento.objects.exclude(empresa__supervisor__isnull=True).exclude(empresa__supervisor='') \
                               .values_list('empresa__supervisor', flat=True).distinct()
        )

    return render(request, 'agendamentos_listar.html', {
        'agendamentos': agendamentos_paginados,
        'query': query,
        'status_filter': status_filter,
        'operadores': operadores,
        'supervisores': supervisores,
        'operador': operador_f,
        'supervisor': supervisor_f,
        'trava_operador': not is_admin,
    })


@login_required
@group_required([2])    
def finalizar_agendamento(request, agendamento_id):
    agendamento = get_object_or_404(Agendamento, id=agendamento_id)
    if request.method == 'POST':
        agendamento.status = 'Finalizado'
        agendamento.save()
        #messages.success(request, 'Agendamento finalizado com sucesso.')
    return redirect('listar_agendamentos')



@login_required
@group_required([2])  

def criar_agendamento(request):
    """
    View para criar um novo agendamento.
    """
    # Recupera os devedores para o template
    devedores = Devedor.objects.select_related('empresa').all()
    devedores_com_empresas = [
        {
            "id": devedor.id,
            "nome": devedor.nome,
            "empresa_id": devedor.empresa.id if devedor.empresa else None,
            "nome_fantasia": devedor.empresa.nome_fantasia if devedor.empresa else "",
            "telefone": devedor.telefone,  # Telefone do devedor
        }
        for devedor in devedores
    ]

    if request.method == 'POST':
        try:
            # Captura os dados do formulário
            devedor_id = request.POST['devedor_id']
            telefone = request.POST['telefone']
            data_abertura = make_aware(datetime.strptime(request.POST['data_abertura'], "%Y-%m-%dT%H:%M"))
            data_retorno = make_aware(datetime.strptime(request.POST['data_retorno'], "%Y-%m-%dT%H:%M"))
            operador = request.POST.get('operador', '')
            assunto = request.POST.get('assunto', '')  # Captura o assunto

            # Recupera o devedor e a empresa associada
            devedor = get_object_or_404(Devedor, id=devedor_id)
            empresa = devedor.empresa

            # Cria o agendamento
            Agendamento.objects.create(
                devedor=devedor,
                empresa=empresa,
                telefone=telefone,
                data_abertura=data_abertura,
                data_retorno=data_retorno,
                operador=operador,
                assunto=assunto,  # Atribui o assunto aqui
                status='Pendente',  # Status inicial
            )

            # Exibe mensagem de sucesso
            #messages.success(request, "Agendamento criado com sucesso!")
            #print("Agendamento criado com sucesso!")  # Log para depuração

            # Redireciona para a lista de agendamentos
            return redirect('listar_agendamentos')

        except Exception as e:
            # Em caso de erro
            messages.error(request, f"Erro ao criar agendamento: {e}")
            print(f"Erro ao criar agendamento: {e}")  # Log para depuração

    return render(request, 'agendamentos_criar.html', {'devedores': devedores_com_empresas})



@login_required
@group_required([2])
def anexar_comprovante(request, parcela_id):
    # permitir anexar em qualquer Titulo (parcela ou não)
    parcela = get_object_or_404(Titulo, id=parcela_id)

    if request.method == 'POST' and 'comprovante' in request.FILES:
        comprovante = request.FILES['comprovante']

        extension = os.path.splitext(comprovante.name)[1]
        unique_filename = f"{uuid.uuid4()}{extension}"

        parcela.comprovante.save(unique_filename, comprovante)
        parcela.save()

        next_url = request.POST.get('next') or request.META.get('HTTP_REFERER') or reverse('acordos_listar')
        return redirect(next_url)

    messages.error(request, "Falha ao anexar o comprovante. Tente novamente.")
    return redirect(request.META.get('HTTP_REFERER', 'acordos_listar'))

import mimetypes
from django.http import FileResponse

def baixar_comprovante(request, titulo_id):
    titulo = get_object_or_404(Titulo, id=titulo_id)
    if not titulo.comprovante:
        return HttpResponse("Nenhum comprovante disponível.", status=404)

    path = titulo.comprovante.path
    mime, _ = mimetypes.guess_type(path)
    return FileResponse(open(path, 'rb'),
                        as_attachment=False,
                        filename=os.path.basename(path),
                        content_type=mime or 'application/octet-stream')




def baixar_comprovante(request, titulo_id):
    titulo = Titulo.objects.get(id=titulo_id)
    if titulo.comprovante:
        comprovante_path = titulo.comprovante.path  # Usando o atributo .path que já considera o MEDIA_ROOT

        with open(comprovante_path, 'rb') as f:
            response = HttpResponse(f.read(), content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="{os.path.basename(comprovante_path)}"'
            return response
    else:
        return HttpResponse("Nenhum comprovante disponível.", status=404)



# views.py
import re
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.utils.timezone import now

from django.db.models import Sum, F, DecimalField, ExpressionWrapper, Value
from django.db.models.functions import Coalesce, Cast

from .models import Titulo, Agendamento, FollowUp
from .decorators import group_required
from .utils import consultar_obito  # helper externo

# ====== helpers/constantes ======
DEC_FIELD = DecimalField(max_digits=12, decimal_places=2)
ZERO_DEC = Value(0, output_field=DEC_FIELD)


def _format_brl(valor):
    try:
        return f"{float(valor):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    except Exception:
        return "0,00"


@login_required
@group_required([2])
def detalhes_devedor(request, titulo_id):
    titulo = get_object_or_404(Titulo, id=titulo_id)
    devedor = titulo.devedor
    hoje = now().date()

    if not devedor:
        messages.error(request, "Devedor associado a este título não foi encontrado.")
        return redirect('lista_titulos')

    # -------------------- QuerySets --------------------
    titulos_entrada = (
        Titulo.objects.filter(devedor=devedor, idTituloRef__isnull=True)
        .select_related('empresa', 'devedor')
    )
    titulos_associados = (
        Titulo.objects.filter(devedor=devedor, idTituloRef__isnull=False)
        .select_related('empresa', 'devedor')
    )
    titulos = (
        Titulo.objects.filter(devedor=devedor)
        .select_related('empresa', 'devedor')
    )

    # -------------------- Juros / atraso --------------------
    for t in titulos:
        if t.dataVencimento and t.dataVencimento < hoje:
            diferenca_dias = (hoje - t.dataVencimento).days
            juros_mensais = 0.08  # 8% a.m.
            juros_totais = (t.valor * juros_mensais) * (diferenca_dias / 30)
            t.juros = round(juros_totais, 2)
            t.dias_atraso = diferenca_dias
            t.save(update_fields=['juros', 'dias_atraso'])

    # -------------------- Totais --------------------
    total_quitado = titulos.filter(statusBaixa=2).aggregate(total=Sum('valor'))['total'] or 0
    total_negociado = titulos.filter(statusBaixa=3).aggregate(total=Sum('valor'))['total'] or 0
    total_pendente_valor = titulos.filter(statusBaixa=0).aggregate(total=Sum('valor'))['total'] or 0

    # Saldo pendente (valor + juros) dos pendentes — tudo decimal
    pendentes = titulos.filter(statusBaixa=0)
    valor_base_expr = ExpressionWrapper(
        Coalesce(Cast(F('valor'), DEC_FIELD), ZERO_DEC) + Coalesce(Cast(F('juros'), DEC_FIELD), ZERO_DEC),
        output_field=DEC_FIELD
    )
    saldo_pendente = pendentes.aggregate(total=Coalesce(Sum(valor_base_expr), ZERO_DEC))['total'] or 0

    forma_pagamento_map = {
        0: "Pix", 1: "Dinheiro", 2: "Cartão de Débito", 3: "Cartão de Crédito",
        4: "Cheque", 5: "Depósito em Conta", 6: "Pagamento na Loja",
        7: "Boleto Bancário", 8: "Duplicata",
    }

    empresa = getattr(devedor, 'empresa', None)
    agendamentos = Agendamento.objects.filter(devedor=devedor)
    follow_ups = FollowUp.objects.filter(devedor=devedor).order_by('-created_at')

    # -------------------- Máscara doc --------------------
    def mascarar_documento(documento):
        documento_limpo = re.sub(r'\D', '', documento or '')
        if len(documento_limpo) == 11:
            return f"{documento_limpo[:3]}.{documento_limpo[3:6]}.xxx.xx"
        elif len(documento_limpo) == 14:
            return f"{documento_limpo[:2]}.{documento_limpo[2:5]}.{documento_limpo[5:6]}xxx-xx"
        return "N/A"

    cpf_cnpj = devedor.cpf or devedor.cnpj
    cpf_cnpj_mascarado = mascarar_documento(cpf_cnpj) if cpf_cnpj else "N/A"
    nome_consultor = request.user.get_full_name() or request.user.username
    nome_credor = (empresa.nome_fantasia if empresa else None) or "NomeCredor"

    # -------------------- PIX (empresa) --------------------
    # Chave: hoje está em empresa.banco (pelo seu template); fallback para 'chave_pix' se existir.
    pix_chave = (getattr(empresa, 'banco', '') or getattr(empresa, 'chave_pix', '') or '').strip() if empresa else ''
    # Tipo: aceita int (Choice) ou string; mapeamos os valores mais comuns
    pix_tipo_raw = (getattr(empresa, 'tipo_chave_pix', '') or getattr(empresa, 'pix_tipo', '') or '').strip() if empresa else ''
    _map_tipo = {0: "CPF", 1: "CNPJ", 2: "E-mail", 3: "Telefone", 4: "Aleatória"}
    try:
        pix_tipo = _map_tipo.get(int(pix_tipo_raw), str(pix_tipo_raw).upper() or 'CHAVE')
    except (TypeError, ValueError):
        pix_tipo = (str(pix_tipo_raw).upper() or 'CHAVE')
    # Favorecido: razão social ou fantasia
    pix_favorecido = (empresa.razao_social or empresa.nome_fantasia or '').strip() if empresa else ''

    # -------------------- Listas p/ mensagens --------------------
    nao_quitados = titulos.exclude(statusBaixa=2)
    vencidas_qs = nao_quitados.filter(dataVencimento__lt=hoje)
    a_vencer_qs = nao_quitados.filter(dataVencimento__gte=hoje)

    qtde_vencidas = vencidas_qs.count()
    lista_vencidas = ", ".join(
        [t.dataVencimento.strftime("%d/%m/%Y") for t in vencidas_qs if t.dataVencimento]
    )
    total_vencidas = vencidas_qs.aggregate(
        total=Coalesce(Sum(Cast(F('valor'), DEC_FIELD)), ZERO_DEC)
    )['total'] or 0

    total_quebra = nao_quitados.aggregate(
        total=Coalesce(Sum(Cast(F('valor'), DEC_FIELD)), ZERO_DEC)
    )['total'] or 0

    # -------------------- Templates --------------------
    tpl_vencidas = (
        "Olá %Nome%\n"
        "CPF: %CpfCnpjMascarado%\n\n"
        "Me chamo %NomeConsultor% e tenho uma mensagem importante para você.\n\n"
        "Lamentamos que não tenha pago a (s) parcela (s) referente ao acordo firmado junto a empresa %NomeCredor%.\n\n"
        "Parcelas a serem acionadas por meios jurídicos em 5 (cinco) dias úteis.\n\n"
        "Parcelas em aberto: %QtdeParcelas%\n"
        "Vencimentos: %ListaVencimentosParcelas%\n\n"
        "Valor total em aberto: R$ %ValorTotalParcelas%\n\n"
        "Pague agora mesmo por PIX (%PixTipo%)\n"
        "Favorecido: %PixFavorecido%\n"
        "Chave: *%PixChave%*\n\n"
        "Nos encaminhe o comprovante e já lhe respondemos com o recibo de pagamento.\n\n"
        "Solicitamos sua imediata atenção, hoje seu acordo está sendo encaminhado para negativação junto aos órgãos de proteção ao crédito SPC/SERASA/BOA VISTA e em seguida para análise e acionamento jurídico junto à comarca de sua cidade\n\n"
        "PROTOCOLO JURÍDICO 160120\n\n"
        "Converse com Negociar Cobranças no WhatsApp: wa.me/5591991600118\n"
        "®NEGOCIAR COBRANÇA\n"
        "CNPJ: 12.855.602/0001-74"
    )

    tpl_a_vencer = (
        "Olá %Nome%\n\n"
        "ME chamo %NomeConsultor%, gostaria de lembrar sobre a (as ) parcela (s) a vencer referente ao acordo firmado junto a empresa %NomeCredor%.\n\n"
        "Solicito que ao efetuar o pagamento da mesma, que encaminhe o comprovante nesse meu contato ou no contato direto do nosso financeiro 91991600118\n"
        "Fico no aguardo do envio do comprovante\n\n"
        "NOSSO CANAL DE ATENDIMENTO\n"
        "(91) 99160-0118\n"
        "®NEGOCIAR COBRANÇA\n"
        "Att:.\n"
        "%NomeConsultor%"
    )

    tpl_padrao = (
        "NOTIFICAÇÃO EXTRAJUDICIAL\n\n"
        "Olá, me Chamo %NomeConsultor%.\n\n"
        "Nesse contato eu falo com %Nome%, portado do documento %CpfCnpjMascarado% ?\n\n"
        "Digite - 01 - Para SIM\n"
        "Digite - 02 - Para NÃO\n\n"
        "Temos uma informação importante referente a empresa %NomeCredor%\n\n"
        "Caso eu demora a responder, me chama no contato abaixo:\n\n"
        "Central de Atendimento e Negociações: wa.me://5591991600118"
    )

    tpl_quebra = (
        "Olá, tudo bem contigo ?\n\n"
        "Notificamos %Nome%, portador do documento %CpfCnpjMascarado%, que por motivo de não identificarmos os pagamentos das parcelas do acordo feito junto a empresa %NomeCredor%, o mesmo esta sendo cancelado e encaminhado para inclusão em protesto cartorial.\n\n"
        "Em segunda estancia, será levado ao acionamento juridico no Forum da comarca de sua residencia.\n\n"
        "A divida hoje encontra-se no valor de *R$ %ValorTotalParcelas%*H\n\n"
        "Fico no aguardo do seu contato nas proxima 24 horas para evitar o procedimento acima informado.\n\n"
        "NOSSO CANAL DE ATENDIMENTO\n"
        "(91) 99160-0118\n"
        "®NEGOCIAR COBRANÇA\n\n\n"
        "atenciosamente\n"
        "%NomeConsultor%"
    )

    # -------------------- Base placeholders (inclui PIX) --------------------
    base_data = {
        "%Nome%": (devedor.nome or "").strip(),
        "%CpfCnpjMascarado%": cpf_cnpj_mascarado,
        "%NomeConsultor%": nome_consultor,
        "%NomeCredor%": nome_credor,
        "%PixTipo%": pix_tipo,
        "%PixChave%": pix_chave or "-",
        "%PixFavorecido%": pix_favorecido or (nome_credor or "-"),
    }

    # -------------------- Montagem das mensagens --------------------
    data_vencidas = {
        **base_data,
        "%QtdeParcelas%": str(qtde_vencidas),
        "%ListaVencimentosParcelas%": lista_vencidas or "-",
        "%ValorTotalParcelas%": _format_brl(total_vencidas),
    }
    msg_vencidas = tpl_vencidas
    for k, v in data_vencidas.items():
        msg_vencidas = msg_vencidas.replace(k, v)

    msg_a_vencer = tpl_a_vencer
    for k, v in base_data.items():
        msg_a_vencer = msg_a_vencer.replace(k, v)

    msg_padrao = tpl_padrao
    for k, v in base_data.items():
        msg_padrao = msg_padrao.replace(k, v)

    data_quebra = {**base_data, "%ValorTotalParcelas%": _format_brl(total_quebra)}
    msg_quebra = tpl_quebra
    for k, v in data_quebra.items():
        msg_quebra = msg_quebra.replace(k, v)

    # -------------------- Consulta de ÓBITO (API) --------------------
    obito_info = {}
    if devedor.cpf:
        try:
            obito_info = consultar_obito(devedor.cpf)
            if obito_info.get('deceased'):
                messages.warning(request, "⚠️ Registro de óbito localizado para este CPF.")
        except Exception as e:
            obito_info = {'checked': False, 'deceased': False, 'status': 'ERROR', 'error': str(e)}

    # -------------------- POST: salvar telefones --------------------
    if request.method == 'POST':
        telefone_fields = [
            'telefone', 'telefone1', 'telefone2', 'telefone3', 'telefone4', 'telefone5',
            'telefone6', 'telefone7', 'telefone8', 'telefone9', 'telefone10'
        ]
        valido_fields = [
            'telefone_valido', 'telefone1_valido', 'telefone2_valido', 'telefone3_valido',
            'telefone4_valido', 'telefone5_valido', 'telefone6_valido', 'telefone7_valido',
            'telefone8_valido', 'telefone9_valido', 'telefone10_valido'
        ]
        try:
            for f in telefone_fields:
                v = (request.POST.get(f, '') or '').strip()
                if v:
                    setattr(devedor, f, v)
            for f in valido_fields:
                v = (request.POST.get(f, '') or '').strip()
                if v:
                    setattr(devedor, f, v)
            devedor.save()
            messages.success(request, "Telefones salvos com sucesso!")
        except Exception as e:
            messages.error(request, f"Erro ao salvar os telefones: {e}")
        return redirect('detalhes_devedor', titulo_id=titulo.id)

    # -------------------- Contexto --------------------
    context = {
        'devedor': devedor,
        'titulos': titulos,
        'empresa': empresa,
        'agendamentos': agendamentos,
        'follow_ups': follow_ups,
        'telefones': [getattr(devedor, f'telefone{i}', '') for i in range(1, 11)],

        # mensagens (4 opções) + compat
        'msg_vencidas': msg_vencidas,
        'msg_a_vencer': msg_a_vencer,
        'msg_padrao': msg_padrao,
        'msg_quebra': msg_quebra,
        'mensagem_whatsapp': msg_padrao,  # compat com o que já havia no template

        'total_quitado': total_quitado,
        'total_negociado': total_negociado,
        'total_pendente': total_pendente_valor,
        'forma_pagamento_map': forma_pagamento_map,
        'today': hoje,
        'titulos_entrada': titulos_entrada,
        'titulos_associados': titulos_associados,

        # proposta
        'saldo_pendente': float(saldo_pendente),
        'percentual_avista_default': 45,
        'percentual_parc_default': 55,

        # óbito
        'obito_info': obito_info,
    }
    return render(request, 'detalhes_devedor.html', context)

    
@login_required
@group_required([2]) 
 

def lista_titulos(request):
    titulos = Titulo.objects.all()  # Ajuste a lógica conforme necessário
    return render(request, 'lista_titulos.html', {'titulos': titulos})

@login_required
@group_required([2])    

def editar_telefones(request, devedor_id):
    devedor = get_object_or_404(Devedor, id=devedor_id)
    if request.method == "POST":
        devedor.telefone1 = request.POST.get('telefone1')
        devedor.telefone2 = request.POST.get('telefone2')
        devedor.telefone3 = request.POST.get('telefone3')
        devedor.telefone4 = request.POST.get('telefone4')
        devedor.telefone5 = request.POST.get('telefone5')
        devedor.telefone6 = request.POST.get('telefone6')
        devedor.telefone7 = request.POST.get('telefone7')
        devedor.telefone8 = request.POST.get('telefone8')
        devedor.telefone9 = request.POST.get('telefone9')
        devedor.telefone10 = request.POST.get('telefone10')
        devedor.save()
       # messages.success(request, "Telefones atualizados com sucesso!")
        return redirect('detalhes_devedor', titulo_id=devedor.id)
    return redirect('detalhes_devedor', titulo_id=devedor.id)


@login_required
@group_required([2])    


def editar_agendamento(request, agendamento_id):
    agendamento = get_object_or_404(Agendamento, id=agendamento_id)
    devedores = Devedor.objects.all()
    empresas = Empresa.objects.all()  # Fetching all core_empresa
    if request.method == 'POST':
        data = request.POST
        agendamento.devedor = Devedor.objects.get(id=data['devedor_id'])
        agendamento.empresa = Empresa.objects.get(id=data['empresa_id'])
        agendamento.data_abertura = data['data_abertura']
        agendamento.data_retorno = data['data_retorno']
        agendamento.assunto = data['assunto']
        agendamento.operador = data.get('operador', '')
        agendamento.save()
        #messages.success(request, 'Agendamento atualizado com sucesso.')
        return redirect('listar_agendamentos')
    return render(request, 'agendamentos_editar.html', {
        'agendamento': agendamento,
        'devedores': devedores,
        'empresas': empresas  # Use 'empresas' here instead of 'core_empresa'
    })


# Excluir Agendamento
@login_required
@group_required([2])    
def excluir_agendamento(request, agendamento_id):
    agendamento = get_object_or_404(Agendamento, id=agendamento_id)
    agendamento.delete()
    #messages.success(request, 'Agendamento excluído com sucesso.')
    return redirect('listar_agendamentos')
 
@login_required
@group_required([2])

def listar_follow_ups(request, devedor_id):
    devedor = get_object_or_404(Devedor, id=devedor_id)
    follow_ups = FollowUp.objects.filter(devedor=devedor).order_by('-created_at')
    return render(request, 'follow_ups_listar.html', {'devedor': devedor, 'follow_ups': follow_ups})

@login_required
@group_required([2])
def adicionar_follow_up(request, devedor_id):
    if request.method == "POST":
        devedor = get_object_or_404(Devedor, id=devedor_id)
        texto = request.POST.get('texto')

        if texto:
            FollowUp.objects.create(devedor=devedor, texto=texto)
            #messages.success(request, "Follow-up adicionado com sucesso.")
        else:
            messages.error(request, "O texto do Follow-up não pode estar vazio.")
        
        # Tentar encontrar um título relacionado
        titulo = Titulo.objects.filter(devedor=devedor).first()
        if titulo:
            return redirect('detalhes_devedor', titulo_id=titulo.id)
        else:
            messages.warning(request, "Nenhum título encontrado para o devedor.")
            return redirect('lista_devedores')  # Substitua por uma view apropriada
    else:
        messages.error(request, "Método inválido.")
        return redirect('lista_devedores')  

from django.contrib.admin.views.decorators import staff_member_required
from django.core.paginator import Paginator
from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from core.models import UserAccessLog

@login_required
@staff_member_required
def listar_logs(request):
    logs = (UserAccessLog.objects
            .select_related('user')
            .order_by('-timestamp'))
    page_obj = Paginator(logs, 30).get_page(request.GET.get('page'))
    return render(request, 'logs_listar.html', {'page_obj': page_obj})



def buscar_devedores(request):
    if request.method == 'GET':
        termo = request.GET.get('termo', '').strip()
        if termo:
            devedores = Devedor.objects.filter(nome__icontains=termo)[:10]  # Limitar os resultados a 10
            resultados = [
                {
                    "id": devedor.id,
                    "nome": devedor.nome,
                    "empresa_nome": devedor.empresa.nome_fantasia if devedor.empresa else "Não associado"
                }
                for devedor in devedores
            ]
            return JsonResponse(resultados, safe=False)
        return JsonResponse([], safe=False)
        
        
def configurar_permissoes_admin():


    # Obter ou criar o grupo Admin
    admin_group, _ = Group.objects.get_or_create(name='Admin')

    # Associar todas as permissões disponíveis ao grupo Admin
    todas_permissoes = Permission.objects.all()
    admin_group.permissions.set(todas_permissoes)
    admin_group.save()

    print(f"O grupo '{admin_group.name}' agora tem todas as permissões.")
    

def listar_mensagens(request):
    mensagens = MensagemWhatsapp.objects.all()
    return render(request, 'mensagens_listar.html', {'mensagens': mensagens})

def adicionar_mensagem(request):
    if request.method == 'POST':
        form = MensagemWhatsappForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('listar_mensagens')
    else:
        form = MensagemWhatsappForm()
    return render(request, 'mensagem_adicionar.html', {'form': form})

def editar_mensagem(request, pk):
    mensagem = get_object_or_404(MensagemWhatsapp, pk=pk)
    if request.method == 'POST':
        form = MensagemWhatsappForm(request.POST, instance=mensagem)
        if form.is_valid():
            form.save()
            return redirect('listar_mensagens')
    else:
        form = MensagemWhatsappForm(instance=mensagem)
    return render(request, 'mensagem_editar.html', {'form': form})

def excluir_mensagem(request, pk):
    mensagem = get_object_or_404(MensagemWhatsapp, pk=pk)
    mensagem.delete()
    return redirect('listar_mensagens')    
    


def tabelas_listar(request):
    query = request.GET.get('q', '')
    tabelas = TabelaRemuneracao.objects.filter(nome__icontains=query) if query else TabelaRemuneracao.objects.all()
    
    paginator = Paginator(tabelas, 10)  # 10 tabelas por página
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    return render(request, 'tabelas_listar.html', {'page_obj': page_obj, 'query': query})



def tabela_adicionar(request):
    if request.method == 'POST':
        nome = request.POST.get('nome')
        if nome:
            TabelaRemuneracao.objects.create(nome=nome)
            #messages.success(request, "Tabela adicionada com sucesso!")
        return redirect('tabelas_listar')
    return render(request, 'tabela_adicionar.html')


def tabela_editar(request, tabela_id):
    tabela = get_object_or_404(TabelaRemuneracao, id=tabela_id)
    if request.method == 'POST':
        tabela.nome = request.POST.get('nome')
        tabela.save()
       # messages.success(request, "Tabela editada com sucesso!")
        return redirect('tabelas_listar')
    return render(request, 'tabela_editar.html', {'tabela': tabela})


def tabela_excluir(request, tabela_id):
    tabela = get_object_or_404(TabelaRemuneracao, id=tabela_id)
    tabela.delete()
    #messages.success(request, "Tabela excluída com sucesso!")
    return redirect('tabelas_listar')


def lista_gerenciar(request, tabela_id):
    tabela = get_object_or_404(TabelaRemuneracao, id=tabela_id)

    if request.method == 'POST':
        # Coleta os dados do formulário
        de_dias = request.POST.get('de_dias')
        ate_dias = request.POST.get('ate_dias')
        percentual_remuneracao = request.POST.get('percentual_remuneracao')

        # Criação de um novo item na tabela
        TabelaRemuneracaoLista.objects.create(
            tabela_remuneracao=tabela,
            de_dias=de_dias,
            ate_dias=ate_dias,
            percentual_remuneracao=percentual_remuneracao
        )

        # Mensagem de sucesso
        messages.success(request, "Item adicionado à lista!")
    
    # Obtém todos os itens relacionados à tabela
    itens = tabela.listas.all()

    # Renderiza o template com o nome da tabela e os itens
    return render(request, 'lista_gerenciar.html', {'tabela': tabela, 'itens': itens})



def lista_editar(request, tabela_id, item_id):
    tabela = get_object_or_404(TabelaRemuneracao, id=tabela_id)
    item = get_object_or_404(TabelaRemuneracaoLista, id=item_id, tabela_remuneracao=tabela)

    if request.method == 'POST':
        item.de_dias = request.POST.get('de_dias')
        item.ate_dias = request.POST.get('ate_dias')

        # Substituir vírgula por ponto no percentual_remuneracao
        percentual_remuneracao = request.POST.get('percentual_remuneracao', '').replace(',', '.')
        
        try:
            item.percentual_remuneracao = Decimal(percentual_remuneracao)
            item.save()
            #messages.success(request, "Item atualizado com sucesso!")
            return redirect('lista_gerenciar', tabela_id=tabela.id)
        except Exception as e:
            messages.error(request, f"Erro ao atualizar item: {e}")

    return render(request, 'lista_editar.html', {'tabela': tabela, 'item': item})
    
def lista_adicionar(request, tabela_id):
    tabela = get_object_or_404(TabelaRemuneracao, id=tabela_id)
    if request.method == 'POST':
        de_dias = request.POST.get('de_dias')
        ate_dias = request.POST.get('ate_dias')
        percentual_remuneracao = request.POST.get('percentual_remuneracao')
        if de_dias and ate_dias and percentual_remuneracao:
            TabelaRemuneracaoLista.objects.create(
                tabela_remuneracao=tabela,
                de_dias=de_dias,
                ate_dias=ate_dias,
                percentual_remuneracao=percentual_remuneracao
            )
            #messages.success(request, "Item adicionado à lista com sucesso!")
        else:
            messages.error(request, "Todos os campos são obrigatórios!")
        return redirect('lista_gerenciar', tabela_id=tabela.id)
    return render(request, 'lista_adicionar.html', {'tabela': tabela})
    
def lista_excluir(request, lista_id):
    lista = get_object_or_404(TabelaRemuneracaoLista, id=lista_id)
    tabela_id = lista.tabela_remuneracao.id  # Captura o ID da tabela para redirecionar após exclusão
    lista.delete()
    #messages.success(request, "Item da lista excluído com sucesso!")
    return redirect('lista_gerenciar', tabela_id=tabela_id)
    
# core/views.py
from django.contrib import messages
from django.shortcuts import redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.views.decorators.http import require_POST
from django.db.models.deletion import ProtectedError
from django.contrib.auth.models import User
from django.http import HttpResponseNotAllowed
from django.urls import reverse

# Torna opcional o decorator group_required para evitar quebrar import se não existir
try:
    from .decorators import group_required  # ajuste o caminho se necessário
except Exception:
    def group_required(_groups):
        def _decorator(func):
            return func
        return _decorator


from django.http import JsonResponse
from django.contrib import messages
from django.shortcuts import redirect
from django.contrib.auth.decorators import login_required
from django.views.decorators.http import require_POST
from django.db.models.deletion import ProtectedError
from django.contrib.auth.models import User
from django.urls import reverse

from .models import UsersLojistas  # <<< ESTE é o modelo listado na tela

def _delete_lojista_ou_user(lojista_id: int):
    # 1) tenta apagar um UsersLojistas (que é o que a sua tela lista)
    obj = UsersLojistas.objects.filter(pk=lojista_id).first()
    if obj:
        try:
            obj.delete()
            return True, None
        except ProtectedError:
            return False, 'Não foi possível excluir: há registros relacionados que impedem a exclusão.'
        except Exception as e:
            return False, str(e)

    # 2) fallback opcional para auth.User (remova se não fizer sentido no seu fluxo)
    user = User.objects.filter(pk=lojista_id).first()
    if user:
        try:
            user.delete()
            return True, None
        except ProtectedError:
            return False, 'Não foi possível excluir: há registros relacionados que impedem a exclusão.'
        except Exception as e:
            return False, str(e)

    return False, 'Lojista não encontrado.'

@login_required
@require_POST
def excluir_lojista(request, lojista_id):
    back = request.GET.get('next') or request.META.get('HTTP_REFERER') or reverse('usuarios_lojista_listar')
    ok, err = _delete_lojista_ou_user(lojista_id)
    if ok:
        messages.success(request, 'Registro excluído com sucesso.')
    else:
        messages.error(request, f'Erro ao excluir: {err}')
    return redirect(back)

@login_required
@require_POST
def excluir_lojista_api(request, lojista_id):
    ok, err = _delete_lojista_ou_user(lojista_id)
    if ok:
        return JsonResponse({'ok': True})
    status = 409 if err and 'impedem' in err else 404 if 'não encontrado' in (err or '').lower() else 500
    return JsonResponse({'ok': False, 'error': err or 'erro'}, status=status)

# core/views.py
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.views.decorators.http import require_POST
from django.shortcuts import redirect
from django.urls import reverse
from django.db.models.deletion import ProtectedError

from .models import UsersLojistas

@login_required
@require_POST
def excluir_lojistas_em_massa(request):
    next_url = request.POST.get('next') or reverse('usuarios_lojista_listar')
    ids = request.POST.getlist('ids')

    if not ids:
        messages.warning(request, 'Nenhum registro selecionado.')
        return redirect(next_url)

    # mantém só números
    ids = [int(x) for x in ids if str(x).isdigit()]

    ok_ids = []
    falhas = []  # (id, motivo)

    for pk in ids:
        obj = UsersLojistas.objects.filter(pk=pk).first()
        if not obj:
            falhas.append((pk, 'não encontrado'))
            continue
        try:
            obj.delete()
            ok_ids.append(pk)
        except ProtectedError:
            falhas.append((pk, 'há registros relacionados que impedem a exclusão'))
        except Exception as e:
            falhas.append((pk, str(e)))

    if ok_ids:
        messages.success(request, f'{len(ok_ids)} registro(s) excluído(s): {", ".join(map(str, ok_ids))}.')

    if falhas:
        # limita a verbosidade
        detalhes = '; '.join([f'ID {pk}: {motivo}' for pk, motivo in falhas[:10]])
        resto = len(falhas) - 10
        if resto > 0:
            detalhes += f' … e mais {resto}.'
        messages.error(request, f'Falha ao excluir {len(falhas)} registro(s): {detalhes}')

    return redirect(next_url)
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db.models.deletion import ProtectedError
from django.shortcuts import get_object_or_404, redirect
from django.urls import reverse
from django.views.decorators.http import require_POST

from .models import Devedor  # ajuste se o nome/modelo for diferente

@login_required
@require_POST
def excluir_devedor(request, id):
    """Exclusão individual por POST + redirect, com mensagens."""
    back = request.GET.get('next') or request.META.get('HTTP_REFERER') or reverse('listar_devedores')
    devedor = get_object_or_404(Devedor, pk=id)
    try:
        devedor.delete()
        messages.success(request, 'Devedor excluído com sucesso.')
    except ProtectedError:
        messages.error(request, 'Não foi possível excluir: existem registros relacionados que impedem a exclusão.')
    except Exception as e:
        messages.error(request, f'Erro ao excluir: {e}')
    return redirect(back)

@login_required
@require_POST
def excluir_devedores_em_massa(request):
    """
    Recebe vários ids via checkbox (name="ids") e exclui um a um, somando sucessos/falhas.
    Redireciona de volta (next ou listagem) com mensagens.
    """
    back = request.POST.get('next') or request.META.get('HTTP_REFERER') or reverse('listar_devedores')
    ids = request.POST.getlist('ids')

    if not ids:
        messages.warning(request, 'Nenhum devedor selecionado.')
        return redirect(back)

    ok = 0
    fail = 0
    for pk in ids:
        try:
            devedor = Devedor.objects.get(pk=pk)
            devedor.delete()
            ok += 1
        except Devedor.DoesNotExist:
            fail += 1
        except ProtectedError:
            fail += 1
        except Exception:
            fail += 1

    if ok and not fail:
        messages.success(request, f'{ok} devedor(es) excluído(s) com sucesso.')
    elif ok and fail:
        messages.warning(request, f'{ok} excluído(s) e {fail} não excluído(s) (possíveis vínculos protegidos).')
    else:
        messages.error(request, 'Nenhum registro foi excluído.')

    return redirect(back)

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.db.models import Q
from django.db.models.deletion import ProtectedError
from django.shortcuts import redirect
from django.urls import reverse

from .models import Devedor

def _aplicar_filtro_devedores(qs, q, status):
    """NÃO decorar helpers com @login_required."""
    q = (q or '').strip()
    status = (status or '').strip()

    if q:
        qs = qs.filter(
            Q(nome__icontains=q) |
            Q(nome_fantasia__icontains=q) |
            Q(cpf__icontains=q) |
            Q(cnpj__icontains=q) |
            Q(telefone1__icontains=q) |
            Q(telefone2__icontains=q)
        )

    if status in ('Pendente', 'Quitado', 'Negociado'):
        qs = qs.filter(status_baixa=status)

    return qs

def _contar_filtrados(q, status):
    """Também sem @login_required."""
    qs = _aplicar_filtro_devedores(Devedor.objects.all(), q, status)
    return qs.count()

@login_required
def excluir_devedores_todos(request):
    if request.method != 'POST':
        messages.error(request, 'Operação inválida.')
        return redirect(reverse('listar_devedores'))

    q = request.POST.get('q', '')
    status = request.POST.get('status', '')
    next_url = request.POST.get('next') or reverse('listar_devedores')
    confirm = (request.POST.get('confirm') or '').strip().upper()

    qs = _aplicar_filtro_devedores(Devedor.objects.all(), q, status)
    total = qs.count()

    if total == 0:
        messages.info(request, 'Não há registros para excluir com o filtro atual.')
        return redirect(next_url)

    if confirm != 'EXCLUIR':
        messages.error(request, f'Para excluir {total} registro(s), digite EXCLUIR no campo de confirmação.')
        return redirect(next_url)

    try:
        with transaction.atomic():
            deletados, _ = qs.delete()
        messages.success(request, f'{deletados} registro(s) excluído(s) com sucesso.')
    except ProtectedError:
        messages.error(request, 'Não foi possível excluir: há registros relacionados que impedem a exclusão.')
    except Exception as e:
        messages.error(request, f'Erro ao excluir: {e}')

    return redirect(next_url)


# core/views.py
from django.http import JsonResponse
from django.views.decorators.http import require_GET
from django.contrib.auth.decorators import login_required
from django.db.models import Q

@login_required
@require_GET
def api_buscar_empresas(request):
    # Se a sessão expirar, login_required tentaria redirecionar.
    # Como estamos em API, garantimos JSON de auth:
    if not request.user.is_authenticated:
        return JsonResponse({'ok': False, 'error': 'auth_required'}, status=401)

    q = (request.GET.get('q') or '').strip()
    try:
        from .models import Empresa  # ajuste se o nome for outro
        qs = Empresa.objects.all()
        if q:
            qs = qs.filter(Q(razao_social__icontains=q) | Q(nome_fantasia__icontains=q) | Q(cnpj__icontains=q))
        qs = qs.order_by('razao_social')[:20]

        data = [
            {'id': e.id, 'razao_social': e.razao_social, 'nome_fantasia': e.nome_fantasia, 'cnpj': e.cnpj}
            for e in qs
        ]
        return JsonResponse({'ok': True, 'results': data})
    except Exception as e:
        # Em qualquer erro, devolva JSON (não HTML)
        return JsonResponse({'ok': False, 'error': str(e)}, status=500)


from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.db import connection
from django.shortcuts import render

@login_required
@group_required([2])
def quitados_listar(request):
    # Filtros básicos
    data_inicio = (request.GET.get('data_inicio') or '').strip()
    data_fim    = (request.GET.get('data_fim') or '').strip()
    tipo        = (request.GET.get('tipo') or '').lower().strip()  # 'parcela' | 'quitacao' | ''
    devedor     = (request.GET.get('devedor') or '').strip()
    empresa     = (request.GET.get('empresa') or '').strip()
    valor_min   = (request.GET.get('valor_min') or '').replace(',', '.').strip()
    valor_max   = (request.GET.get('valor_max') or '').replace(',', '.').strip()

    # Filtros de operador/supervisor
    operador    = (request.GET.get('operador') or '').strip()
    supervisor  = (request.GET.get('supervisor') or '').strip()

    is_admin  = request.user.is_staff or request.user.is_superuser
    user_name = (request.user.get_full_name() or request.user.username).strip()

    # Consulta base — já traz operador e supervisor da empresa
    sql = """
        SELECT
            t.data_baixa,
            t.dataVencimento,
            t.valorRecebido,
            d.nome,
            d.cpf,
            d.cnpj,
            e.nome_fantasia,
            t.idTituloRef,
            e.operador,
            e.supervisor
        FROM titulo t
        JOIN devedores d    ON d.id = t.devedor_id
        JOIN core_empresa e ON e.id = d.empresa_id
        WHERE t.data_baixa IS NOT NULL
          AND t.valorRecebido IS NOT NULL
          AND e.status_empresa = 1
    """
    params = []

    # Filtros de período/tipo
    if data_inicio:
        sql += " AND t.data_baixa >= %s"
        params.append(data_inicio)
    if data_fim:
        sql += " AND t.data_baixa <= %s"
        params.append(data_fim)
    if tipo == 'parcela':
        sql += " AND t.idTituloRef IS NOT NULL"
    elif tipo == 'quitacao':
        sql += " AND t.idTituloRef IS NULL"

    # Filtros de texto/valor
    if devedor:
        sql += " AND d.nome LIKE %s"
        params.append(f"%{devedor}%")
    if empresa:
        sql += " AND e.nome_fantasia LIKE %s"
        params.append(f"%{empresa}%")
    if valor_min:
        try:
            vmin = float(valor_min); sql += " AND t.valorRecebido >= %s"; params.append(vmin)
        except ValueError:
            valor_min = ''
    if valor_max:
        try:
            vmax = float(valor_max); sql += " AND t.valorRecebido <= %s"; params.append(vmax)
        except ValueError:
            valor_max = ''

    # Visibilidade e filtros de Operador/Supervisor
    if not is_admin:
        # trava: mostra apenas onde o logado é operador OU supervisor
        operador = user_name
        sql += " AND (LOWER(e.operador)=LOWER(%s) OR LOWER(COALESCE(e.supervisor,''))=LOWER(%s))"
        params += [operador, operador]
    else:
        if operador:
            sql += " AND LOWER(e.operador) = LOWER(%s)"
            params.append(operador)
        if supervisor:
            sql += " AND LOWER(e.supervisor) = LOWER(%s)"
            params.append(supervisor)

    # Ordenação
    sql += " ORDER BY t.data_baixa DESC, t.id DESC"

    # Executa
    with connection.cursor() as cur:
        cur.execute(sql, params)
        rows = cur.fetchall()

    # Mapeia
    quitados = [
        {
            "data_baixa":      r[0].strftime("%d/%m/%Y") if r[0] else "",
            "data_vencimento": r[1].strftime("%d/%m/%Y") if r[1] else "",
            "valor_recebido":  float(r[2]) if r[2] is not None else 0.0,
            "nome":            r[3] or "",
            "cpf":             r[4] or "",
            "cnpj":            r[5] or "",
            "empresa":         r[6] or "",
            "idTituloRef":     r[7],
            "operador":        r[8] or "",
            "supervisor":      r[9] or "",
        }
        for r in rows
    ]

    # Paginação e soma
    paginator = Paginator(quitados, 10)
    page_obj  = paginator.get_page(request.GET.get("page"))
    soma_total = sum(item["valor_recebido"] for item in quitados)

    # Listas para selects (apenas admin)
    operadores = supervisores = []
    if is_admin:
        # pega distintos da tabela de empresas ativas
        with connection.cursor() as cur:
            cur.execute("""
                SELECT DISTINCT e.operador
                FROM core_empresa e
                WHERE e.status_empresa=1 AND e.operador IS NOT NULL AND e.operador <> ''
                ORDER BY 1
            """)
            operadores = [r[0] for r in cur.fetchall()]
            cur.execute("""
                SELECT DISTINCT e.supervisor
                FROM core_empresa e
                WHERE e.status_empresa=1 AND e.supervisor IS NOT NULL AND e.supervisor <> ''
                ORDER BY 1
            """)
            supervisores = [r[0] for r in cur.fetchall()]

    context = {
        "page_obj": page_obj,
        "soma_total": soma_total,
        "data_inicio": data_inicio,
        "data_fim": data_fim,
        "tipo": tipo,
        "devedor": devedor,
        "empresa": empresa,
        "valor_min": valor_min,
        "valor_max": valor_max,
        "operador": operador,
        "supervisor": supervisor,
        "operadores": operadores,
        "supervisores": supervisores,
        "trava_operador": not is_admin,
    }
    return render(request, "quitados.html", context)


@login_required
@group_required([2])
def anexar_contrato(request, titulo_id):
    titulo = get_object_or_404(Titulo, pk=titulo_id)
    if request.method == 'POST' and 'contrato' in request.FILES:
        contrato_file = request.FILES['contrato']
        extension = os.path.splitext(contrato_file.name)[1]
        unique_filename = f"{uuid.uuid4()}{extension}"
        titulo.contrato.save(unique_filename, contrato_file)
        titulo.save()
        #messages.success(request, "Contrato anexado com sucesso!")
        return redirect('acordos_listar')
    else:
        messages.error(request, "Falha ao anexar o contrato. Tente novamente.")
        return redirect('acordos_listar')



@login_required
@group_required([2])
def baixar_contrato_view(request, titulo_id):
    titulo = get_object_or_404(Titulo, pk=titulo_id)
    contrato_path = titulo.contrato.path
    with open(contrato_path, 'rb') as fh:
        response = HttpResponse(fh.read(), content_type="application/force-download")
        response['Content-Disposition'] = f'attachment; filename="{os.path.basename(contrato_path)}"'
        return response
        
@login_required
@group_required([2])
def usuarios_lojista_listar(request):
    query = request.GET.get('q', '')  # Obtem o valor da pesquisa
    usuarios = UsersLojistas.objects.filter(
        Q(name__icontains=query) | Q(email__icontains=query) | Q(empresa__razao_social__icontains=query)
    )
    
    # Paginação
    paginator = Paginator(usuarios, 10)  # Mostra 10 usuários por página
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'usuarios_lojista_listar.html', {
        'page_obj': page_obj,
        'query': query
    })
    
@login_required
@group_required([2])
def usuarios_lojista_criar(request):
    if request.method == 'POST':
        nome = request.POST.get('name')
        email = request.POST.get('email')
        password = request.POST.get('password')
        empresa_id = request.POST.get('empresa_id')

        if UsersLojistas.objects.filter(email=email).exists():
            messages.error(request, "Email já cadastrado.")
        else:
            try:
                empresa = Empresa.objects.get(id=empresa_id)

                # Hash da senha com bcrypt
                hashed_password = bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
                hashed_password = hashed_password.replace("$2b$", "$2y$", 1)

                lojista = UsersLojistas(
                    name=nome,
                    email=email,
                    password=hashed_password,
                    empresa=empresa
                )
                lojista.save()
                #messages.success(request, "Usuário criado com sucesso.")
                return redirect('usuarios_lojista_listar')
            except Empresa.DoesNotExist:
                messages.error(request, "Empresa inválida.")

    empresas = Empresa.objects.all()
    return render(request, 'usuarios_lojista_criar.html', {'empresas': empresas})

@login_required
@group_required([2])
def usuarios_lojista_editar(request, user_id):
    user = get_object_or_404(UsersLojistas, id=user_id)

    if request.method == 'POST':
        user.name = request.POST.get('name')
        user.email = request.POST.get('email')
        new_password = request.POST.get('password')

        if new_password:
            # Hash da nova senha
            hashed_password = bcrypt.hashpw(new_password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
            hashed_password = hashed_password.replace("$2b$", "$2y$", 1)
            user.password = hashed_password

        user.save()
        #messages.success(request, "Usuário atualizado com sucesso.")
        return redirect('usuarios_lojista_listar')

    empresas = Empresa.objects.all()
    return render(request, 'usuarios_lojista_editar.html', {'user': user, 'empresas': empresas})

from django.views.decorators.http import require_http_methods
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.shortcuts import render, get_object_or_404, redirect
from django.http import JsonResponse
from django.contrib import messages

@login_required
@require_http_methods(["GET", "POST"])
def usuarios_lojista_excluir(request, usuario_id):
    usuario = get_object_or_404(User, id=usuario_id)

    if request.method == "GET":
        return render(request, "core/usuarios_lojista_confirm_delete.html", {"usuario": usuario})

    if usuario == request.user:
        return _resp(request, ok=False, error="Você não pode excluir seu próprio usuário.",
                     redirect_name="usuarios_lojista_listar", status=400)

    if usuario.is_superuser:
        return _resp(request, ok=False, error="Não é permitido excluir um superusuário.",
                     redirect_name="usuarios_lojista_listar", status=400)

    try:
        usuario.delete()
    except Exception as e:
        return _resp(request, ok=False, error=str(e),
                     redirect_name="usuarios_lojista_listar", status=500)

    return _resp(request, ok=True, redirect_name="usuarios_lojista_listar")


def _resp(request, ok, error=None, redirect_name=None, status=200):
    # Se for AJAX, responde JSON
    if request.headers.get("x-requested-with") == "XMLHttpRequest":
        return JsonResponse({"ok": ok, "error": error}, status=status)

    # Senão, usa mensagens e redireciona
    if ok:
        messages.success(request, "Usuário excluído com sucesso.")
    else:
        messages.error(request, error or "Erro ao excluir.")
    return redirect(redirect_name or "usuarios_lojista_listar")

    
    
@login_required
@group_required([2])
def emails_envio_criar(request):
    if request.method == "POST":
        email = request.POST.get("email")
        autenticacao = request.POST.get("autenticacao")
        porta = request.POST.get("porta")
        servidor_smtp = request.POST.get("servidor_smtp")
        tipo_envio = request.POST.get("tipo_envio")
        provedor = request.POST.get("provedor")
        senha = request.POST.get("senha")

        EmailEnvio.objects.create(
            email=email,
            autenticacao=autenticacao,
            porta=porta,
            servidor_smtp=servidor_smtp,
            tipo_envio=tipo_envio,
            provedor=provedor,
            senha=senha
        )

        messages.success(request, "E-mail de envio cadastrado com sucesso.")
        return redirect("emails_envio_listar")

    return render(request, "emails_envio_criar.php")

@login_required
@group_required([2])
def emails_envio_editar(request, id):
    email_envio = get_object_or_404(EmailEnvio, id=id)

    if request.method == "POST":
        email_envio.email = request.POST.get("email")
        email_envio.autenticacao = request.POST.get("autenticacao")
        email_envio.porta = request.POST.get("porta")
        email_envio.servidor_smtp = request.POST.get("servidor_smtp")
        email_envio.tipo_envio = request.POST.get("tipo_envio")
        email_envio.provedor = request.POST.get("provedor")
        email_envio.senha = request.POST.get("senha")
        email_envio.save()

        messages.success(request, "E-mail de envio atualizado com sucesso.")
        return redirect("emails_envio_listar")

    return render(request, "emails_envio_editar.php", {"email_envio": email_envio})

@login_required
@group_required([2])
def emails_envio_listar(request):
    emails = EmailEnvio.objects.all()
    return render(request, "emails_envio_listar.php", {"emails": emails})
    
    
@login_required
@group_required([2])
def email_template_listar(request):
    templates = EmailTemplate.objects.all()
    return render(request, "email_template_listar.html", {"templates": templates})

@login_required
@group_required([2])
def email_template_criar(request):
    if request.method == "POST":
        tipo_envio = request.POST.get("tipo_envio")
        mensagem = request.POST.get("mensagem")
        EmailTemplate.objects.create(tipo_envio=tipo_envio, mensagem=mensagem)
        messages.success(request, "Template de e-mail criado com sucesso.")
        return redirect("email_template_listar")
    return render(request, "email_template_criar.html")

@login_required
@group_required([2])
def email_template_editar(request, id):
    template = get_object_or_404(EmailTemplate, id=id)
    if request.method == "POST":
        template.tipo_envio = request.POST.get("tipo_envio")
        template.mensagem = request.POST.get("mensagem")
        template.save()
        messages.success(request, "Template de e-mail atualizado com sucesso.")
        return redirect("email_template_listar")
    return render(request, "email_template_editar.html", {"template": template})    
# --- IMPORTS necessários ---
# core/views.py — emitir_boletos_view (CORA)

import os, re, logging, requests
from datetime import datetime, timedelta, date
from decimal import Decimal
from urllib.parse import quote

from django.conf import settings
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db import connection
from django.shortcuts import render, redirect
from django.utils import timezone
from django.views.decorators.csrf import csrf_exempt

from core.decorators import group_required
from core.integracao_cora import cora_criar_boleto

logger = logging.getLogger(__name__)


def _clean_phone_to_e164_br(phone: str) -> str:
    if not phone:
        return ""
    digits = re.sub(r"\D", "", phone)
    digits = re.sub(r"^0+", "", digits)
    if not digits.startswith("55"):
        digits = "55" + digits
    return digits if len(digits) >= 12 else ""


def _saudacao_credor(nome: str) -> str:
    try:
        parts = [p.strip() for p in (nome or "").split(" - ")]
        return parts[-1] if len(parts) >= 2 else (nome or "")
    except Exception:
        return nome or ""


def _sexta_da_semana(ref_date):
    # 4 = sexta
    delta = (ref_date.weekday() - 4) % 7
    return ref_date - timedelta(days=delta)


def _norm_url(v: str) -> str:
    v = (v or "").strip()
    return v if v.lower().startswith(("http://", "https://")) else ""


@login_required
@group_required([2])
@csrf_exempt
def emitir_boletos_view(request):
    # -------- Janela por DATA (sexta -> +7 dias) --------
    qs_from = (request.GET.get("from") or "").strip()
    try:
        ref_date = datetime.strptime(qs_from, "%Y-%m-%d").date() if qs_from else timezone.localdate()
    except ValueError:
        ref_date = timezone.localdate()

    sexta_ref = _sexta_da_semana(ref_date)
    dti_date  = sexta_ref
    dtf_date  = sexta_ref + timedelta(days=7)

    dti_str = dti_date.strftime("%Y-%m-%d")
    dtf_str = dtf_date.strftime("%Y-%m-%d")

    # Mesmas regras com aliases diferentes
    COND_T  = """
        (
            ((t.statusBaixa = 2 OR t.statusBaixaGeral = 2) AND COALESCE(t.valorRecebido,0) > 0)
         OR ((t.statusBaixa = 3 OR t.statusBaixaGeral = 3) AND t.num_titulo = 1)
        )
    """
    COND_T2 = """
        (
            ((t2.statusBaixa = 2 OR t2.statusBaixaGeral = 2) AND COALESCE(t2.valorRecebido,0) > 0)
         OR ((t2.statusBaixa = 3 OR t2.statusBaixaGeral = 3) AND t2.num_titulo = 1)
        )
    """

    # -------- SQL principal (agregado por EMPRESA) --------
    LIST_SQL = f"""
    WITH titulos_semana AS (
      SELECT t.id, t.devedor_id, d.empresa_id,
             COALESCE(t.valorRecebido, 0) AS valor_base
      FROM titulo t
      JOIN devedores d ON t.devedor_id = d.id
      WHERE
        (t.id_cobranca IS NULL OR t.id_cobranca = '')
        AND (t.email_enviado IS NULL OR t.email_enviado = 'NAO' OR t.email_enviado = '')
        AND DATE(t.data_baixa) >= %s AND DATE(t.data_baixa) < %s
        AND {COND_T}
    ),
    base_por_devedor AS (
      SELECT empresa_id, devedor_id, SUM(valor_base) AS base_devedor
      FROM titulos_semana
      GROUP BY empresa_id, devedor_id
    ),
    hist_devedor AS (
      SELECT d.empresa_id, d.id AS devedor_id,
             MAX(GREATEST(
                 0,
                 DATEDIFF(
                   t2.data_baixa,
                   COALESCE(t2.dataVencimentoReal, t2.dataVencimento, t2.dataVencimentoPrimeira)
                 )
             )) AS dias_max_hist
      FROM titulo t2
      JOIN devedores d ON d.id = t2.devedor_id
      WHERE {COND_T2}
      GROUP BY d.empresa_id, d.id
    ),
    comissao_por_devedor AS (
      SELECT b.empresa_id, b.devedor_id, b.base_devedor, h.dias_max_hist,
             ROUND(
               b.base_devedor * (
                 CASE
                   WHEN h.dias_max_hist BETWEEN  30 AND   90 THEN 0.09
                   WHEN h.dias_max_hist BETWEEN  91 AND  180 THEN 0.15
                   WHEN h.dias_max_hist BETWEEN 181 AND  720 THEN 0.21
                   WHEN h.dias_max_hist BETWEEN 721 AND 1825 THEN 0.30
                   WHEN h.dias_max_hist >= 1826             THEN 0.40
                   ELSE 0
                 END
               ), 2
             ) AS comissao_devedor
      FROM base_por_devedor b
      JOIN hist_devedor h ON h.empresa_id = b.empresa_id AND h.devedor_id = b.devedor_id
    ),
    empresa_aggr AS (
      SELECT empresa_id,
             SUM(base_devedor)      AS valor_recebido_total,
             MAX(dias_max_hist)     AS dias_max_emp_hist,
             SUM(comissao_devedor)  AS comissao_total
      FROM comissao_por_devedor
      GROUP BY empresa_id
    )
    SELECT
      e.id AS empresa_id, e.razao_social AS razao, e.cnpj AS cnpj,
      e.endereco AS end, e.bairro AS bairro, e.cidade AS cidade, e.uf AS uf,
      e.cep AS cep, e.telefone AS fone,
      (
        SELECT GROUP_CONCAT(DISTINCT d2.id ORDER BY d2.id SEPARATOR ',')
        FROM devedores d2
        JOIN titulo t2 ON t2.devedor_id = d2.id
        WHERE d2.empresa_id = e.id
          AND (t2.id_cobranca IS NULL OR t2.id_cobranca = '')
          AND (t2.email_enviado IS NULL OR t2.email_enviado = 'NAO' OR t2.email_enviado = '')
          AND DATE(t2.data_baixa) >= %s AND DATE(t2.data_baixa) < %s
          AND {COND_T2}
      ) AS devedores,
      ea.dias_max_emp_hist AS dias_max,
      ea.valor_recebido_total AS valor_recebido,
      ea.comissao_total AS comissao_total
    FROM empresa_aggr ea
    JOIN core_empresa e ON e.id = ea.empresa_id
    ORDER BY e.razao_social ASC
    """

    with connection.cursor() as cur:
        cur.execute(LIST_SQL, [dti_str, dtf_str, dti_str, dtf_str])
        results = cur.fetchall()

    # Totais
    tot_valor_recebido = Decimal("0.00")
    tot_comissao = Decimal("0.00")
    for r in results:
        tot_valor_recebido += Decimal(str(r[11] or 0))
        tot_comissao += Decimal(str(r[12] or 0))
    tot_empresas = len(results)

    # -------- Detalhes por empresa (para o modal) --------
    detalhes_por_empresa = {}
    empresa_ids = [row[0] for row in results]
    if empresa_ids:
        ph = ",".join(["%s"] * len(empresa_ids))
        DETAILS_SQL = f"""
        SELECT
          e.id AS empresa_id,
          t.id,
          COALESCE(NULLIF(d.nome,''), NULLIF(d.nome_fantasia,''), d.razao_social) AS nome_devedor,
          t.num_titulo,
          COALESCE(t.valorRecebido,0) AS valor,
          DATE_FORMAT(COALESCE(t.dataVencimentoReal, t.dataVencimento, t.dataVencimentoPrimeira), '%%d/%%m/%%Y') AS vencimento,
          CASE
            WHEN ((t.statusBaixa=2 OR t.statusBaixaGeral=2) AND COALESCE(t.valorRecebido,0)>0) THEN 'Quitado'
            WHEN  (t.statusBaixa=3 OR t.statusBaixaGeral=3) THEN 'Negociado'
            ELSE 'Pendente'
          END AS status_txt,
          DATE_FORMAT(t.data_baixa, '%%d/%%m/%%Y') AS data_baixa_fmt
        FROM titulo t
        JOIN devedores d ON d.id = t.devedor_id
        JOIN core_empresa e ON e.id = d.empresa_id
        WHERE e.id IN ({ph})
          AND (t.id_cobranca IS NULL OR t.id_cobranca = '')
          AND (t.email_enviado IS NULL OR t.email_enviado = '' OR t.email_enviado = 'NAO')
          AND {COND_T}
          AND DATE(t.data_baixa) >= %s AND DATE(t.data_baixa) < %s
        ORDER BY e.id, t.data_baixa DESC, t.id DESC
        """
        with connection.cursor() as cur:
            cur.execute(DETAILS_SQL, [*empresa_ids, dti_str, dtf_str])
            rows = cur.fetchall()
        for (emp_id, t_id, nome_dev, num_tit, valor, venc, status_txt, data_bx) in rows:
            detalhes_por_empresa.setdefault(emp_id, []).append({
                "id": t_id,
                "nome": nome_dev,
                "num_titulo": num_tit,
                "valor": float(valor or 0),
                "vencimento": venc,
                "status": status_txt,
                "data_baixa": data_bx,
            })

    # =========================
    #   POST: EMITIR (CORA)
    # =========================
    if request.method == "POST" and request.POST.get("emitir"):
        try:
            empresa_id = (request.POST.get("empresa_id") or "").strip()
            if not empresa_id:
                raise ValueError("Empresa não informada.")

            row = next((r for r in results if str(r[0]) == empresa_id), None)
            if not row:
                raise ValueError("Empresa fora da janela atual. Atualize a página.")

            (emp_id, razao, cnpj, endereco, bairro, cidade, uf, cep,
             telefone, devedor_ids, dias_max, valor_recebido, comissao_valor) = row

            comissao_dec = Decimal(str(comissao_valor or 0)).quantize(Decimal("0.01"))
            if comissao_dec <= 0:
                messages.error(request, "A comissão calculada é zero. Nada a emitir.")
                return redirect(request.get_full_path())

            # ------- CORA (boleto com Pix opcional) -------
            resp = cora_criar_boleto(
                valor=float(comissao_dec),
                pagador_nome=razao,
                pagador_documento=cnpj,
                vencimento=date.today().strftime("%Y-%m-%d"),
                mensagem=f"Honorários {razao}",
            )

            boleto_id  = resp.get("id") or resp.get("nosso_numero")
            linha      = (resp.get("digitable_line") or resp.get("linha_digitavel") or "").strip()
            barras     = (resp.get("barcode") or resp.get("codigo_barras") or "").strip()
            boleto_url = _norm_url(resp.get("boleto_url") or resp.get("url") or resp.get("link"))
            pdf_url    = _norm_url(resp.get("pdf_url") or resp.get("pdf"))
            pix_cc     = (resp.get("pix_brcode") or "").strip()
            pix_txid   = (resp.get("pix_txid") or "").strip()

            codigo_solicitacao = f"BOL:{(boleto_id or date.today().strftime('%Y%m%d'))}-{emp_id}"

            # Marca os títulos elegíveis
            with connection.cursor() as cur:
                cur.execute(
                    f"""
                    UPDATE titulo t
                    JOIN devedores d ON d.id = t.devedor_id
                    SET t.id_cobranca = %s
                    WHERE d.empresa_id = %s
                      AND (t.id_cobranca IS NULL OR t.id_cobranca = '')
                      AND (t.email_enviado IS NULL OR t.email_enviado = '' OR t.email_enviado = 'NAO')
                      AND {COND_T}
                      AND DATE(t.data_baixa) >= %s AND DATE(t.data_baixa) < %s
                    """,
                    [codigo_solicitacao, emp_id, dti_str, dtf_str]
                )

            # Upsert em core_boleto
            try:
                with connection.cursor() as cur:
                    cur.execute("""
                        UPDATE core_boleto
                           SET valor_nominal=%s, origem_recebimento='CORA', tipo_cobranca='BOLETO',
                               pagador_nome=%s, pagador_cpf_cnpj=%s,
                               nosso_numero=%s, linha_digitavel=%s, codigo_barras=%s,
                               boleto_url=%s, pdf_url=%s,
                               pix_copia_e_cola=%s, txid=%s
                         WHERE codigo_solicitacao=%s
                    """, [str(comissao_dec), razao, cnpj,
                          boleto_id, linha, barras, boleto_url, pdf_url,
                          pix_cc, pix_txid, codigo_solicitacao])
                    if cur.rowcount == 0:
                        cur.execute("""
                            INSERT INTO core_boleto
                                (empresa_id, codigo_solicitacao, seu_numero, situacao,
                                 data_situacao, data_emissao, data_vencimento,
                                 valor_nominal, valor_total_recebido,
                                 origem_recebimento, tipo_cobranca,
                                 pagador_nome, pagador_cpf_cnpj,
                                 nosso_numero, linha_digitavel, codigo_barras,
                                 boleto_url, pdf_url, pix_copia_e_cola, txid, cobranca_enviada_whatsapp)
                            VALUES
                                (%s, %s, NULL, 'A_RECEBER',
                                 CURDATE(), CURDATE(), CURDATE(),
                                 %s, 0,
                                 'CORA', 'BOLETO',
                                 %s, %s,
                                 %s, %s, %s,
                                 %s, %s, %s, %s, 'NAO')
                        """, [emp_id, codigo_solicitacao, str(comissao_dec),
                              razao, cnpj, boleto_id, linha, barras,
                              boleto_url, pdf_url, pix_cc, pix_txid])
            except Exception:
                logger.exception("Falha core_boleto (BOLETO/CORA)")

            # Salva PDF local (se houver URL)
            if pdf_url:
                try:
                    r = requests.get(pdf_url, timeout=30)
                    r.raise_for_status()
                    base_dir = "/home/app_admin/core/boletos"
                    os.makedirs(base_dir, exist_ok=True)
                    with open(os.path.join(base_dir, f"boleto_{codigo_solicitacao}.pdf"), "wb") as f:
                        f.write(r.content)
                except Exception as e:
                    logger.warning("PDF indisponível (%s): %s", pdf_url, e)

            # WhatsApp + contexto (inclui Pix se veio)
            nome_limpo = re.sub(r'^\s*\d+\s*-\s*', '', razao or '').strip()
            credor_exib = _saudacao_credor(nome_limpo)

            msg = [
                f"Olá, {credor_exib}\n\n",
                "Segue relatório e boleto referente aos honorários da semana.\n",
            ]
            if pix_cc:
                msg.append("Você pode pagar via *Pix Copia-e-Cola* colando o código abaixo no app do banco.\n\n")
                msg.append(pix_cc + "\n\n")
            msg.append("Se preferir, pague pelo boleto (linha digitável abaixo).\n\n")
            if linha:
                msg.append(f"Linha digitável: {linha}\n")
            if boleto_url:
                msg.append(f"Acesse: {boleto_url}\n\n")
            msg.append("Atenciosamente\nFrancisco Bordin")
            wa_msg = "".join(msg)

            request.session["wa_ctx"] = {
                "empresa_id": str(emp_id),
                "wa_numero": _clean_phone_to_e164_br(telefone or ""),
                "wa_razao": credor_exib,
                "valor": f"{comissao_dec:.2f}",
                "boleto_url": boleto_url,
                "linha_digitavel": linha,
                "pix_brcode": pix_cc,
                "wa_msg": wa_msg,
            }
            messages.success(request, f"Boleto emitido para {credor_exib}.")
            return redirect(request.get_full_path())

        except Exception as e:
            logger.exception("Falha ao emitir cobrança")
            messages.error(request, f"Falha ao emitir cobrança: {e}")
            return redirect(request.get_full_path())

    # -------- POST 2: Confirmar número e abrir WhatsApp --------
    if request.method == "POST" and request.POST.get("confirmar_wa"):
        try:
            numero_raw = (request.POST.get("numero_whatsapp") or "").strip()
            empresa_id = (request.POST.get("empresa_id") or "").strip()
            wa_msg = (request.POST.get("wa_msg") or "").strip()

            numero = _clean_phone_to_e164_br(numero_raw)
            if not numero:
                messages.error(request, "Informe um número de WhatsApp válido.")
                return redirect(request.get_full_path())

            # (Opcional) atualizar telefone da empresa
            if request.POST.get("salvar_tel_empresa") == "1" and empresa_id:
                try:
                    with connection.cursor() as cur:
                        cur.execute("UPDATE core_empresa SET telefone=%s WHERE id=%s", [numero_raw, empresa_id])
                except Exception:
                    logger.warning("Não foi possível atualizar telefone da empresa %s", empresa_id)

            wa_url = f"https://wa.me/{numero}?text={quote(wa_msg)}"
            request.session["wa_open"] = wa_url
            messages.success(request, "Número confirmado. Abriremos o WhatsApp em uma nova aba.")
            return redirect(request.get_full_path())

        except Exception as e:
            logger.exception("Falha ao confirmar WhatsApp")
            messages.error(request, f"Falha ao confirmar/enviar WhatsApp: {e}")
            return redirect(request.get_full_path())

    # -------- Contexto --------
    wa_ctx = request.session.pop("wa_ctx", None)
    wa_open_url = request.session.pop("wa_open", None)

    context = {
        "headers": ["ID Empresa","Razão Social","CNPJ","Endereço","Bairro","Cidade","UF","CEP",
                    "Telefone","ID Devedor(es)","Dias de Atraso (máx.)","Valor Recebido","Comissão (R$)"],
        "results": results,
        "detalhes_por_empresa": detalhes_por_empresa,
        "dt_from": dti_date,
        "dt_to": dtf_date,
        "tot_valor_recebido": tot_valor_recebido,
        "tot_comissao": tot_comissao,
        "tot_empresas": tot_empresas,

        # Bloco de confirmação/whatsapp
        "wa_razao": (wa_ctx or {}).get("wa_razao"),
        "wa_valor": (wa_ctx or {}).get("valor"),
        "boleto_url": (wa_ctx or {}).get("boleto_url"),
        "linha_digitavel": (wa_ctx or {}).get("linha_digitavel"),
        "pix_brcode": (wa_ctx or {}).get("pix_brcode"),
        "wa_numero": (wa_ctx or {}).get("wa_numero") or "",
        "wa_msg": (wa_ctx or {}).get("wa_msg") or "",
        "wa_empresa_id": (wa_ctx or {}).get("empresa_id") or "",
        "wa_open_url": wa_open_url,
    }
    return render(request, "boletos_emitir.html", context)


# imports no topo (alguns você já tem)


import logging
from django.shortcuts import render
from django.views.decorators.csrf import csrf_exempt

from core.integracao_banco_inter import (
    consultar_cob_por_txid,
    qrcode_png_base64,
)

logger = logging.getLogger(__name__)

@csrf_exempt  # é só leitura; deixe sem login para compartilhar o link
def pix_qr_page(request, txid: str):
    """
    Página pública: mostra o QR (PNG base64) e o Pix Copia-e-Cola de uma COB do Inter.
    Gera o QR localmente a partir do copia-e-cola, sem depender do host /qrcode do banco.
    """
    try:
        cob = consultar_cob_por_txid(txid)  # chama a API do Inter
        brcode = cob.get("pixCopiaECola") or ""
        if not brcode:
            return render(request, "pix_qr.html", {
                "erro": "Cobrança sem Pix Copia-e-Cola. Tente novamente em alguns segundos.",
                "txid": txid,
            }, status=404)

        qr_png_b64 = qrcode_png_base64(brcode, box_size=8, border=2)  # PNG base64
        valor = (cob.get("valor") or {}).get("original", "")
        desc  = cob.get("solicitacaoPagador") or ""
        return render(request, "pix_qr.html", {
            "txid": txid,
            "valor": valor,
            "descricao": desc,
            "brcode": brcode,
            "qr_png_b64": qr_png_b64,
        })
    except Exception as e:
        logger.exception("pix_qr_page erro")
        return render(request, "pix_qr.html", {
            "erro": f"Falha ao consultar a cobrança: {e}",
            "txid": txid,
        }, status=500)

@login_required
@group_required([2])
def baixar_boleto(request, codigo_solicitacao):
    """
    Faz o download do PDF do boleto com base no código de solicitação, buscando o arquivo localmente.
    """
    try:
        # Tentar buscar um único boleto, mas lidar com múltiplos resultados
        boletos = Boleto.objects.filter(codigo_solicitacao=codigo_solicitacao)

        if not boletos.exists():
            logger.error(f"Nenhum boleto encontrado para o código de solicitação: {codigo_solicitacao}")
            return HttpResponse("Boleto não encontrado no servidor.", status=404)

        if boletos.count() > 1:
            logger.warning(
                f"Múltiplos boletos encontrados para o código de solicitação {codigo_solicitacao}. Usando o primeiro registro.")

        # Usar o primeiro boleto encontrado
        boleto = boletos.first()

        # Caminho base onde os boletos estão armazenados
        boleto_path = os.path.join('/home/app_admin/core/boletos', f'boleto_{codigo_solicitacao}.pdf')

        # Verificar se o arquivo existe no caminho especificado
        if not os.path.exists(boleto_path):
            logger.error(f"Arquivo boleto não encontrado para o código de solicitação: {codigo_solicitacao}")
            return HttpResponse("Boleto não encontrado no servidor.", status=404)

        # Retornar o arquivo para download
        return FileResponse(
            open(boleto_path, 'rb'),
            as_attachment=True,
            filename=f"boleto_{codigo_solicitacao}.pdf",
        )
    except Exception as e:
        logger.error(f"Erro ao baixar boleto para código {codigo_solicitacao}: {e}")
        return HttpResponse(f"Erro ao baixar boleto: {str(e)}", status=500)

def dictfetchall(cursor):
    """Converte os resultados da query em uma lista de dicionários."""
    columns = [col[0] for col in cursor.description]
    return [dict(zip(columns, row)) for row in cursor.fetchall()]

def format_date(value):
    """Garante que todas as datas sejam formatadas no padrão dd/mm/yyyy"""
    if isinstance(value, datetime):
        return value.strftime('%d/%m/%Y')  # Se já for um datetime
    elif isinstance(value, str):
        try:
            return datetime.strptime(value, '%Y-%m-%d').strftime('%d/%m/%Y')  # Formato YYYY-MM-DD
        except ValueError:
            return value  # Se não conseguir converter, retorna o original
    return value

from django.db import connection
from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.contrib.auth.decorators import user_passes_test

def group_required(group_names):
    def in_groups(user):
        if user.is_authenticated:
            if user.groups.filter(name__in=group_names).exists() or user.is_superuser:
                return True
        return False
    return user_passes_test(in_groups)

def dictfetchall(cursor):
    columns = [col[0] for col in cursor.description]
    return [dict(zip(columns, row)) for row in cursor.fetchall()]

def format_date(date_str):
    from datetime import datetime
    if date_str:
        try:
            # Tenta converter a data com hora
            return datetime.strptime(date_str, "%Y-%m-%d %H:%M:%S").strftime("%d/%m/%Y")
        except ValueError:
            try:
                # Tenta converter a data sem hora
                return datetime.strptime(date_str, "%Y-%m-%d").strftime("%d/%m/%Y")
            except ValueError:
                # Se falhar, retorna o valor original ou um placeholder
                return date_str  # ou "---" se preferir
    return "---"
# imports (no topo do arquivo)
from django.http import JsonResponse
from core.integracao_banco_inter import consultar_cob_por_txid, qrcode_png_base64
@login_required
def pix_detalhes(request, ident: str):
    """
    Retorna JSON com status, copia-e-cola e QR (base64) para o txid informado.
    'ident' deve ser o txid (26-35 chars). Se você salvar loc_id no banco,
    ajuste para mapear loc_id->txid conforme sua persistência.
    """
    ident = (ident or "").strip()
    if not ident:
        return JsonResponse({"error": "ident vazio."}, status=400)

    try:
        cob = consultar_cob_por_txid(ident)
        brcode = cob.get("pixCopiaECola") or ""
        qr_b64 = qrcode_png_base64(brcode) if brcode else ""
        loc_id = (cob.get("loc") or {}).get("id")
        status = cob.get("status") or "DESCONHECIDO"
        valor = (cob.get("valor") or {}).get("original")
        criacao = (cob.get("calendario") or {}).get("criacao")
        expiracao = (cob.get("calendario") or {}).get("expiracao")

        return JsonResponse({
            "ok": True,
            "status": status,
            "brcode": brcode,
            "qr_image_b64": qr_b64,   # "data:image/png;base64,{{...}}" no front
            "txid": cob.get("txid"),
            "loc_id": loc_id,
            "valor": valor,
            "criacao": criacao,
            "expiracao": expiracao,
        })
    except Exception as e:
        return JsonResponse({"ok": False, "error": str(e)}, status=400)

# --- IMPORTS NO TOPO DO ARQUIVO (garanta que estes existam) ---
import re
from urllib.parse import quote

from django.conf import settings
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db import connection
from django.shortcuts import render, redirect
from django.utils import timezone

# Se você já tem o decorator original, pode removê-lo daqui
def group_required(_):
    def deco(fn): return login_required(fn)
    return deco

# Utilitário pra normalizar número do WhatsApp
def _clean_phone_to_e164_br(phone: str) -> str:
    if not phone:
        return ""
    digits = re.sub(r"\D", "", phone)
    digits = re.sub(r"^0+", "", digits)
    if not digits.startswith("55"):
        digits = "55" + digits
    return digits if len(digits) >= 12 else ""

# =======================
#   LISTAGEM DE PIX/INTER
# =======================
@group_required([2])
def boletos_listar_emitidos(request):
    """
    Lista cobranças (core_boleto) de origem INTER / tipo PIX que ainda
    estão vinculadas a títulos por id_cobranca. Inclui:
      - QR/link
      - Copiar brcode
      - Reenviar WhatsApp
      - Modal de detalhes (devedor/credor/títulos)
    """
    import math
    from datetime import timedelta

    # --- POST: preparar WhatsApp ---
    if request.method == "POST" and request.POST.get("reenviar_wa"):
        try:
            numero_raw = (request.POST.get("numero_whatsapp") or "").strip()
            wa_msg     = (request.POST.get("wa_msg") or "").strip()
            numero     = _clean_phone_to_e164_br(numero_raw)
            if not numero:
                messages.error(request, "Informe um número de WhatsApp válido.")
                return redirect(request.get_full_path())

            wa_url = f"https://wa.me/{numero}?text={quote(wa_msg)}"
            request.session["wa_open"] = wa_url
            messages.success(request, "WhatsApp preparado. Clique em 'Abrir WhatsApp' no topo.")
            return redirect(request.get_full_path())
        except Exception as e:
            messages.error(request, f"Falha ao preparar o WhatsApp: {e}")
            return redirect(request.get_full_path())

    # --- paginação / filtros ---
    page_number    = int(request.GET.get("page", 1))
    per_page       = 10
    offset         = (page_number - 1) * per_page

    search_query = (request.GET.get("q") or "").strip()
    data_inicio  = (request.GET.get("data_inicio") or "").strip()
    data_fim     = (request.GET.get("data_fim") or "").strip()

    COLL = "utf8mb4_unicode_ci"   # use UMA collation em ambos os lados

    where = [
        "b.origem_recebimento = 'INTER'",
        "b.tipo_cobranca = 'PIX'",
        # garantir que segue vinculado a algum título por id_cobranca
        f"""EXISTS (
             SELECT 1
             FROM titulo t
             WHERE t.id_cobranca IS NOT NULL AND t.id_cobranca <> ''
               AND (
                     -- amarra 'PIX:<txid>'
                     (b.txid IS NOT NULL AND b.txid <> '' AND
                      (CONVERT(t.id_cobranca USING utf8mb4) COLLATE {COLL}) =
                      (CONCAT('PIX:', b.txid) COLLATE {COLL})
                     )
                     -- ou usa b.codigo_solicitacao como 'PIX:<txid>'
                  OR (b.codigo_solicitacao IS NOT NULL AND b.codigo_solicitacao <> '' AND
                      (CONVERT(t.id_cobranca USING utf8mb4) COLLATE {COLL}) =
                      (CONVERT(b.codigo_solicitacao USING utf8mb4) COLLATE {COLL})
                     )
               )
           )"""
    ]
    params = []

    if search_query:
        like = f"%{search_query}%"
        where.append("(b.situacao LIKE %s OR b.pagador_nome LIKE %s OR b.pagador_cpf_cnpj LIKE %s OR b.codigo_solicitacao LIKE %s)")
        params += [like, like, like, like]

    # janela padrão: últimos 60 dias por emissão
    if data_inicio:
        where.append("b.data_emissao >= %s")
        params.append(data_inicio)
    if data_fim:
        where.append("b.data_emissao <= %s")
        params.append(data_fim)
    if not data_inicio and not data_fim:
        from datetime import date
        where.append("b.data_emissao >= %s")
        params.append((timezone.now().date() - timedelta(days=60)).isoformat())

    base_sql = f"""
        FROM core_boleto b
        LEFT JOIN core_empresa e ON e.id = b.empresa_id
        WHERE {" AND ".join(where)}
    """

    # ATENÇÃO: 'PIX:%' precisa ser 'PIX:%%' para escapar o % no MySQLdb.
    select_sql = """
        SELECT
            b.id,
            b.situacao,
            CAST(b.data_emissao    AS CHAR) AS data_emissao,
            CAST(b.data_vencimento AS CHAR) AS data_vencimento,
            b.valor_nominal,
            COALESCE(b.valor_total_recebido, 0) AS valor_total_recebido,
            b.origem_recebimento,
            b.tipo_cobranca,
            b.pagador_nome,
            b.pagador_cpf_cnpj,
            b.codigo_solicitacao,
            CAST(b.atualizado_em AS CHAR) AS atualizado_em,
            e.nome_fantasia AS empresa_nome,
            CASE
              WHEN COALESCE(NULLIF(b.txid,''), '') <> '' THEN b.txid
              WHEN b.codigo_solicitacao LIKE 'PIX:%%' THEN SUBSTRING_INDEX(b.codigo_solicitacao, ':', -1)
              ELSE NULL
            END AS txid_effective,
            b.pix_copia_e_cola AS brcode
    """

    list_sql    = f"{select_sql} {base_sql} ORDER BY b.data_emissao DESC, b.id DESC LIMIT %s OFFSET %s"
    list_params = params + [per_page, offset]

    with connection.cursor() as cur:
        cur.execute(list_sql, list_params)
        cols = [c[0] for c in cur.description]
        rows = [dict(zip(cols, r)) for r in cur.fetchall()]

    # --- totais (cards) ---
    totals_sql = f"""
        SELECT
            COALESCE(SUM(CASE WHEN b.situacao='RECEBIDO' THEN b.valor_total_recebido END),0),
            COALESCE(SUM(CASE WHEN b.situacao='A_RECEBER' THEN b.valor_nominal END),0),
            COALESCE(SUM(CASE WHEN b.situacao='ATRASADO' THEN b.valor_nominal END),0)
        {base_sql}
    """
    with connection.cursor() as cur:
        cur.execute(totals_sql, params)
        total_recebidos, total_a_receber, total_atrasados = cur.fetchone()

    # --- contagem para paginação ---
    count_sql = f"SELECT COUNT(*) {base_sql}"
    with connection.cursor() as cur:
        cur.execute(count_sql, params)
        total_boletos = cur.fetchone()[0]
    total_pages = max(1, math.ceil(total_boletos / per_page))

    # --- DETALHES por boleto (para modal) ---
    boleto_ids = [r["id"] for r in rows]
    detalhes_map = {bid: [] for bid in boleto_ids}

    if boleto_ids:
        ph = ",".join(["%s"] * len(boleto_ids))
        detalhes_sql = f"""
            SELECT
              b.id AS boleto_id,
              t.id AS titulo_id,
              COALESCE(NULLIF(d.nome,''), NULLIF(d.nome_fantasia,''), d.razao_social) AS devedor_nome,
              e.nome_fantasia AS credor_nome,
              e.cnpj         AS credor_cnpj,
              t.num_titulo,
              COALESCE(t.valorRecebido,0) AS valor_recebido,
              DATE_FORMAT(COALESCE(t.dataVencimentoReal, t.dataVencimento, t.dataVencimentoPrimeira), '%%d/%%m/%%Y') AS vencimento,
              DATE_FORMAT(t.data_baixa,'%%d/%%m/%%Y') AS data_baixa,
              CASE
                WHEN ((t.statusBaixa=2 OR t.statusBaixaGeral=2) AND COALESCE(t.valorRecebido,0)>0) THEN 'Quitado'
                WHEN  (t.statusBaixa=3 OR t.statusBaixaGeral=3) THEN 'Negociado'
                ELSE 'Pendente'
              END AS status_txt
            FROM core_boleto b
            JOIN titulo t
                 ON t.id_cobranca IS NOT NULL AND t.id_cobranca <> ''
                AND (CONVERT(t.id_cobranca USING utf8mb4) COLLATE {COLL}) =
                    (CASE
                       WHEN COALESCE(NULLIF(b.txid,''),'') <> '' THEN (CONCAT('PIX:', b.txid) COLLATE {COLL})
                       ELSE (CONVERT(b.codigo_solicitacao USING utf8mb4) COLLATE {COLL})
                     END)
            JOIN devedores d ON d.id = t.devedor_id
            JOIN core_empresa e ON e.id = d.empresa_id
            WHERE b.id IN ({ph})
            ORDER BY b.id, t.data_baixa DESC, t.id DESC
        """
        with connection.cursor() as cur:
            cur.execute(detalhes_sql, boleto_ids)
            for (boleto_id, titulo_id, dev_nome, credor_nome, credor_cnpj,
                 num_titulo, valor_rec, venc, data_baixa, status_txt) in cur.fetchall():
                detalhes_map[boleto_id].append({
                    "titulo_id": titulo_id,
                    "devedor": dev_nome,
                    "credor": credor_nome,
                    "credor_cnpj": credor_cnpj,
                    "num_titulo": num_titulo,
                    "valor": float(valor_rec or 0),
                    "vencimento": venc,
                    "data_baixa": data_baixa,
                    "status": status_txt,
                })

    for r in rows:
        r["detalhes"] = detalhes_map.get(r["id"], [])

    context = {
        "page_obj": {
            "boletos": rows,
            "has_previous": page_number > 1,
            "has_next": page_number < total_pages,
            "previous_page_number": page_number - 1 if page_number > 1 else None,
            "next_page_number": page_number + 1 if page_number < total_pages else None,
            "number": page_number,
            "paginator": {"num_pages": total_pages},
        },
        "valor_quitado": total_recebidos or 0,
        "valor_aberto": total_a_receber or 0,
        "valor_atrasado": total_atrasados or 0,
        "search_query": search_query,
        "data_inicio": data_inicio,
        "data_fim": data_fim,
        "pix_public_base": getattr(settings, "PIX_PUBLIC_BASE_URL", ""),
        "wa_open_url": request.session.pop("wa_open", None),
    }
    return render(request, "boletos_emitidos.html", context)

@login_required
@group_required([2])
def some_secure_view(request):
    # Your secure view logic here
    return render(request, 'some_secure_template.html')

@login_required
@csrf_exempt  # 🔴 REMOVA APÓS TESTES
def alterar_operador(request, titulo_id):
    if request.method != "POST":
        return JsonResponse({"status": "error", "message": "Método não permitido"}, status=405)

    operador_username = request.POST.get("operador")
    if not operador_username:
        return JsonResponse({"status": "error", "message": "Operador não selecionado!"}, status=400)

    titulo = get_object_or_404(Titulo, id=titulo_id)
    
    # Verifica se o operador existe
    if not User.objects.filter(username=operador_username).exists():
        return JsonResponse({"status": "error", "message": "Operador inválido!"}, status=400)

    # Atualiza o operador do título
    titulo.operador = operador_username
    titulo.save()

    return JsonResponse({
        "status": "success",
        "message": f"Operador alterado para {titulo.operador}!",
        "novo_operador": titulo.operador
    })
    
@login_required
@group_required([2])
def excluir_titulo_devedor(request, titulo_id):
    """
    Exclui um título associado a um devedor.
    """
    titulo = get_object_or_404(Titulo, id=titulo_id)

    if titulo.statusBaixa == 2:  # Impedir exclusão se já estiver quitado
        messages.error(request, "Não é possível excluir um título já quitado.")
        return redirect('listar_titulos_por_devedor', devedor_id=titulo.devedor.id)

    titulo.delete()
    messages.success(request, "Título excluído com sucesso!")
    return redirect('listar_titulos_por_devedor', devedor_id=titulo.devedor.id)
from decimal import Decimal, ROUND_HALF_UP
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.core.paginator import Paginator
from django.db import connection
from django.http import HttpResponse
from django.shortcuts import render
from core.decorators import group_required
from core.models import Empresa
from openpyxl import Workbook


@login_required
@group_required([2])
def honorarios(request):
    """
    Relatório por TÍTULO:
      Devedor | CPF/CNPJ | Credor | Consultor | Vencto. | Pagto. | Forma | Parc. |
      Valor principal | Valor pago | Honorários Credor | Valor liq.

    Regra de remuneração (maior atraso HISTÓRICO do DEVEDOR):
      30–90  = 9%
      91–180 = 15%
      181–720 = 21%
      721–1825 = 30%
      >=1826 = 40%
    """
    # -------- Filtros --------
    data_inicio  = (request.GET.get('data_inicio') or '').strip()
    data_fim     = (request.GET.get('data_fim') or '').strip()
    operador_sel = (request.GET.get('operador') or '').strip()
    empresa_filt = (request.GET.get('empresa') or '').strip()
    devedor_filt = (request.GET.get('devedor') or '').strip()
    cpf_cnpj     = (request.GET.get('cpf_cnpj') or '').strip()
    export_excel = (request.GET.get('exportar_excel') or '').strip()

    # Trava operador para não-admin
    trava_operador = False
    if not (request.user.is_staff or request.user.is_superuser):
        operador_sel = (request.user.get_full_name() or request.user.username).strip()
        trava_operador = True

    # -------- WHERE --------
    where_parts = [
        "(t.statusBaixa = 2 OR t.statusBaixaGeral = 2)",
        "t.valorRecebido > 0",
        "t.data_baixa IS NOT NULL",
        "e.status_empresa = 1",
    ]
    params = []

    if data_inicio:
        where_parts.append("t.data_baixa >= %s")
        params.append(f"{data_inicio} 00:00:00")
    if data_fim:
        where_parts.append("t.data_baixa < DATE_ADD(%s, INTERVAL 1 DAY)")
        params.append(data_fim)
    if operador_sel:
        where_parts.append("LOWER(e.operador) = LOWER(%s)")
        params.append(operador_sel)
    if empresa_filt:
        where_parts.append("(e.razao_social LIKE %s OR e.nome_fantasia LIKE %s)")
        params += [f"%{empresa_filt}%", f"%{empresa_filt}%"]
    if devedor_filt:
        where_parts.append("(d.nome LIKE %s OR d.nome_fantasia LIKE %s OR d.razao_social LIKE %s)")
        params += [f"%{devedor_filt}%"] * 3
    if cpf_cnpj:
        where_parts.append("(d.cpf LIKE %s OR d.cnpj LIKE %s)")
        params += [f"%{cpf_cnpj}%", f"%{cpf_cnpj}%"]

    where_sql = " AND ".join(where_parts)

    # -------- Query (usa MAIOR atraso histórico do devedor) --------
    QUERY = f"""
    WITH max_atraso AS (
      SELECT
        d.id AS devedor_id,
        MAX(GREATEST(
          0,
          DATEDIFF(
            t2.data_baixa,
            COALESCE(t2.dataVencimentoReal, t2.dataVencimento, t2.dataVencimentoPrimeira)
          )
        )) AS max_dias
      FROM titulo t2
      JOIN devedores d    ON d.id = t2.devedor_id
      JOIN core_empresa e ON e.id = d.empresa_id
      WHERE t2.data_baixa IS NOT NULL
        AND e.status_empresa = 1
      GROUP BY d.id
    )
    SELECT
      COALESCE(NULLIF(d.nome,''), NULLIF(d.nome_fantasia,''), d.razao_social) AS devedor_exib,
      COALESCE(NULLIF(d.cpf,''), d.cnpj)                                      AS doc_exib,
      CONCAT(e.id, ' - ', e.nome_fantasia)                                    AS credor_exib,
      e.operador                                                               AS consultor_exib,

      COALESCE(t.dataVencimentoReal, t.dataVencimento, t.dataVencimentoPrimeira) AS dt_venc,
      t.data_baixa                                                                 AS dt_pagto,
      t.forma_pag_Id                                                               AS forma_id,
      t.nPrc                                                                        AS n_prc,
      t.qtde_parcelas                                                               AS qtde_prc,

      t.valor                                                                       AS valor_principal,
      t.valorRecebido                                                               AS valor_pago,

      ma.max_dias                                                                    AS dias_atraso_hist,

      ROUND(
        t.valorRecebido * (
          CASE
            WHEN ma.max_dias BETWEEN  30 AND   90 THEN 0.09
            WHEN ma.max_dias BETWEEN  91 AND  180 THEN 0.15
            WHEN ma.max_dias BETWEEN 181 AND  720 THEN 0.21
            WHEN ma.max_dias BETWEEN 721 AND 1825 THEN 0.30
            WHEN ma.max_dias >= 1826             THEN 0.40
            ELSE 0
          END
        ), 2
      ) AS comissao_valor

    FROM titulo t
    JOIN devedores d    ON d.id = t.devedor_id
    JOIN core_empresa e ON e.id = d.empresa_id
    JOIN max_atraso ma  ON ma.devedor_id = d.id
    WHERE {where_sql}
    ORDER BY t.data_baixa DESC, e.nome_fantasia ASC, d.nome ASC
    """

    with connection.cursor() as cur:
        cur.execute(QUERY, params)
        rows = cur.fetchall()

    # -------- Mapeamento & Totais --------
    forma_map = {
        0: "Pix", 1: "Dinheiro", 2: "Cartão Débito", 3: "Cartão Crédito",
        4: "Cheque", 5: "Depósito", 6: "Na Loja", 7: "Boleto", 8: "Duplicata",
    }

    itens = []
    tot_valor_principal = Decimal('0.00')
    tot_valor_pago      = Decimal('0.00')
    tot_comissao        = Decimal('0.00')
    tot_liquido         = Decimal('0.00')

    for (devedor_exib, doc_exib, credor_exib, consultor_exib,
         dt_venc, dt_pagto, forma_id, n_prc, qtde_prc,
         v_princ, v_pago, dias_atraso_hist, comissao_valor) in rows:

        vp  = Decimal(str(v_princ or 0)).quantize(Decimal('0.01'))
        vg  = Decimal(str(v_pago or 0)).quantize(Decimal('0.01'))
        cm  = Decimal(str(comissao_valor or 0)).quantize(Decimal('0.01'))
        liq = (vg - cm).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)

        itens.append({
            "devedor": devedor_exib or "-",
            "doc": doc_exib or "-",
            "credor": credor_exib or "-",
            "consultor": consultor_exib or "-",
            "vencto": dt_venc,
            "pagto": dt_pagto,
            "forma": forma_map.get(forma_id, "Não informada"),
            "parc": f"{n_prc}/{qtde_prc}" if n_prc and qtde_prc else ("1/1" if vg > 0 else "-"),
            "valor_principal": vp,
            "valor_pago": vg,
            "honorarios_credor": cm,
            "valor_liq": liq,
        })

        tot_valor_principal += vp
        tot_valor_pago      += vg
        tot_comissao        += cm
        tot_liquido         += liq

    # Honorário do operador (25% da comissão total)
    honorario_total = (tot_comissao * Decimal('0.25')).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)

    # -------- Exportar Excel --------
    if export_excel:
        wb = Workbook(); ws = wb.active; ws.title = "Honorários"
        ws.append([
            "Devedor","CPF/CNPJ","Credor","Consultor","Vencto.","Pagto.","Forma","Parc.",
            "Valor principal","Valor pago","Honorários Credor","Valor liq."
        ])
        for h in itens:
            ws.append([
                h["devedor"], h["doc"], h["credor"], h["consultor"],
                (h["vencto"].strftime("%d/%m/%Y") if h["vencto"] else ""),
                (h["pagto"].strftime("%d/%m/%Y") if h["pagto"] else ""),
                h["forma"], h["parc"],
                float(h["valor_principal"]), float(h["valor_pago"]),
                float(h["honorarios_credor"]), float(h["valor_liq"]),
            ])
        ws.append(["","","","","","","","",
                   float(tot_valor_principal), float(tot_valor_pago),
                   float(tot_comissao), float(tot_liquido)])
        resp = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        resp['Content-Disposition'] = 'attachment; filename="honorarios.xlsx"'
        wb.save(resp)
        return resp

    # -------- Combos e paginação --------
    if request.user.is_staff or request.user.is_superuser:
        operadores = User.objects.filter(is_active=True).values_list('username', flat=True).distinct()
    else:
        operadores = [operador_sel]
    empresas = Empresa.objects.filter(status_empresa=1).values_list('nome_fantasia', flat=True).distinct()

    page_obj = Paginator(itens, 20).get_page(request.GET.get('page'))

    return render(request, 'honorarios.html', {
        "page_obj": page_obj,
        "tot_valor_principal": tot_valor_principal,
        "tot_valor_pago": tot_valor_pago,
        "tot_comissao": tot_comissao,
        "tot_liquido": tot_liquido,
        "honorario_total": honorario_total,

        "data_inicio": data_inicio,
        "data_fim": data_fim,
        "operador": operador_sel,
        "empresa": empresa_filt,
        "devedor": devedor_filt,
        "cpf_cnpj": cpf_cnpj,
        "operadores": operadores,
        "empresas": empresas,
        "trava_operador": trava_operador,
    })
